from unittest.mock import patch, sentinel
from pathlib import Path

from django.test import RequestFactory, SimpleTestCase
from django.urls import reverse
from openpyxl import load_workbook

from .views import F4_TEMPLATE_PATH, _export_filters, export_resumen_horas_excel


class ExportFilterTests(SimpleTestCase):
    def test_accepts_numeric_filters(self):
        request = RequestFactory().get('/', {'periodo': '12', 'carrera': '4'})
        self.assertEqual(_export_filters(request), ('12', '4'))

    def test_ignores_invalid_filters(self):
        request = RequestFactory().get('/', {'periodo': 'x', 'carrera': '-1'})
        self.assertEqual(_export_filters(request), (None, None))

    def test_report_center_has_a_dedicated_route(self):
        self.assertEqual(reverse('reportes:centro_reportes'), '/reportes/')

    @patch('reportes.views.export_planificacion_general_excel')
    def test_legacy_hours_export_uses_canonical_general_report(self, general_export):
        general_export.return_value = sentinel.response
        request = RequestFactory().get('/', {'periodo': '12'})

        result = export_resumen_horas_excel.__wrapped__(request)

        self.assertIs(result, sentinel.response)
        general_export.assert_called_once_with(request)


class InstitutionalF4TemplateTests(SimpleTestCase):
    def test_official_template_is_packaged_with_required_sheet_and_logos(self):
        self.assertTrue(Path(F4_TEMPLATE_PATH).is_file())
        workbook = load_workbook(F4_TEMPLATE_PATH)
        self.assertIn('MATRIZ F4 V1', workbook.sheetnames)
        worksheet = workbook['MATRIZ F4 V1']
        self.assertGreaterEqual(len(worksheet._images), 2)
        self.assertTrue(any(
            cell.data_type == 'f'
            for row in worksheet.iter_rows()
            for cell in row
        ))
