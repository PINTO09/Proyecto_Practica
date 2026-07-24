from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from copy import copy
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from docentes.models import DocenteFcacc, DocenteTituloAcademico, DocenteCampoAfinidad
from planificacion.models import (
    PlanificacionActividadDocente, PlanificacionAsignacionDocente,
    PlanificacionDemandaAcademica, PlanificacionMatrizF4,
)
from planificacion.services import (
    activity_workload_key, assignment_hour_category,
    build_docente_workload_map, knowledge_field_maps,
    normalize_parallel, normalize_workload_text, parallel_labels,
    registered_activity_keys,
)
from curriculo.models import CurriculoAsignatura
from catalogos.models import (
    CatalogoCarrera, CatalogoPeriodoAcademico, LimiteHorario,
)
from accounts.decorators import module_permission_required, allowed_career_ids


THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
HEADER_FILL = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
F4_TEMPLATE_PATH = (
    Path(__file__).resolve().parent
    / 'templates_excel'
    / 'planificacion_mkt_2026_2_v1.xlsx'
)


def _style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def _style_data(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center')


def _finish_sheet(ws, widths=None):
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    widths = widths or {}
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        ws.column_dimensions[letter].width = widths.get(column, 20)


def _export_filters(request):
    periodo = (request.GET.get('periodo') or '').strip()
    carrera = (request.GET.get('carrera') or '').strip()
    result = (
        periodo if periodo.isdigit() else None,
        carrera if carrera.isdigit() else None,
    )
    user = getattr(request, 'user', None)
    permitted = allowed_career_ids(user) if user is not None else None
    if permitted is not None and result[1] and int(result[1]) not in permitted:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied
    return result


def _excel_response(workbook, prefix):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'{prefix}_{timezone.localtime():%Y%m%d_%H%M%S}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    workbook.save(response)
    return response


def _filtered_assignments(periodo_id=None, carrera_id=None, user=None):
    qs = PlanificacionAsignacionDocente.objects.select_related(
        'id_docente', 'id_asignatura', 'id_carrera', 'id_periodo', 'id_campo'
    )
    if periodo_id:
        qs = qs.filter(id_periodo_id=periodo_id)
    if carrera_id:
        qs = qs.filter(id_carrera_id=carrera_id)
    permitted = allowed_career_ids(user) if user is not None else None
    if permitted is not None:
        qs = qs.filter(id_carrera_id__in=permitted)
    return qs


def _teacher_ids_for_scope(periodo_id=None, carrera_id=None, user=None):
    """Docentes vinculados a la carrera; None significa sin filtro de carrera."""
    permitted = allowed_career_ids(user) if user is not None else None
    if not carrera_id and permitted is None:
        return None
    assignment_ids = PlanificacionAsignacionDocente.objects.all()
    f4_ids = PlanificacionMatrizF4.objects.all()
    if carrera_id:
        assignment_ids = assignment_ids.filter(id_carrera_id=carrera_id)
        f4_ids = f4_ids.filter(id_carrera_id=carrera_id)
    if permitted is not None:
        assignment_ids = assignment_ids.filter(id_carrera_id__in=permitted)
        f4_ids = f4_ids.filter(id_carrera_id__in=permitted)
    if periodo_id:
        assignment_ids = assignment_ids.filter(id_periodo_id=periodo_id)
        f4_ids = f4_ids.filter(id_periodo_id=periodo_id)
    return set(assignment_ids.values_list('id_docente_id', flat=True)) | set(
        f4_ids.values_list('id_docente_id', flat=True)
    )


@login_required
@module_permission_required('reportes', 'view')
def centro_reportes(request):
    """Interfaz única para consultar y descargar reportes de planificación."""
    periodo_id, carrera_id = _export_filters(request)
    periodos = CatalogoPeriodoAcademico.objects.order_by(
        '-periodo_activo', '-fecha_inicio_periodo', '-id_periodo'
    )
    if not periodo_id:
        active_period = periodos.filter(periodo_activo=True).first()
        if active_period:
            periodo_id = str(active_period.id_periodo)

    assignments = _filtered_assignments(periodo_id, carrera_id, request.user)
    teacher_ids = _teacher_ids_for_scope(periodo_id, carrera_id, request.user)
    activities = PlanificacionActividadDocente.objects.all()
    f4_rows = PlanificacionMatrizF4.objects.all()
    if periodo_id:
        activities = activities.filter(id_periodo_id=periodo_id)
        f4_rows = f4_rows.filter(id_periodo_id=periodo_id)
    if carrera_id:
        f4_rows = f4_rows.filter(id_carrera_id=carrera_id)
        activities = activities.filter(id_docente_id__in=teacher_ids)

    workload = build_docente_workload_map(periodo_id=periodo_id)
    if teacher_ids is not None:
        workload = {key: value for key, value in workload.items() if key in teacher_ids}
    teachers_with_load = sum(1 for item in workload.values() if item.get('total_horas', 0) > 0)
    query = f'periodo={periodo_id or ""}&carrera={carrera_id or ""}'
    context = {
        'active_section': 'centro_reportes',
        'periodos': periodos,
        'carreras': CatalogoCarrera.objects.filter(
            carrera_activa=True,
            **({'id_carrera__in': allowed_career_ids(request.user)} if allowed_career_ids(request.user) is not None else {})
        ).order_by('nombre_carrera'),
        'periodo_id': int(periodo_id) if periodo_id else None,
        'carrera_id': int(carrera_id) if carrera_id else None,
        'export_query': query,
        'assignment_count': assignments.count(),
        'activity_count': activities.count(),
        'f4_count': f4_rows.count(),
        'teacher_count': teachers_with_load,
        'total_hours': sum(item.get('total_horas', 0) for item in workload.values()),
    }
    return render(request, 'reportes/centro_reportes.html', context)


# ——— API: Reporte Carga Docente ————————————————————————————

@login_required
@module_permission_required('reportes', 'view')
def reporte_carga_docente(request):
    periodo_id = request.GET.get('periodo')
    carrera_id = request.GET.get('carrera')
    qs = PlanificacionAsignacionDocente.objects.select_related(
        'id_docente', 'id_asignatura', 'id_carrera', 'id_periodo', 'id_campo'
    )
    if periodo_id:
        qs = qs.filter(id_periodo_id=periodo_id)
    if carrera_id:
        qs = qs.filter(id_carrera_id=carrera_id)

    data = []
    for a in qs[:500]:
        data.append({
            'docente': str(a.id_docente),
            'cedula': a.id_docente.cedula_docente if a.id_docente_id else '',
            'asignatura': str(a.id_asignatura) if a.id_asignatura_id else '',
            'carrera': str(a.id_carrera.nombre_carrera) if a.id_carrera_id else '',
            'periodo': str(a.id_periodo.nombre_periodo) if a.id_periodo_id else '',
            'horas': a.horas_clase,
            'semanas': a.semanas_planificadas,
            'horas_periodo': a.horas_clase_periodo,
            'campo': str(a.id_campo.nombre_campo_conocimiento) if a.id_campo_id else '',
        })
    return JsonResponse({'data': data})


# ——— API: Reporte Resumen Horas (Matriz F4) ————————————————

@login_required
@module_permission_required('reportes', 'view')
def reporte_resumen_horas(request):
    periodo_id = request.GET.get('periodo')
    carrera_id = request.GET.get('carrera')
    qs = PlanificacionMatrizF4.objects.select_related('id_docente', 'id_carrera', 'id_periodo')
    if periodo_id:
        qs = qs.filter(id_periodo_id=periodo_id)
    if carrera_id:
        qs = qs.filter(id_carrera_id=carrera_id)

    data = []
    for m in qs[:500]:
        data.append({
            'docente': str(m.id_docente),
            'cedula': m.id_docente.cedula_docente if m.id_docente_id else '',
            'carrera': str(m.id_carrera.nombre_carrera) if m.id_carrera_id else '',
            'periodo': str(m.id_periodo.nombre_periodo) if m.id_periodo_id else '',
            'horas_actividad': float(m.horas_actividad or 0),
            'total_horas': float(m.horas_actividad or 0),
        })
    return JsonResponse({'data': data})


# ——— API: Reporte Malla Curricular —————————————————————————

@login_required
@module_permission_required('reportes', 'view')
def reporte_malla_curricular(request):
    carrera_id = request.GET.get('carrera')
    qs = CurriculoAsignatura.objects.select_related('id_carrera').all()
    if carrera_id:
        qs = qs.filter(id_carrera_id=carrera_id)

    data = []
    for a in qs.order_by('id_carrera_id', 'nivel_semestre', 'nombre_asignatura')[:500]:
        data.append({
            'carrera': str(a.id_carrera.nombre_carrera) if a.id_carrera_id else '',
            'asignatura': a.nombre_asignatura,
            'codigo': a.codigo_asignatura or '',
            'nivel': a.nivel_semestre or 0,
            'horas': a.horas_semanales_asignatura or 0,
        })
    return JsonResponse({'data': data})


# ——— API: Reporte Docentes por Formación —————————————————————

@login_required
@module_permission_required('reportes', 'view')
def reporte_docentes_formacion(request):
    qs = DocenteTituloAcademico.objects.select_related('id_docente', 'id_posgrado').all()[:500]
    data = []
    for t in qs:
        data.append({
            'docente': str(t.id_docente),
            'cedula': t.id_docente.cedula_docente if t.id_docente_id else '',
            'titulo': t.nombre_titulo or '',
            'posgrado': str(t.id_posgrado.nombre_titulo_posgrado) if t.id_posgrado_id else '',
            'fecha_obtencion': str(t.fecha_obtencion_titulo) if t.fecha_obtencion_titulo else '',
            'registro_senescyt': t.numero_registro_senescyt or '',
        })
    return JsonResponse({'data': data})


# ——— API: Reporte Docentes por Campo Conocimiento ——————————

@login_required
@module_permission_required('reportes', 'view')
def reporte_docentes_campos(request):
    qs = DocenteCampoAfinidad.objects.select_related('id_docente', 'id_campo').all()[:500]
    data = []
    for dc in qs:
        data.append({
            'docente': str(dc.id_docente),
            'cedula': dc.id_docente.cedula_docente if dc.id_docente_id else '',
            'campo': str(dc.id_campo.nombre_campo_conocimiento) if dc.id_campo_id else '',
        })
    return JsonResponse({'data': data})


# ═══════════════════════════════════════════════════════════════
# EXPORTACIONES EXCEL
# ═══════════════════════════════════════════════════════════════

@login_required
@module_permission_required('reportes', 'view')
def export_carga_docente_excel(request):
    periodo_id, carrera_id = _export_filters(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Carga Docente'

    headers = ['Docente', 'Cédula', 'Asignatura', 'Carrera', 'Período', 'Horas semanales', 'Semanas', 'Horas período', 'Campo']
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    qs = _filtered_assignments(periodo_id, carrera_id, request.user).order_by(
        'id_docente__nombres_completos', 'id_asignatura__nombre_asignatura'
    )
    for i, a in enumerate(qs, start=2):
        ws.append([
            str(a.id_docente),
            a.id_docente.cedula_docente if a.id_docente_id else '',
            str(a.id_asignatura) if a.id_asignatura_id else '',
            str(a.id_carrera.nombre_carrera) if a.id_carrera_id else '',
            str(a.id_periodo.nombre_periodo) if a.id_periodo_id else '',
            a.horas_clase,
            a.semanas_planificadas,
            a.horas_clase_periodo,
            str(a.id_campo.nombre_campo_conocimiento) if a.id_campo_id else '',
        ])
        _style_data(ws, i, len(headers))

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    _finish_sheet(ws, {1: 36, 2: 14, 3: 36, 4: 30, 5: 22, 9: 30})
    return _excel_response(wb, 'asignaciones_docentes')


@login_required
@module_permission_required('reportes', 'view')
def export_malla_excel(request):
    _, carrera_id = _export_filters(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Malla Curricular'

    headers = ['Carrera', 'Asignatura', 'Código', 'Nivel', 'Horas Semanales']
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    qs = CurriculoAsignatura.objects.select_related('id_carrera').order_by(
        'id_carrera_id', 'nivel_semestre', 'nombre_asignatura'
    )
    if carrera_id:
        qs = qs.filter(id_carrera_id=carrera_id)
    permitted = allowed_career_ids(request.user)
    if permitted is not None:
        qs = qs.filter(id_carrera_id__in=permitted)
    for i, a in enumerate(qs, start=2):
        ws.append([
            str(a.id_carrera.nombre_carrera) if a.id_carrera_id else '',
            a.nombre_asignatura,
            a.codigo_asignatura or '',
            a.nivel_semestre or 0,
            a.horas_semanales_asignatura or 0,
        ])
        _style_data(ws, i, len(headers))

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

    _finish_sheet(ws, {1: 32, 2: 38, 3: 18, 4: 12, 5: 18})
    return _excel_response(wb, 'malla_curricular')


@login_required
@module_permission_required('reportes', 'view')
def export_resumen_horas_excel(request):
    # Compatibilidad con el botón antiguo: el resumen correcto es ahora la
    # planificación general, que incluye clases y todas las actividades.
    return export_planificacion_general_excel(request)


@login_required
@module_permission_required('reportes', 'view')
def export_planificacion_general_excel(request):
    """Exporta una fila consolidada por docente con su carga total."""
    periodo_id, carrera_id = _export_filters(request)
    workload_map = build_docente_workload_map(periodo_id=periodo_id)
    limites = {
        limite.id_modalidad_id: limite
        for limite in LimiteHorario.objects.filter(activo=True)
    }

    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen general'
    headers = [
        'Docente', 'Cédula', 'Modalidad', 'Dedicación', 'Horas afinidad',
        'Horas no afinidad', 'Horas unidad básica', 'Total horas clase',
        'Horas complementarias', 'Investigación', 'Gestión', 'Vinculación',
        'Otras actividades', 'Total actividades', 'Total horas', 'Límite',
        'Disponible', 'Cumplimiento %', 'Estado',
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    docentes = DocenteFcacc.objects.filter(docente_activo=True).select_related(
        'id_modalidad', 'id_dedicacion'
    ).order_by('nombres_completos')
    teacher_ids = _teacher_ids_for_scope(periodo_id, carrera_id, request.user)
    if teacher_ids is not None:
        docentes = docentes.filter(id_docente__in=teacher_ids)
    for row_number, docente in enumerate(docentes, start=2):
        workload = workload_map.get(docente.id_docente, {})
        afinidad = workload.get('horas_afinidad', 0) or 0
        no_afinidad = workload.get('horas_no_afinidad', 0) or 0
        unidad_basica = workload.get('horas_unidad_basica', 0) or 0
        clase = workload.get('total_horas_clase', 0) or 0
        complementarias = workload.get('horas_complementarias', 0) or 0
        investigacion = workload.get('horas_investigacion', 0) or 0
        gestion = workload.get('horas_gestion', 0) or 0
        vinculacion = workload.get('horas_vinculacion', 0) or 0
        actividades = workload.get('horas_actividad', 0) or 0
        total_actividades = workload.get('horas_actividades_total', 0) or 0
        total = workload.get('total_horas', clase + total_actividades) or 0
        limite = limites.get(docente.id_modalidad_id)
        maximo = (
            (limite.horas_maximas or 0) +
            (limite.horas_complementarias_maximas or 0)
        ) if limite else 0
        disponible = max(0, maximo - total)
        porcentaje = round(total * 100 / maximo, 1) if maximo else 0
        if porcentaje > 100:
            estado = 'Sobrecargado'
        elif porcentaje >= 80:
            estado = 'Alerta'
        elif total:
            estado = 'Balanceado'
        else:
            estado = 'Sin carga'
        ws.append([
            docente.nombres_completos,
            docente.cedula_docente,
            str(docente.id_modalidad) if docente.id_modalidad_id else '',
            str(docente.id_dedicacion) if docente.id_dedicacion_id else '',
            afinidad, no_afinidad, unidad_basica, clase, complementarias,
            investigacion, gestion, vinculacion, actividades, total_actividades,
            total, maximo, disponible, porcentaje, estado,
        ])
        _style_data(ws, row_number, len(headers))
        ws.cell(row_number, 5).fill = PatternFill('solid', fgColor='DCFCE7')
        ws.cell(row_number, 6).fill = PatternFill('solid', fgColor='FEE2E2')
        ws.cell(row_number, 7).fill = PatternFill('solid', fgColor='DBEAFE')
        ws.cell(row_number, 19).fill = PatternFill(
            'solid', fgColor='FECACA' if estado == 'Sobrecargado' else
            'FEF3C7' if estado == 'Alerta' else 'DCFCE7' if estado == 'Balanceado' else 'E5E7EB'
        )

    _finish_sheet(ws, {1: 36, 2: 14, 3: 22, 4: 18, 19: 16})
    return _excel_response(wb, 'planificacion_general')


@login_required
@module_permission_required('reportes', 'view')
def export_planificacion_detallada_excel(request):
    """Exporta asignaturas, actividades y Matriz F4 en hojas separadas."""
    periodo_id, carrera_id = _export_filters(request)
    teacher_ids = _teacher_ids_for_scope(periodo_id, carrera_id, request.user)
    wb = Workbook()

    assignments = PlanificacionAsignacionDocente.objects.select_related(
        'id_docente', 'id_asignatura', 'id_carrera', 'id_periodo', 'id_campo'
    ).order_by('id_docente__nombres_completos', 'id_asignatura__nombre_asignatura')
    if periodo_id:
        assignments = assignments.filter(id_periodo_id=periodo_id)
    if carrera_id:
        assignments = assignments.filter(id_carrera_id=carrera_id)
    permitted = allowed_career_ids(request.user)
    if permitted is not None:
        assignments = assignments.filter(id_carrera_id__in=permitted)
    ws = wb.active
    ws.title = 'Asignaturas'
    headers = [
        'Docente', 'Cédula', 'Carrera', 'Período', 'Código', 'Asignatura',
        'Nivel', 'Paralelo', 'Campo', 'Tipo de horas', 'Horas semanales',
        'Semanas', 'Horas período',
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    subject_fields, teacher_fields = knowledge_field_maps()
    category_labels = {
        'afinidad': 'Afinidad',
        'no_afinidad': 'No afinidad',
        'unidad_basica': 'Unidad básica',
    }
    for row_number, item in enumerate(assignments.iterator(), start=2):
        category = assignment_hour_category(
            item.id_docente_id, item.id_asignatura_id,
            item.id_campo.nombre_campo_conocimiento,
            subject_fields, teacher_fields,
        )
        ws.append([
            item.id_docente.nombres_completos, item.id_docente.cedula_docente,
            item.id_carrera.nombre_carrera, item.id_periodo.nombre_periodo,
            item.id_asignatura.codigo_asignatura, item.id_asignatura.nombre_asignatura,
            item.nivel_semestre_asignado, item.paralelo_asignado,
            item.id_campo.nombre_campo_conocimiento, category_labels[category],
            item.horas_clase, item.semanas_planificadas, item.horas_clase_periodo,
        ])
        _style_data(ws, row_number, len(headers))
        ws.cell(row_number, 10).fill = PatternFill(
            'solid', fgColor={
                'afinidad': 'DCFCE7',
                'no_afinidad': 'FEE2E2',
                'unidad_basica': 'DBEAFE',
            }[category]
        )
    _finish_sheet(ws, {1: 36, 3: 30, 6: 36, 9: 30, 10: 18})

    activities = PlanificacionActividadDocente.objects.select_related(
        'id_docente', 'id_periodo', 'id_actividad'
    ).order_by('id_docente__nombres_completos', 'id_actividad__nombre_actividad')
    if periodo_id:
        activities = activities.filter(id_periodo_id=periodo_id)
    if teacher_ids is not None:
        activities = activities.filter(id_docente_id__in=teacher_ids)
    ws = wb.create_sheet('Actividades')
    headers = ['Docente', 'Cédula', 'Período', 'Código', 'Actividad', 'Tipo', 'Horas', 'Observaciones']
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for row_number, item in enumerate(activities.iterator(), start=2):
        ws.append([
            item.id_docente.nombres_completos, item.id_docente.cedula_docente,
            item.id_periodo.nombre_periodo, item.id_actividad.codigo_actividad,
            item.id_actividad.nombre_actividad, item.id_actividad.get_tipo_actividad_display(),
            item.horas_asignadas, item.observaciones or '',
        ])
        _style_data(ws, row_number, len(headers))
    _finish_sheet(ws, {1: 36, 3: 24, 5: 38, 8: 45})

    f4_rows = PlanificacionMatrizF4.objects.select_related(
        'id_docente', 'id_carrera', 'id_periodo', 'id_grado_afinidad'
    ).order_by('id_docente__nombres_completos', 'tipo_actividad')
    if periodo_id:
        f4_rows = f4_rows.filter(id_periodo_id=periodo_id)
    if carrera_id:
        f4_rows = f4_rows.filter(id_carrera_id=carrera_id)
    if permitted is not None:
        f4_rows = f4_rows.filter(id_carrera_id__in=permitted)
    ws = wb.create_sheet('F4 adicional')
    headers = [
        'Docente', 'Cédula', 'Carrera', 'Período', 'Tipo', 'Detalle',
        'Nivel', 'Horas', 'Paralelos', 'Total', 'Observaciones',
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    activity_keys = registered_activity_keys(periodo_id)
    seen_f4 = set()
    row_number = 2
    for item in f4_rows.iterator():
        total = (item.horas_actividad or 0) * (item.numero_paralelos_actividad or 1)
        duplicate_key = activity_workload_key(
            item.id_docente_id, item.id_periodo_id,
            item.nombre_asignatura_actividad, total,
        )
        if duplicate_key in activity_keys:
            continue
        f4_key = (
            item.id_docente_id, item.id_periodo_id, item.tipo_actividad,
            normalize_workload_text(item.nombre_asignatura_actividad),
            item.horas_actividad, item.numero_paralelos_actividad,
        )
        if not carrera_id and f4_key in seen_f4:
            continue
        seen_f4.add(f4_key)
        ws.append([
            item.id_docente.nombres_completos, item.id_docente.cedula_docente,
            item.id_carrera.nombre_carrera, item.id_periodo.nombre_periodo,
            item.tipo_actividad, item.nombre_asignatura_actividad or '',
            item.nivel_semestre_actividad or '', item.horas_actividad,
            item.numero_paralelos_actividad, total, item.observaciones or '',
        ])
        _style_data(ws, row_number, len(headers))
        row_number += 1
    _finish_sheet(ws, {1: 36, 3: 30, 5: 24, 6: 38, 11: 45})

    return _excel_response(wb, 'planificacion_detallada')


@login_required
@module_permission_required('planificacion')
def export_matriz_f4_mkt_excel(request):
    """Exporta Matriz F4 en formato MKT (16 columnas) para el periodo y carrera actual."""
    periodo_id = request.GET.get('periodo')
    carrera_id = request.GET.get('carrera')
    wb = Workbook()
    ws = wb.active
    ws.title = 'MATRIZ F4 V1'

    headers = [
        'N°', 'Cédula', 'Apellidos y Nombres', 'Sexo', 'Unidad Orgánica',
        'Titular', 'Categoría', 'Título de tercer nivel', 'Título de cuarto nivel',
        'Carrera o Tipo de Actividad', 'Asignatura o Actividad',
        'Grado de afinidad', 'Horas', 'Paralelos', 'Total Horas', 'Carga Horaria',
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    qs = PlanificacionMatrizF4.objects.select_related(
        'id_docente', 'id_carrera', 'id_periodo', 'id_grado_afinidad',
    ).order_by('id_docente__nombres_completos', 'tipo_actividad')

    if periodo_id:
        qs = qs.filter(id_periodo__codigo_periodo=periodo_id)
    if carrera_id:
        qs = qs.filter(id_carrera__codigo_carrera=carrera_id)

    permitted = allowed_career_ids(request.user)
    if permitted is not None:
        qs = qs.filter(id_carrera__codigo_carrera__in=permitted)

    # Pre-fetch titulos por docente
    from django.db.models import Q
    titulos_qs = DocenteTituloAcademico.objects.filter(
        id_docente__in=qs.values_list('id_docente_id', flat=True).distinct()
    ).order_by('id_docente_id', '-nivel_titulo')
    titulos_por_docente = {}
    for t in titulos_qs:
        titulos_por_docente.setdefault(t.id_docente_id, []).append(t)

    row_number = 2
    docente_actual = None
    secuencia = 0
    from django.db.models import Sum, F, Q
    carga_filter = Q()
    if periodo_id:
        carga_filter &= Q(id_periodo__codigo_periodo=periodo_id)
    carga_por_docente = dict(
        PlanificacionMatrizF4.objects.filter(carga_filter)
        .values('id_docente_id')
        .annotate(total=Sum(F('horas_actividad') * F('numero_paralelos_actividad')))
        .values_list('id_docente_id', 'total')
    )

    for item in qs.iterator():
        docente = item.id_docente
        total_linea = (item.horas_actividad or 0) * (item.numero_paralelos_actividad or 1)

        if docente.id_docente != docente_actual:
            secuencia += 1
            docente_actual = docente.id_docente

            titulos = titulos_por_docente.get(docente.id_docente, [])
            tercer_nivel = ''
            cuarto_nivel = ''
            for t in titulos:
                nt = t.nivel_titulo or 0
                if nt >= 4:
                    cuarto_nivel = (cuarto_nivel + '; ' + t.nombre_titulo) if cuarto_nivel else t.nombre_titulo
                else:
                    tercer_nivel = (tercer_nivel + '; ' + t.nombre_titulo) if tercer_nivel else t.nombre_titulo
            if not tercer_nivel and titulos:
                tercer_nivel = titulos[0].nombre_titulo

            modalidad_text = str(docente.id_modalidad or '')
            dedicacion_text = str(docente.id_dedicacion or '')
            categoria = f'{modalidad_text} {dedicacion_text}'.strip()

            carga = carga_por_docente.get(docente.id_docente, '')

            ws.append([
                secuencia,
                docente.cedula_docente or '',
                docente.nombres_completos or '',
                '',
                docente.unidad_organica or '',
                '',
                categoria,
                tercer_nivel,
                cuarto_nivel,
                item.id_carrera.nombre_carrera if item.id_carrera else '',
                item.nombre_asignatura_actividad or '',
                str(item.id_grado_afinidad or ''),
                item.horas_actividad or 0,
                item.numero_paralelos_actividad or 1,
                total_linea,
                carga if carga else '',
            ])
        else:
            ws.append([
                '', '', '', '', '', '', '', '', '',
                item.id_carrera.nombre_carrera if item.id_carrera else '',
                item.nombre_asignatura_actividad or '',
                str(item.id_grado_afinidad or ''),
                item.horas_actividad or 0,
                item.numero_paralelos_actividad or 1,
                total_linea,
                '',
            ])

        _style_data(ws, row_number, len(headers))
        row_number += 1

    _finish_sheet(ws, {1: 5, 2: 14, 3: 36, 5: 30, 7: 30, 8: 30, 9: 30, 10: 30, 11: 40, 12: 30})
    return _excel_response(wb, 'matriz_f4_mkt')


@login_required
@module_permission_required('reportes', 'view')
def descargar_planificacion_original(request):
    """Genera el F4 final consolidado del módulo de planificación."""
    from collections import OrderedDict
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from docentes.models import DocenteTituloAcademico
    from planificacion.models import (
        CatalogoActividadComplementaria,
        PlanificacionActividadDocente,
        PlanificacionMatrizF4,
    )
    from planificacion.services import docente_tiene_afinidad
    from django.http import Http404
    import re
    import unicodedata

    if not F4_TEMPLATE_PATH.exists():
        raise Http404('La plantilla institucional de la Matriz F4 no está instalada.')

    periodo_id, carrera_id = _export_filters(request)
    wb_dst = load_workbook(F4_TEMPLATE_PATH)
    if 'MATRIZ F4 V1' not in wb_dst.sheetnames:
        raise Http404('La plantilla no contiene la hoja MATRIZ F4 V1.')
    ws_dst = wb_dst['MATRIZ F4 V1']
    for worksheet in list(wb_dst.worksheets):
        if worksheet.title != 'MATRIZ F4 V1':
            wb_dst.remove(worksheet)

    def normalize_name(value):
        value = unicodedata.normalize('NFKD', str(value or ''))
        value = ''.join(ch for ch in value if not unicodedata.combining(ch))
        return re.sub(r'[^A-Z0-9]+', ' ', value.upper()).strip()

    # 1) Asignaturas: una línea por docente/asignatura, agrupando paralelos.
    assignments = list(
        _filtered_assignments(periodo_id, carrera_id, request.user)
        .select_related(
            'id_docente__id_tipo_docente',
            'id_docente__id_modalidad',
            'id_docente__id_dedicacion',
            'id_asignatura', 'id_carrera', 'id_periodo',
        )
        .order_by(
            'id_docente__nombres_completos',
            'id_carrera__nombre_carrera',
            'id_asignatura__nivel_semestre',
            'id_asignatura__nombre_asignatura',
            'paralelo_asignado',
        )
    )
    grouped_assignments = OrderedDict()
    represented_names = set()
    for item in assignments:
        affinity = docente_tiene_afinidad(item.id_docente, item.id_asignatura)
        key = (
            item.id_docente_id, item.id_periodo_id, item.id_carrera_id,
            item.id_asignatura_id, item.horas_clase, affinity,
        )
        entry = grouped_assignments.setdefault(key, {
            'teacher': item.id_docente,
            'period_id': item.id_periodo_id,
            'rank': 1,
            'type': 'ASIGNATURA',
            'career_or_type': item.id_carrera.nombre_carrera,
            'name': (
                f'{item.id_asignatura.codigo_asignatura} - '
                f'{item.id_asignatura.nombre_asignatura} '
                f'(Nivel {item.nivel_semestre_asignado})'
            ),
            'affinity': 'AFÍN' if affinity else 'NO AFÍN',
            'hours': item.horas_clase or 0,
            'parallels': set(),
        })
        entry['parallels'].add(item.paralelo_asignado or str(item.pk))
        represented_names.add((
            item.id_docente_id, item.id_periodo_id,
            normalize_name(item.id_asignatura.nombre_asignatura),
        ))
        represented_names.add((
            item.id_docente_id, item.id_periodo_id,
            normalize_name(
                f'{item.id_asignatura.codigo_asignatura} '
                f'{item.id_asignatura.nombre_asignatura}'
            ),
        ))

    final_rows = []
    for entry in grouped_assignments.values():
        entry['parallels'] = len(entry['parallels'])
        final_rows.append(entry)

    # 2) Actividades complementarias registradas en el sistema.
    teacher_scope = _teacher_ids_for_scope(
        periodo_id, carrera_id, request.user
    )
    activities = (
        PlanificacionActividadDocente.objects
        .select_related(
            'id_docente__id_tipo_docente',
            'id_docente__id_modalidad',
            'id_docente__id_dedicacion',
            'id_periodo', 'id_actividad',
        )
        .order_by(
            'id_docente__nombres_completos',
            'id_actividad__tipo_actividad',
            'id_actividad__nombre_actividad',
        )
    )
    if periodo_id:
        activities = activities.filter(id_periodo_id=periodo_id)
    if teacher_scope is not None:
        activities = activities.filter(id_docente_id__in=teacher_scope)
    activity_labels = dict(CatalogoActividadComplementaria.TIPOS)
    for item in activities:
        activity_name = item.id_actividad.nombre_actividad
        final_rows.append({
            'teacher': item.id_docente,
            'period_id': item.id_periodo_id,
            'rank': 2,
            'type': item.id_actividad.tipo_actividad,
            'career_or_type': activity_labels.get(
                item.id_actividad.tipo_actividad,
                item.id_actividad.tipo_actividad,
            ),
            'name': (
                f'{item.id_actividad.codigo_actividad} - {activity_name}'
            ),
            'affinity': 'NO APLICA',
            'hours': item.horas_asignadas or 0,
            'parallels': 1,
        })
        represented_names.add((
            item.id_docente_id, item.id_periodo_id,
            normalize_name(activity_name),
        ))
        represented_names.add((
            item.id_docente_id, item.id_periodo_id,
            normalize_name(
                f'{item.id_actividad.codigo_actividad} {activity_name}'
            ),
        ))

    # 3) Registros históricos F4 que no están ya representados.
    historical = (
        PlanificacionMatrizF4.objects
        .select_related(
            'id_docente__id_tipo_docente',
            'id_docente__id_modalidad',
            'id_docente__id_dedicacion',
            'id_carrera', 'id_periodo', 'id_grado_afinidad',
        )
        .order_by(
            'id_docente__nombres_completos',
            'tipo_actividad', 'nombre_asignatura_actividad',
        )
    )
    if periodo_id:
        historical = historical.filter(id_periodo_id=periodo_id)
    if carrera_id:
        historical = historical.filter(id_carrera_id=carrera_id)
    permitted = allowed_career_ids(request.user)
    if permitted is not None:
        historical = historical.filter(id_carrera_id__in=permitted)

    historical_seen = set()
    for item in historical:
        name = item.nombre_asignatura_actividad or ''
        represented_key = (
            item.id_docente_id, item.id_periodo_id, normalize_name(name),
        )
        history_key = (
            represented_key, item.horas_actividad,
            item.numero_paralelos_actividad,
        )
        if represented_key in represented_names or history_key in historical_seen:
            continue
        historical_seen.add(history_key)
        final_rows.append({
            'teacher': item.id_docente,
            'period_id': item.id_periodo_id,
            'rank': 3,
            'type': item.tipo_actividad,
            'career_or_type': item.tipo_actividad or item.id_carrera.nombre_carrera,
            'name': name,
            'affinity': (
                item.id_grado_afinidad.nombre_grado_afinidad
                if item.id_grado_afinidad else 'NO APLICA'
            ),
            'hours': item.horas_actividad or 0,
            'parallels': item.numero_paralelos_actividad or 1,
        })

    final_rows.sort(key=lambda item: (
        item['teacher'].nombres_completos,
        item['teacher'].id_docente,
        item['rank'],
        item['career_or_type'],
        item['name'],
    ))
    rows_by_teacher = OrderedDict()
    for row in final_rows:
        rows_by_teacher.setdefault(row['teacher'].id_docente, []).append(row)

    # Títulos académicos para las columnas personales.
    teacher_ids = list(rows_by_teacher)
    titles_map = {}
    for title in (
        DocenteTituloAcademico.objects
        .filter(id_docente_id__in=teacher_ids)
        .order_by('id_docente_id', 'nivel_titulo', 'nombre_titulo')
    ):
        level = 'cuarto' if title.nivel_titulo >= 4 else 'tercer'
        titles_map.setdefault(
            title.id_docente_id, {'tercer': [], 'cuarto': []}
        )[level].append(title.nombre_titulo)

    # 4) Conservar la plantilla institucional y retirar únicamente el contenido
    # manual: resaltadores, apuntes laterales y filas diligenciadas.
    original_widths = {
        column: ws_dst.column_dimensions[get_column_letter(column)].width
        for column in range(1, 17)
    }
    original_hidden_columns = {
        column: bool(
            ws_dst.column_dimensions[get_column_letter(column)].hidden
        )
        for column in range(1, 17)
    }
    personal_styles = {
        column: copy(ws_dst.cell(18, column)._style)
        for column in range(1, 10)
    }
    detail_styles = {
        column: copy(ws_dst.cell(12, column)._style)
        for column in range(10, 16)
    }
    load_style = copy(ws_dst.cell(18, 16)._style)

    for merged in list(ws_dst.merged_cells.ranges):
        if merged.min_row >= 8:
            ws_dst.unmerge_cells(str(merged))
    if ws_dst.max_row > 7:
        ws_dst.delete_rows(8, ws_dst.max_row - 7)
    if ws_dst.max_column > 16:
        ws_dst.delete_cols(17, ws_dst.max_column - 16)
    ws_dst.freeze_panes = 'A8'
    if hasattr(ws_dst, 'conditional_formatting'):
        ws_dst.conditional_formatting._cf_rules.clear()
    if getattr(ws_dst, 'data_validations', None):
        ws_dst.data_validations.dataValidation = []
    for row in ws_dst.iter_rows():
        for cell in row:
            if cell.comment is not None:
                cell.comment = None

    # Actualizar solo los datos variables. Los títulos, logos, combinaciones y
    # estilos de las filas 1 a 7 permanecen como en la plantilla original.
    period = (
        CatalogoPeriodoAcademico.objects.filter(pk=periodo_id).first()
        if periodo_id else None
    )
    career = (
        CatalogoCarrera.objects.filter(pk=carrera_id).first()
        if carrera_id else None
    )
    period_text = period.nombre_periodo if period else 'TODOS LOS PERÍODOS'
    unit_text = (
        career.nombre_carrera if career
        else 'CIENCIAS ADMINISTRATIVAS, CONTABLES Y COMERCIO'
    )
    ws_dst['A6'] = f'PERÍODO ACADÉMICO:   {period_text}'
    ws_dst['G6'] = f'UNIDAD ACADÉMICA:    {unit_text}'
    ws_dst['I7'] = 'Título de cuarto nivel'

    # 5) Diligenciar la matriz con los datos consolidados del sistema usando
    # estilos originales, pero sin resaltadores manuales.
    neutral_fill = PatternFill(fill_type=None)
    current_row = 8
    for sequence, teacher_rows in enumerate(rows_by_teacher.values(), start=1):
        teacher = teacher_rows[0]['teacher']
        start_row = current_row
        title_values = titles_map.get(
            teacher.id_docente, {'tercer': [], 'cuarto': []}
        )
        teacher_type = teacher.id_tipo_docente
        modality = teacher.id_modalidad
        dedication = teacher.id_dedicacion
        is_titular = bool(
            teacher_type
            and teacher_type.codigo_tipo_docente.upper() == 'TITULAR'
        )
        category = ' · '.join(filter(None, [
            modality.nombre_modalidad if modality else '',
            dedication.codigo_dedicacion if dedication else '',
        ]))

        end_row = current_row + len(teacher_rows) - 1
        for i, item in enumerate(teacher_rows):
            row = current_row
            personal_values = {
                1: sequence if i == 0 else '',
                2: teacher.cedula_docente,
                3: teacher.nombres_completos,
                4: '',
                5: teacher.unidad_organica or '',
                6: 'SÍ' if is_titular else 'NO',
                7: category,
                8: '; '.join(title_values['tercer']),
                9: '; '.join(title_values['cuarto']),
            }
            detail_values = {
                10: item['career_or_type'],
                11: item['name'],
                12: item['affinity'],
                13: item['hours'],
                14: item['parallels'],
            }
            for column, value in {**personal_values, **detail_values}.items():
                ws_dst.cell(row, column).value = value
            ws_dst.cell(row, 15).value = f'=M{row}*N{row}'
            ws_dst.cell(row, 16).value = f'=SUM(O{start_row}:O{end_row})' if i == 0 else ''

            for column in range(1, 10):
                cell = ws_dst.cell(row, column)
                cell._style = copy(personal_styles[column])
                cell.fill = copy(neutral_fill)
            for column in range(10, 16):
                cell = ws_dst.cell(row, column)
                cell._style = copy(detail_styles[column])
                cell.fill = copy(neutral_fill)
            load_cell = ws_dst.cell(row, 16)
            load_cell._style = copy(load_style)
            load_cell.fill = copy(neutral_fill)
            ws_dst.row_dimensions[row].height = 30
            current_row += 1

        if end_row > start_row:
            for column in [1, 4, 8, 9]:
                ws_dst.merge_cells(
                    start_row=start_row, start_column=column,
                    end_row=end_row, end_column=column,
                )
            ws_dst.merge_cells(
                start_row=start_row, start_column=16,
                end_row=end_row, end_column=16,
            )

    data_end = current_row - 1
    for column, width in original_widths.items():
        dimension = ws_dst.column_dimensions[get_column_letter(column)]
        dimension.width = width
        dimension.hidden = original_hidden_columns[column]
    # El título se conserva en el archivo, pero no forma parte de la vista
    # principal solicitada para la entrega.
    ws_dst.column_dimensions['I'].hidden = True
    for row in range(1, data_end + 1):
        ws_dst.row_dimensions[row].hidden = False
    ws_dst.auto_filter.ref = (
        f'A7:P{data_end}' if data_end >= 8 else 'A7:P7'
    )
    ws_dst.print_title_rows = '1:7'
    ws_dst.print_area = f'A1:P{data_end}' if data_end >= 8 else 'A1:P7'
    ws_dst.sheet_properties.pageSetUpPr.fitToPage = True
    ws_dst.page_setup.fitToWidth = 1
    ws_dst.page_setup.fitToHeight = 0

    # Forzar recálculo de las fórmulas al abrir el archivo.
    try:
        wb_dst.calculation.fullCalcOnLoad = True
        wb_dst.calculation.forceFullCalc = True
        wb_dst.calculation.calcMode = 'auto'
    except AttributeError:
        pass

    from io import BytesIO
    output = BytesIO()
    wb_dst.save(output)
    output.seek(0)
    content = output.read()

    response = HttpResponse(content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'MATRIZ_F4_V1_{timezone.localtime():%Y%m%d_%H%M%S}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Length'] = len(content)
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@login_required
@module_permission_required('reportes', 'view')
def export_reporte_asignacion_carreras(request):
    periodo_id, carrera_id = _export_filters(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Asignación por carreras'
    headers = ['Carrera', 'Asignatura', 'Nivel', 'Paralelo', 'Horas', 'Docente asignado']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    permitted = allowed_career_ids(request.user)

    # Solo asignaturas con Demanda
    demandas_qs = PlanificacionDemandaAcademica.objects.select_related('id_asignatura', 'id_carrera').all()
    if periodo_id:
        demandas_qs = demandas_qs.filter(id_periodo_id=periodo_id)
    if permitted is not None:
        demandas_qs = demandas_qs.filter(id_carrera_id__in=permitted)
    if carrera_id:
        demandas_qs = demandas_qs.filter(id_carrera_id=carrera_id)
    demandas_qs = demandas_qs.order_by(
        'id_carrera__nombre_carrera', 'id_asignatura__nivel_semestre', 'id_asignatura__nombre_asignatura'
    )

    # Asignaciones del periodo
    asignaciones_qs = PlanificacionAsignacionDocente.objects.all()
    if periodo_id:
        asignaciones_qs = asignaciones_qs.filter(id_periodo_id=periodo_id)
    if permitted is not None:
        asignaciones_qs = asignaciones_qs.filter(id_carrera_id__in=permitted)
    if carrera_id:
        asignaciones_qs = asignaciones_qs.filter(id_carrera_id=carrera_id)
    asignadas = {}
    asignaciones_horas = {}
    for asig_id, carr_id, paralelo, horas in asignaciones_qs.values_list(
        'id_asignatura_id', 'id_carrera_id', 'paralelo_asignado', 'horas_clase'
    ):
        key = (asig_id, carr_id, normalize_parallel(paralelo))
        asignadas[key] = True
        if key not in asignaciones_horas:
            asignaciones_horas[key] = horas

    row = 2
    for d in demandas_qs:
        paralelos = parallel_labels(d.numero_paralelos)
        for p in paralelos:
            a_key = (d.id_asignatura_id, d.id_carrera_id, p)
            tiene = a_key in asignadas
            horas = asignaciones_horas.get(a_key, d.id_asignatura.horas_semanales_asignatura or 0)
            docente = '✅' if tiene else '🚫'
            ws.cell(row=row, column=1, value=str(d.id_carrera))
            ws.cell(row=row, column=2, value=str(d.id_asignatura))
            ws.cell(row=row, column=3, value=d.id_asignatura.nivel_semestre)
            ws.cell(row=row, column=4, value=p)
            ws.cell(row=row, column=5, value=horas)
            ws.cell(row=row, column=6, value=docente)
            _style_data(ws, row, len(headers))
            row += 1

    _finish_sheet(ws, {1: 30, 2: 45, 3: 8, 4: 10, 5: 8, 6: 18})
    return _excel_response(wb, 'asignacion_por_carreras')
