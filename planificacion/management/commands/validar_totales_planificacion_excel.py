from collections import defaultdict
from pathlib import Path
import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db.models import Sum, F, IntegerField, ExpressionWrapper
from openpyxl import load_workbook

from catalogos.models import CatalogoCarrera, CatalogoPeriodoAcademico
from docentes.models import DocenteFcacc
from planificacion.models import PlanificacionAsignacionDocente, PlanificacionMatrizF4


def _normalize_text(value):
    text = str(value or "").replace("\xa0", " ").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def _clean_text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _positive_int(value):
    try:
        number = int(float(value or 0))
    except (TypeError, ValueError):
        return 0
    return number if number > 0 else 0


def _canonical_name(value):
    text = _normalize_text(value)
    text = re.sub(r"^\d+\s+", "", text)
    text = re.sub(r"\s*\([^)]*\)\s*", " ", text)
    text = text.replace(" - ", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


CAREER_ALIASES = {
    "CONTABILIDAD Y ADITORIA 2024": "CONTABILIDAD Y AUDITORIA 2024 NS",
    "AUDITORIA Y CONTROL DE GESTION 2024": "AUDITORIA Y CONTROL DE GESTION 2024 NS",
}


F4_ACTIVITY_CAREER_CODES = {
    "FCACC-1-DO-CL",
    "FCACC-4-DO-TU-TI",
    "FCACC-5-DO-PE-IN",
    "FCACC-6-IVN",
    "FCACC-8-VI-SO",
    "FCACC-9-GE_ED",
}


class Command(BaseCommand):
    help = "Compara los totales docentes de los Excel contra las horas cargadas en la base."

    def add_arguments(self, parser):
        parser.add_argument(
            "--base-dir",
            default=None,
            help="Directorio base con los Excel extraidos. Por defecto usa _excel_input del proyecto.",
        )
        parser.add_argument(
            "--periodo-codigo",
            default="2026-2",
            help="Codigo del periodo academico a validar.",
        )
        parser.add_argument(
            "--solo-diferencias",
            action="store_true",
            help="Muestra solamente docentes con diferencias entre Excel y base.",
        )
        parser.add_argument(
            "--global-docente",
            action="store_true",
            help="Agrega una comparacion global por docente tomando el mayor total declarado en los Excel.",
        )

    def handle(self, *args, **options):
        base_dir = Path(options["base_dir"] or Path(__file__).resolve().parents[3] / "_excel_input")
        periodo = CatalogoPeriodoAcademico.objects.filter(codigo_periodo=options["periodo_codigo"]).first()
        if not periodo:
            raise CommandError(f"No existe el periodo {options['periodo_codigo']}.")

        sources = [
            base_dir / "OneDrive_1" / "FCACC-PLANIFICACION.xlsx",
            base_dir / "OneDrive_1" / "PLANIFICACION_ADMINISTRACION.xlsx",
            base_dir / "OneDrive_1" / "PLANIFICACION_COMERCIO.xlsx",
            base_dir / "OneDrive_1" / "PLANIFICACION_CONTABILIDAD.xlsx",
        ]
        for path in sources:
            if not path.exists():
                raise CommandError(f"No se encontro el archivo: {path}")

        careers_by_code = {obj.codigo_carrera: obj for obj in CatalogoCarrera.objects.all()}
        careers_by_name = {
            _canonical_name(obj.nombre_carrera): obj
            for obj in CatalogoCarrera.objects.all()
        }
        teachers_by_name = {
            _normalize_text(obj.nombres_completos): obj
            for obj in DocenteFcacc.objects.all()
        }

        expected = []
        expected_global = {}
        for path in sources:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = workbook["ASIGNACION"]
                rows = sheet.iter_rows(values_only=True)
                headers = list(next(rows))
                teacher_idx = headers.index("NOMBRE_DOCENT")
                career_idx = headers.index("CARRERA")
                career_code_idx = headers.index("ID_CARRERA")
                total_idx = headers.index("TOTAL_HORAS_DOCENTE") if "TOTAL_HORAS_DOCENTE" in headers else headers.index("TOTAL HORAS")
                is_main_book = path.name == "FCACC-PLANIFICACION.xlsx"
                nivel_idx = headers.index("NIVEL")
                subject_idx = headers.index("ASIGNATURA")
                row_total_idx = headers.index("TOTAL")
                field_idx = headers.index("CAMPO") if "CAMPO" in headers else None
                horas_idx = headers.index("HORAS")
                horas_semanas_idx = headers.index("HORAS_SEMANAS") if "HORAS_SEMANAS" in headers else None
                h_adicional_1_idx = headers.index("H.ADI_01") if "H.ADI_01" in headers else None
                h_adicional_2_idx = headers.index("H.ADI_02") if "H.ADI_02" in headers else None
                h_investigacion_idx = headers.index("H.INVESTIGACION") if "H.INVESTIGACION" in headers else None

                bucket = {}
                for row in rows:
                    if not row or len(row) <= max(teacher_idx, career_idx, career_code_idx, total_idx):
                        continue
                    teacher_name = _clean_text(row[teacher_idx])
                    total = _positive_int(row[total_idx])
                    if not teacher_name or total <= 0:
                        continue
                    global_item = expected_global.setdefault(
                        _normalize_text(teacher_name),
                        {
                            "teacher_name": teacher_name,
                            "expected_total": 0,
                            "detail_class": 0,
                            "detail_f4": 0,
                            "detail_f4_seen": set(),
                            "sources": set(),
                        },
                    )
                    global_item["expected_total"] = max(global_item["expected_total"], total)
                    global_item["sources"].add(path.name)

                    subject_name = _clean_text(row[subject_idx]) if len(row) > subject_idx else ""
                    career_code = _clean_text(row[career_code_idx]) if len(row) > career_code_idx else ""
                    try:
                        level = int(row[nivel_idx] or 0)
                    except (TypeError, ValueError):
                        level = 0

                    if is_main_book:
                        row_total = _positive_int(row[row_total_idx]) if len(row) > row_total_idx else 0
                        field_name = _normalize_text(row[field_idx]) if field_idx is not None and len(row) > field_idx else ""
                        if career_code in F4_ACTIVITY_CAREER_CODES or field_name == "ACTIVIDAD":
                            f4_key = ("Actividad", subject_name, row_total)
                            if row_total > 0 and f4_key not in global_item["detail_f4_seen"]:
                                global_item["detail_f4_seen"].add(f4_key)
                                global_item["detail_f4"] += row_total
                        elif 1 <= level <= 8:
                            global_item["detail_class"] += row_total
                    else:
                        horas_semanas = _positive_int(row[horas_semanas_idx]) if horas_semanas_idx is not None and len(row) > horas_semanas_idx else 0
                        horas_clase = _positive_int(row[horas_idx]) if len(row) > horas_idx else 0
                        if horas_semanas > 0 and 1 <= level <= 8:
                            global_item["detail_class"] += horas_clase

                        additions = (
                            ("Actividad adicional 1", h_adicional_1_idx),
                            ("Actividad adicional 2", h_adicional_2_idx),
                            ("Investigacion", h_investigacion_idx),
                        )
                        for tipo_actividad, idx in additions:
                            hours = _positive_int(row[idx]) if idx is not None and len(row) > idx else 0
                            f4_key = (tipo_actividad, subject_name, hours)
                            if hours > 0 and f4_key not in global_item["detail_f4_seen"]:
                                global_item["detail_f4_seen"].add(f4_key)
                                global_item["detail_f4"] += hours

                    career = careers_by_code.get(_clean_text(row[career_code_idx]))
                    if not career:
                        canonical_career = CAREER_ALIASES.get(_canonical_name(row[career_idx]), _canonical_name(row[career_idx]))
                        career = careers_by_name.get(canonical_career)
                    key = (
                        _normalize_text(teacher_name),
                        career.id_carrera if career else _canonical_name(row[career_idx]),
                    )
                    item = bucket.setdefault(
                        key,
                        {
                            "source": path.name,
                            "teacher_name": teacher_name,
                            "expected_total": 0,
                            "career_ids": set(),
                            "career_labels": set(),
                        },
                    )
                    item["expected_total"] = max(item["expected_total"], total)
                    if career:
                        item["career_ids"].add(career.id_carrera)
                        item["career_labels"].add(career.nombre_carrera)
                expected.extend(bucket.values())
            finally:
                workbook.close()

        rows = []
        for item in expected:
            teacher = teachers_by_name.get(_normalize_text(item["teacher_name"]))
            actual_class = 0
            actual_f4 = 0
            if teacher and item["career_ids"]:
                assignment_totals = PlanificacionAsignacionDocente.objects.filter(
                    id_docente=teacher,
                    id_periodo=periodo,
                    id_carrera_id__in=item["career_ids"],
                ).aggregate(
                    class_total=Sum("horas_clase"),
                    complementary_total=Sum("horas_complementarias"),
                )
                actual_class = (assignment_totals["class_total"] or 0) + (assignment_totals["complementary_total"] or 0)

                f4_totals = PlanificacionMatrizF4.objects.filter(
                    id_docente=teacher,
                    id_periodo=periodo,
                    id_carrera_id__in=item["career_ids"],
                ).annotate(
                    horas_equivalentes=ExpressionWrapper(
                        F("horas_actividad") * F("numero_paralelos_actividad"),
                        output_field=IntegerField(),
                    )
                ).aggregate(total=Sum("horas_equivalentes"))
                actual_f4 = f4_totals["total"] or 0

            actual_total = actual_class + actual_f4
            diff = actual_total - item["expected_total"]
            rows.append({
                **item,
                "teacher_found": bool(teacher),
                "actual_class": actual_class,
                "actual_f4": actual_f4,
                "actual_total": actual_total,
                "diff": diff,
            })

        visible_rows = [row for row in rows if row["diff"] != 0 or not row["teacher_found"]] if options["solo_diferencias"] else rows
        matched = sum(1 for row in rows if row["diff"] == 0 and row["teacher_found"])
        different = sum(1 for row in rows if row["diff"] != 0 and row["teacher_found"])
        missing = sum(1 for row in rows if not row["teacher_found"])

        self.stdout.write(f"Periodo validado: {periodo.codigo_periodo}")
        self.stdout.write(f"Registros comparados: {len(rows)}")
        self.stdout.write(self.style.SUCCESS(f"Coinciden: {matched}"))
        self.stdout.write(self.style.WARNING(f"Con diferencia: {different}"))
        if missing:
            self.stdout.write(self.style.ERROR(f"Docentes no encontrados: {missing}"))

        if not visible_rows:
            return

        self.stdout.write("")
        self.stdout.write("Fuente | Docente | Excel | Base | Clase | F4 | Diferencia | Carreras")
        for row in visible_rows:
            style = self.style.SUCCESS if row["diff"] == 0 and row["teacher_found"] else self.style.WARNING
            if not row["teacher_found"]:
                style = self.style.ERROR
            carreras = ", ".join(sorted(row["career_labels"])) or "Sin carrera mapeada"
            self.stdout.write(style(
                f"{row['source']} | {row['teacher_name']} | "
                f"{row['expected_total']} | {row['actual_total']} | "
                f"{row['actual_class']} | {row['actual_f4']} | {row['diff']} | {carreras}"
            ))

        if not options["global_docente"]:
            return

        global_rows = []
        for item in expected_global.values():
            teacher = teachers_by_name.get(_normalize_text(item["teacher_name"]))
            actual_class = 0
            actual_f4 = 0
            if teacher:
                assignment_totals = PlanificacionAsignacionDocente.objects.filter(
                    id_docente=teacher,
                    id_periodo=periodo,
                ).aggregate(
                    class_total=Sum("horas_clase"),
                    complementary_total=Sum("horas_complementarias"),
                )
                actual_class = (assignment_totals["class_total"] or 0) + (assignment_totals["complementary_total"] or 0)

                f4_seen = set()
                for f4_row in PlanificacionMatrizF4.objects.filter(
                    id_docente=teacher,
                    id_periodo=periodo,
                ).values(
                    "tipo_actividad",
                    "nombre_asignatura_actividad",
                    "horas_actividad",
                    "numero_paralelos_actividad",
                ):
                    dedupe_key = (
                        f4_row["tipo_actividad"],
                        f4_row["nombre_asignatura_actividad"],
                        f4_row["horas_actividad"],
                        f4_row["numero_paralelos_actividad"],
                    )
                    if dedupe_key in f4_seen:
                        continue
                    f4_seen.add(dedupe_key)
                    actual_f4 += (
                        (f4_row["horas_actividad"] or 0) *
                        (f4_row["numero_paralelos_actividad"] or 1)
                    )

            actual_total = actual_class + actual_f4
            diff = actual_total - item["expected_total"]
            global_rows.append({
                **item,
                "teacher_found": bool(teacher),
                "detail_total": item["detail_class"] + item["detail_f4"],
                "actual_class": actual_class,
                "actual_f4": actual_f4,
                "actual_total": actual_total,
                "diff": diff,
                "detail_diff": actual_total - (item["detail_class"] + item["detail_f4"]),
            })

        visible_global_rows = [
            row for row in global_rows
            if row["diff"] != 0 or not row["teacher_found"]
        ] if options["solo_diferencias"] else global_rows
        matched_global = sum(1 for row in global_rows if row["diff"] == 0 and row["teacher_found"])
        different_global = sum(1 for row in global_rows if row["diff"] != 0 and row["teacher_found"])
        missing_global = sum(1 for row in global_rows if not row["teacher_found"])

        self.stdout.write("")
        self.stdout.write("Resumen global por docente")
        self.stdout.write(f"Docentes comparados: {len(global_rows)}")
        self.stdout.write(self.style.SUCCESS(f"Coinciden: {matched_global}"))
        self.stdout.write(self.style.WARNING(f"Con diferencia: {different_global}"))
        if missing_global:
            self.stdout.write(self.style.ERROR(f"Docentes no encontrados: {missing_global}"))

        if not visible_global_rows:
            return

        self.stdout.write("")
        self.stdout.write("Docente | Excel declarado | Excel detalle | Base global | Clase | F4 | Dif. declarado | Dif. detalle | Fuentes")
        for row in sorted(visible_global_rows, key=lambda item: item["teacher_name"]):
            style = self.style.SUCCESS if row["diff"] == 0 and row["teacher_found"] else self.style.WARNING
            if row["detail_diff"] == 0 and row["teacher_found"]:
                style = self.style.SUCCESS
            if not row["teacher_found"]:
                style = self.style.ERROR
            sources = ", ".join(sorted(row["sources"]))
            self.stdout.write(style(
                f"{row['teacher_name']} | {row['expected_total']} | "
                f"{row['detail_total']} | {row['actual_total']} | "
                f"{row['actual_class']} | {row['actual_f4']} | "
                f"{row['diff']} | {row['detail_diff']} | {sources}"
            ))
