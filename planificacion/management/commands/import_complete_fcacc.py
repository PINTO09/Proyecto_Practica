from collections import defaultdict
from datetime import date
import hashlib
from pathlib import Path
import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from catalogos.models import (
    CatalogoCampoConocimiento,
    CatalogoCarrera,
    CatalogoDedicacionHoraria,
    CatalogoGradoAfinidad,
    CatalogoModalidadContratacion,
    CatalogoPeriodoAcademico,
    CatalogoTipoDocente,
    CatalogoTituloPosgrado,
    CatalogoPais,
)
from curriculo.models import (
    CurriculoAsignatura,
    CurriculoAsignaturaCampo,
    RelacionPosgradoCampo,
)
from docentes.models import (
    DocenteCampoAfinidad,
    DocenteFcacc,
    DocenteTituloAcademico,
)
from planificacion.models import (
    CatalogoActividadComplementaria,
    CargaHistorial,
    PlanificacionActividadDocente,
    PlanificacionAsignacionDocente,
    PlanificacionDemandaAcademica,
    PlanificacionMatrizF4,
)


def _normalize_text(value):
    text = str(value or "").replace("\xa0", " ").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def _clean_text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _fit_code(value, prefix, max_length=20):
    code = _clean_text(value)
    if len(code) <= max_length:
        return code
    digest = hashlib.sha1(code.encode("utf-8")).hexdigest().upper()[: max_length - len(prefix)]
    return f"{prefix}{digest}"


def _sheet_rows(path, sheet_name):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    return rows


def _valid_level(value):
    try:
        level = int(value or 0)
    except (TypeError, ValueError):
        return None
    return level if 1 <= level <= 8 else None


def _normalize_cedula(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if not digits:
        return None
    if len(digits) > 10:
        digits = digits[:10]
    return digits.zfill(10)


def _normalize_phone(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if not digits:
        return None
    if len(digits) == 9:
        digits = "0" + digits
    digits = digits[:15]
    return digits if 10 <= len(digits) <= 15 else None


def _positive_int(value):
    try:
        number = int(float(value or 0))
    except (TypeError, ValueError):
        return 0
    return number if number > 0 else 0


def _valid_email(value):
    email = _clean_text(value).lower()
    email = email.replace("@uleam.eud.ec", "@uleam.edu.ec")
    if not email or not re.match(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$', email):
        return None
    return email


def _valid_tipo_sangre(value):
    VALID_BLOOD_TYPES = {"A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"}
    cleaned = _clean_text(value).upper()
    return cleaned if cleaned in VALID_BLOOD_TYPES else None


def _canonical_name(value):
    text = _normalize_text(value)
    text = re.sub(r"^\d+\s+", "", text)
    text = re.sub(r"\s*\([^)]*\)\s*", " ", text)
    text = text.replace(" - ", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


CAREER_ALIASES = {
    "CONTABILIDAD Y ADITORIA 2024": "CONTABILIDAD Y AUDITORIA 2024 NS",
    "CONTABILIDAD Y AUDITORIA": "CONTABILIDAD Y AUDITORIA 2024 NS",
    "AUDITORIA Y CONTROL DE GESTION 2024": "AUDITORIA Y CONTROL DE GESTION 2024 NS",
    "GESTION DE LA INFORMACION GERENCIAL": "GESTION DE LA INFORMACION GERENCIAL 2024 NS",
    "GESTION DE LA INFORMACION GERENCIAL 2024": "GESTION DE LA INFORMACION GERENCIAL 2024 NS",
    "COMERCIO EXTERIOR": "COMERCIO EXTERIOR 2026 NS",
    "COMERCIO EXTERIOR 2026": "COMERCIO EXTERIOR 2026 NS",
    "MARKETING": "MARKETING 2024 NS",
    "ADMINISTRACION DE EMPRESAS": "ADMINISTRACION DE EMPRESAS",
}

FIELD_ALIASES = {
    "ADMINISTRACION": "ADMINISTRACION",
    "COMPUTACION": "CIENCIAS COMPUTACIONALES",
    "CONTABILIDAD": "CONTABILIDAD Y AUDITORIA",
    "DERECHO": "DERECHO",
    "ESTADISTICA": "ECONOMIA MATEMATICA",
    "INFORMATICA": "CIENCIAS COMPUTACIONALES",
    "MATEMATICA": "ECONOMIA MATEMATICA",
    "SOCIALES": "UNIDAD BASICA",
}

F4_ACTIVITY_CAREER_CODES = {
    "FCACC-1-DO-CL",
    "FCACC-4-DO-TU-TI",
    "FCACC-5-DO-PE-IN",
    "FCACC-6-IVN",
    "FCACC-8-VI-SO",
    "FCACC-9-GE_ED",
    "FCACC-8-PRAC",
    "FCACC-8-TITU",
}


def _log_carga(tipo_carga, archivo_origen, total, creados, actualizados, omitidos, errores=None, estado="COMPLETADO"):
    try:
        # El savepoint evita dejar inservible la transaccion principal si una
        # instalacion antigua aun no tiene la tabla de historial.
        with transaction.atomic():
            CargaHistorial.objects.create(
                tipo_carga=tipo_carga,
                archivo_origen=archivo_origen,
                total_registros=total,
                registros_creados=creados,
                registros_actualizados=actualizados,
                registros_omitidos=omitidos,
                detalle_errores=errores,
                estado=estado,
            )
    except Exception:
        pass


def _planning_sources(main_book, coor_files):
    """Fuentes en orden de prioridad: cada coordinacion reemplaza al consolidado."""
    sources = [(main_book.name, _sheet_rows(main_book, "ASIGNACION"))]
    for subdir, path in coor_files:
        sources.append((f"{subdir}/{path.name}", _sheet_rows(path, "ASIGNACION")))
    return sources


class Command(BaseCommand):
    help = "Importacion completa FCACC: 12 pasos desde los 18+ Excel en PLANIFICACION Copiar3"

    def add_arguments(self, parser):
        parser.add_argument(
            "--base-dir",
            default=None,
            help="Directorio base con los archivos Excel (PLANIFICACION Copiar3).",
        )
        parser.add_argument(
            "--periodo-codigo",
            default="2026-2",
            help="Codigo del periodo academico.",
        )
        parser.add_argument(
            "--periodo-nombre",
            default="2026-2",
            help="Nombre del periodo academico.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Analiza y valida sin guardar cambios.",
        )
        parser.add_argument(
            "--sync-asignaciones",
            action="store_true",
            help="Elimina asignaciones del periodo que no aparecen en la carga actual.",
        )
        parser.add_argument(
            "--step",
            type=int,
            default=0,
            help="Ejecutar solo un paso especifico (1-12). 0 = todos.",
        )

    def handle(self, *args, **options):
        project_dir = Path(__file__).resolve().parents[3]
        base_dir = Path(options["base_dir"] or project_dir / "_excel_input" / "PLANIFICACION Copiar3")
        dry_run = options["dry_run"]
        sync_assignments = options["sync_asignaciones"]
        only_step = options["step"]

        if not base_dir.exists():
            raise CommandError(f"No se encontro el directorio base: {base_dir}")

        main_book = base_dir / "FCACC-PLANIFICACION.xlsx"
        revision_book = base_dir / "revision" / "REVISIONDocentes.xlsx"
        malla_book = base_dir / "revision" / "MALLA.xlsx"
        nivel_docente_book = base_dir / "NIVEL_DOCENTE.xlsx"
        docentes_book = base_dir / "Docentes.xlsx"
        maestria_book = base_dir / "Maestria.xlsx"
        detallado_maestria_book = base_dir / "DETALLADO_MAESTRIA.xlsx"
        detallado_book = base_dir / "DETALLADO.xlsx"
        detallado_asig_book = base_dir / "DETALLADO_ASIGNATURA.xlsx"

        for book in [
            main_book, revision_book, nivel_docente_book, docentes_book,
            maestria_book, detallado_maestria_book, detallado_book,
            detallado_asig_book,
        ]:
            if not book.exists():
                raise CommandError(f"No se encontro: {book}")

        coor_files = []
        coor_dirnames = {
            "COOR_CEXT": "FCACC-PLANIFICACION COM_MAR.xlsx",
            "COOR_CONT": "FCACC-PLANIFICACION CONTABILIDAD.xlsx",
            "COOR_EMPRE": "FCACC-PLANIFICACION EMPRESAS.xlsx",
        }
        for subdir, fname in coor_dirnames.items():
            fp = base_dir / subdir / fname
            if fp.exists():
                coor_files.append((subdir, fp))

        self.stdout.write(f"Leyendo datos desde: {base_dir}")
        self.stdout.write(f"Archivo principal: {main_book.name}")
        self.stdout.write(f"Archivos COOR: {[s for s, _ in coor_files]}")
        self.stdout.write(f"Periodo: {options['periodo_codigo']} | Dry-run: {dry_run}")

        tipo_docente_default = CatalogoTipoDocente.objects.filter(
            nombre_tipo_docente__icontains="TITULAR"
        ).first() or CatalogoTipoDocente.objects.filter(
            codigo_tipo_docente__icontains="TIT"
        ).first() or CatalogoTipoDocente.objects.first()
        if not tipo_docente_default:
            raise CommandError("No existe ningun catalogo de tipo docente.")
        grado_afinidad_default = CatalogoGradoAfinidad.objects.filter(
            nombre_grado_afinidad__icontains="MEDIA"
        ).first() or CatalogoGradoAfinidad.objects.filter(
            codigo_grado_afinidad__icontains="MEDIA"
        ).first() or CatalogoGradoAfinidad.objects.order_by("nivel_prioridad").first()
        if not grado_afinidad_default:
            raise CommandError("No existe ningun catalogo de grado de afinidad.")

        modalidad_map = {}
        for obj in CatalogoModalidadContratacion.objects.all():
            modalidad_map[_normalize_text(obj.codigo_modalidad)] = obj
            modalidad_map[_normalize_text(obj.nombre_modalidad)] = obj
        dedicacion_map = {}
        for obj in CatalogoDedicacionHoraria.objects.all():
            dedicacion_map[_normalize_text(obj.codigo_dedicacion)] = obj
            dedicacion_map[_normalize_text(obj.nombre_dedicacion)] = obj
        modalidad_aliases = {
            "CONTRATOS OCASIONALES": "CONTRATO",
            "CONTRATO OCASIONAL": "CONTRATO",
            "NOMBRAMIENTO PROVISIONAL": "NOMBRAMIENTO",
            "NOMBRAMIENTO AUTORIDAD UNIVERS": "NOMBRAMIENTO",
        }

        pais_default = CatalogoPais.objects.filter(codigo_iso_pais="EC").first()
        if not pais_default:
            pais_default, _ = CatalogoPais.objects.get_or_create(
                codigo_iso_pais="EC",
                defaults={"nombre_pais": "Ecuador", "nombre_nacionalidad": "Ecuatoriana"},
            )

        with transaction.atomic():
            periodo, _ = CatalogoPeriodoAcademico.objects.update_or_create(
                codigo_periodo=options["periodo_codigo"],
                defaults={
                    "nombre_periodo": options["periodo_nombre"],
                    "periodo_activo": True,
                    "fecha_inicio_periodo": date(2026, 5, 1),
                    "fecha_fin_periodo": date(2026, 9, 30),
                },
            )

            steps_results = {}

            # ── STEP 1: CARRERAS ──
            malla_career_by_name = {}
            if only_step in (0, 1):
                malla_career_by_name = self._step1_carreras(detallado_asig_book, steps_results)
                self.stdout.write(self.style.SUCCESS(f"  Step 1: {steps_results.get('step1', {}).get('msg', 'OK')}"))

            # ── STEP 2: CAMPOS DE CONOCIMIENTO ──
            if only_step in (0, 2):
                self._step2_campos(detallado_book, steps_results)
                self.stdout.write(self.style.SUCCESS(f"  Step 2: {steps_results.get('step2', {}).get('msg', 'OK')}"))

            # ── STEP 3: POSGRADOS ──
            if only_step in (0, 3):
                self._step3_posgrados(maestria_book, main_book, steps_results)
                self.stdout.write(self.style.SUCCESS(f"  Step 3: {steps_results.get('step3', {}).get('msg', 'OK')}"))

            # Build lookups needed by subsequent steps
            career_by_code = {obj.codigo_carrera: obj for obj in CatalogoCarrera.objects.all()}
            careers_by_name = {_canonical_name(obj.nombre_carrera): obj for obj in CatalogoCarrera.objects.all()}
            if not malla_career_by_name and malla_book.exists():
                try:
                    for row in _sheet_rows(malla_book, "CARRERAS_FCACC")[1:]:
                        if row and row[0] and row[1]:
                            key = _canonical_name(row[1])
                            code = _clean_text(row[0])
                            if key not in malla_career_by_name and code in career_by_code:
                                malla_career_by_name[key] = code
                except Exception:
                    pass
            field_by_code = {obj.codigo_campo: obj for obj in CatalogoCampoConocimiento.objects.all()}
            field_by_name = {_normalize_text(obj.nombre_campo_conocimiento): obj for obj in CatalogoCampoConocimiento.objects.all()}
            posgrado_by_code = {obj.codigo_posgrado: obj for obj in CatalogoTituloPosgrado.objects.all()}
            posgrado_by_name = {_normalize_text(obj.nombre_titulo_posgrado): obj for obj in CatalogoTituloPosgrado.objects.all()}

            # ── STEP 4: ASIGNATURAS ──
            subject_code_map = {}
            subject_by_code = {}
            subjects_by_lookup = defaultdict(list)
            if only_step in (0, 4):
                self._step4_asignaturas(main_book, coor_files, detallado_asig_book, career_by_code, careers_by_name, subject_code_map, subject_by_code, subjects_by_lookup, steps_results)
                self.stdout.write(self.style.SUCCESS(f"  Step 4: {steps_results.get('step4', {}).get('msg', 'OK')}"))

            subject_by_code = {obj.codigo_asignatura: obj for obj in CurriculoAsignatura.objects.select_related("id_carrera")}
            subjects_by_lookup = defaultdict(list)
            for obj in subject_by_code.values():
                key = (
                    _canonical_name(obj.id_carrera.nombre_carrera),
                    int(obj.nivel_semestre or 0),
                    _canonical_name(obj.nombre_asignatura),
                )
                subjects_by_lookup[key].append(obj)

            # ── STEP 5: ASIGNATURA-CAMPO ──
            if only_step in (0, 5):
                self._step5_asignatura_campo(detallado_asig_book, subject_code_map, subject_by_code, field_by_code, steps_results)
                self.stdout.write(self.style.SUCCESS(f"  Step 5: {steps_results.get('step5', {}).get('msg', 'OK')}"))

            # ── STEP 6: POSGRADO-CAMPO ──
            if only_step in (0, 6):
                self._step6_posgrado_campo(detallado_maestria_book, posgrado_by_code, posgrado_by_name, field_by_code, field_by_name, steps_results)
                self.stdout.write(self.style.SUCCESS(f"  Step 6: {steps_results.get('step6', {}).get('msg', 'OK')}"))

            # ── STEP 7: DOCENTES ──
            if only_step in (0, 7):
                self._step7_docentes(docentes_book, modalidad_map, modalidad_aliases, dedicacion_map, tipo_docente_default, steps_results)
                self.stdout.write(self.style.SUCCESS(f"  Step 7: {steps_results.get('step7', {}).get('msg', 'OK')}"))

            teacher_by_name = {_normalize_text(obj.nombres_completos): obj for obj in DocenteFcacc.objects.all()}
            teacher_by_cedula = {_normalize_cedula(obj.cedula_docente): obj for obj in DocenteFcacc.objects.all() if _normalize_cedula(obj.cedula_docente)}

            # ── STEP 8: TITULOS DOCENTES ──
            if only_step in (0, 8):
                self._step8_titulos(nivel_docente_book, main_book, teacher_by_name, teacher_by_cedula, posgrado_by_code, posgrado_by_name, pais_default, steps_results)
                self.stdout.write(self.style.SUCCESS(f"  Step 8: {steps_results.get('step8', {}).get('msg', 'OK')}"))

            # ── STEP 9: DOCENTE-CAMPO AFINIDAD ──
            if only_step in (0, 9):
                self._step9_afinidad(revision_book, teacher_by_name, field_by_name, steps_results)
                self.stdout.write(self.style.SUCCESS(f"  Step 9: {steps_results.get('step9', {}).get('msg', 'OK')}"))

            # ── STEP 10: DEMANDAS ACADEMICAS ──
            if only_step in (0, 10):
                self._step10_demandas(main_book, coor_files, careers_by_name, subjects_by_lookup, periodo, steps_results)
                self.stdout.write(self.style.SUCCESS(f"  Step 10: {steps_results.get('step10', {}).get('msg', 'OK')}"))

            subject_field_by_id = defaultdict(set)
            for relation in CurriculoAsignaturaCampo.objects.select_related("id_campo", "id_asignatura"):
                subject_field_by_id[relation.id_asignatura_id].add(relation.id_campo_id)
            teacher_field_by_id = defaultdict(set)
            for teacher_id, field_id in DocenteCampoAfinidad.objects.values_list("id_docente_id", "id_campo_id"):
                teacher_field_by_id[teacher_id].add(field_id)
            posgrado_field_by_id = defaultdict(set)
            for posgrado_id, field_id in RelacionPosgradoCampo.objects.values_list("id_posgrado_id", "id_campo_id"):
                posgrado_field_by_id[posgrado_id].add(field_id)
            for teacher_id, posgrado_id in DocenteTituloAcademico.objects.filter(
                id_posgrado__isnull=False
            ).values_list("id_docente_id", "id_posgrado_id"):
                teacher_field_by_id[teacher_id].update(posgrado_field_by_id.get(posgrado_id, set()))

            # ── STEP 11: ASIGNACIONES DOCENTES ──
            if only_step in (0, 11):
                self._step11_asignaciones(main_book, coor_files, careers_by_name, subjects_by_lookup, teacher_by_name, teacher_by_cedula, field_by_name, field_by_code, subject_field_by_id, teacher_field_by_id, periodo, sync_assignments, steps_results)
                self.stdout.write(self.style.SUCCESS(f"  Step 11: {steps_results.get('step11', {}).get('msg', 'OK')}"))

            # ── STEP 12: MATRIZ F4 ──
            if only_step in (0, 12):
                self._step12_actividades(main_book, coor_files, teacher_by_name, teacher_by_cedula, periodo, steps_results)
                self.stdout.write(self.style.SUCCESS(f"  Step 12: {steps_results.get('step12', {}).get('msg', 'OK')}"))

            if dry_run:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("Dry run completado sin guardar cambios"))
            else:
                self.stdout.write(self.style.SUCCESS("Importacion completada"))

        for step_num in sorted(steps_results.keys()):
            r = steps_results[step_num]
            self.stdout.write(f"  {r['label']}: +{r['creados']} ~{r['actualizados']} -{r['omitidos']} (total {r['total']})")

    # ── STEP 1: CARRERAS ──
    def _step1_carreras(self, catalog_book, results):
        seen = set()
        creados = 0
        actualizados = 0
        omitidos = 0
        malla_career_by_name = {}
        for sheet_name, book in [("MAE_CARRERA", catalog_book)]:
            try:
                rows = _sheet_rows(book, sheet_name)
            except Exception:
                omitidos += 1
                continue
            for row in rows[1:]:
                if not row or not row[0] or not row[1]:
                    omitidos += 1
                    continue
                codigo = _clean_text(row[0])
                if codigo in seen:
                    omitidos += 1
                    continue
                seen.add(codigo)
                nombre = _clean_text(row[1])
                activa = True
                if len(row) > 6:
                    activa = _normalize_text(row[6]) != "NO VIGENTE"
                es_actividad = codigo in F4_ACTIVITY_CAREER_CODES
                _, created = CatalogoCarrera.objects.update_or_create(
                    codigo_carrera=codigo,
                    defaults={
                        "nombre_carrera": nombre,
                        "carrera_activa": activa and not es_actividad,
                        "es_actividad": es_actividad,
                    },
                )
                if created:
                    creados += 1
                else:
                    actualizados += 1
        total = creados + actualizados + omitidos
        _log_carga("CARRERAS", catalog_book.name, total, creados, actualizados, omitidos)
        results["step1"] = {"label": "Carreras", "total": total, "creados": creados, "actualizados": actualizados, "omitidos": omitidos, "msg": f"{creados} creadas, {actualizados} actualizadas, {omitidos} omitidas"}
        return malla_career_by_name

    # ── STEP 2: CAMPOS DE CONOCIMIENTO ──
    def _step2_campos(self, detallado_book, results):
        seen = set()
        creados = 0
        actualizados = 0
        omitidos = 0
        sources = [("MAE_CONOCIMIENTO", detallado_book)]
        for sheet_name, book in sources:
            rows = _sheet_rows(book, sheet_name)
            for row in rows[1:]:
                if not row or not row[0] or not row[1]:
                    omitidos += 1
                    continue
                codigo = _clean_text(row[0])
                if codigo in seen:
                    continue
                seen.add(codigo)
                nombre = _clean_text(row[1])
                _, created = CatalogoCampoConocimiento.objects.update_or_create(
                    codigo_campo=codigo,
                    defaults={"nombre_campo_conocimiento": nombre},
                )
                if created:
                    creados += 1
                else:
                    actualizados += 1
        total = creados + actualizados + omitidos
        _log_carga("CAMPOS_CONOCIMIENTO", detallado_book.name, total, creados, actualizados, omitidos)
        results["step2"] = {"label": "Campos Conocimiento", "total": total, "creados": creados, "actualizados": actualizados, "omitidos": omitidos, "msg": f"{creados} creados, {actualizados} actualizados, {omitidos} omitidos"}

    # ── STEP 3: POSGRADOS ──
    def _step3_posgrados(self, maestria_book, main_book, results):
        seen_names = set()
        creados = 0
        actualizados = 0
        omitidos = 0
        existing_by_name = {
            _normalize_text(obj.nombre_titulo_posgrado): obj
            for obj in CatalogoTituloPosgrado.objects.all()
        }
        code_owner = {
            obj.codigo_posgrado: _normalize_text(obj.nombre_titulo_posgrado)
            for obj in CatalogoTituloPosgrado.objects.all()
        }
        sources = [("MAESTRIA", maestria_book)]
        for sheet_name, book in sources:
            rows = _sheet_rows(book, sheet_name)
            for row in rows[1:]:
                if not row or not row[1]:
                    omitidos += 1
                    continue
                nombre = _clean_text(row[1])
                name_key = _normalize_text(nombre)
                if name_key in seen_names:
                    continue
                seen_names.add(name_key)
                obj = existing_by_name.get(name_key)
                if obj:
                    obj.nombre_titulo_posgrado = nombre
                    obj.save(update_fields=["nombre_titulo_posgrado"])
                    created = False
                else:
                    source_code = _clean_text(row[0]) if row[0] else nombre
                    codigo = _fit_code(source_code, "POS-")
                    if codigo in code_owner and code_owner[codigo] != name_key:
                        codigo = _fit_code(f"{source_code}|{name_key}", "POS-")
                    obj = CatalogoTituloPosgrado.objects.create(
                        codigo_posgrado=codigo,
                        nombre_titulo_posgrado=nombre,
                    )
                    existing_by_name[name_key] = obj
                    code_owner[codigo] = name_key
                    created = True
                if created:
                    creados += 1
                else:
                    actualizados += 1
        total = creados + actualizados + omitidos
        _log_carga("POSGRADOS", f"Maestria.xlsx, {main_book.name}", total, creados, actualizados, omitidos)
        results["step3"] = {"label": "Posgrados", "total": total, "creados": creados, "actualizados": actualizados, "omitidos": omitidos, "msg": f"{creados} creados, {actualizados} actualizados, {omitidos} omitidos"}

    # ── STEP 4: ASIGNATURAS ──
    # ---- STEP 4: ASIGNATURAS (desde FCACC + COOR + MALLA + DETALLADO_ASIGNATURA) ----
    def _step4_asignaturas(self, main_book, coor_files, detallado_asig_book, career_by_code, careers_by_name, subject_code_map, subject_by_code, subjects_by_lookup, results):
        creados = 0
        actualizados = 0
        omitidos = 0
        seen_codes = set()
        # El catalogo consolidado es la fuente maestra. Los libros operativos
        # solo completan codigos que aun no aparecen en el catalogo.
        sources = [("MAE_ASIGNATURA", detallado_asig_book)]
        for _subdir, fp in coor_files:
            sources.append(("MAE_ASIGNATURA", fp))
        sources.append(("MAE_ASIGNATURA", main_book))
        for sheet_name, book in sources:
            try:
                rows = _sheet_rows(book, sheet_name)
            except Exception:
                omitidos += 1
                continue
            for row in rows[1:]:
                if not row or not row[0] or not row[1]:
                    omitidos += 1
                    continue
                carrera_code = _clean_text(row[4]) if len(row) > 4 and row[4] else ""
                if not carrera_code:
                    omitidos += 1
                    continue
                carrera_code_key = _canonical_name(carrera_code)
                carrera = career_by_code.get(carrera_code) or careers_by_name.get(carrera_code_key)
                if not carrera:
                    omitidos += 1
                    continue
                subject_es_act = carrera.es_actividad if hasattr(carrera, 'es_actividad') else False
                if subject_es_act:
                    # Las actividades ya no se modelan como asignaturas.
                    # Se cargan en el paso 12 en su catalogo independiente.
                    omitidos += 1
                    continue
                level = _valid_level(row[6]) if len(row) > 6 else None
                if level is None:
                    omitidos += 1
                    continue
                source_code = _clean_text(row[0])
                if source_code in seen_codes:
                    continue
                seen_codes.add(source_code)
                compact_code = _fit_code(source_code, "ASG-")
                subject_code_map[source_code] = compact_code
                obj, created = CurriculoAsignatura.objects.update_or_create(
                    codigo_asignatura=compact_code,
                    defaults={
                        "id_carrera": carrera,
                        "nombre_asignatura": _clean_text(row[1]),
                        "horas_semanales_asignatura": int(row[2] or 0),
                        "nivel_semestre": level,
                        "es_actividad": subject_es_act,
                    },
                )
                if created:
                    creados += 1
                else:
                    actualizados += 1
                subject_by_code[compact_code] = obj
                key = (
                    _canonical_name(obj.id_carrera.nombre_carrera),
                    int(obj.nivel_semestre or 0),
                    _canonical_name(obj.nombre_asignatura),
                )
                subjects_by_lookup[key].append(obj)
        total = creados + actualizados + omitidos
        _log_carga("ASIGNATURAS", f"{detallado_asig_book.name} + catalogos operativos", total, creados, actualizados, omitidos)
        results["step4"] = {"label": "Asignaturas", "total": total, "creados": creados, "actualizados": actualizados, "omitidos": omitidos, "msg": f"{creados} creadas, {actualizados} actualizadas, {omitidos} omitidas"}
    # ---- STEP 5: ASIGNATURA-CAMPO (desde FCACC + COOR + DETALLADO_ASIGNATURA) ----
    def _step5_asignatura_campo(self, detallado_asig_book, subject_code_map, subject_by_code, field_by_code, results):
        creados = 0
        ya_existentes = 0
        omitidos = 0
        seen_pairs = set()
        sources = [("DET_ASIG", detallado_asig_book)]
        for sheet_name, book in sources:
            try:
                rows = _sheet_rows(book, sheet_name)
            except Exception:
                omitidos += 1
                continue
            for row in rows[1:]:
                if not row or len(row) < 8 or not row[5] or not row[7]:
                    omitidos += 1
                    continue
                source_subject_code = _clean_text(row[5])
                compact_subject_code = subject_code_map.get(source_subject_code, _fit_code(source_subject_code, "ASG-"))
                subject = subject_by_code.get(compact_subject_code)
                if not subject:
                    omitidos += 1
                    continue
                field_code = _clean_text(row[7])
                field = field_by_code.get(field_code)
                if not field:
                    omitidos += 1
                    continue
                pair = (compact_subject_code, field_code)
                if pair in seen_pairs:
                    continue
                seen_pairs.add(pair)
                _, created = CurriculoAsignaturaCampo.objects.get_or_create(
                    id_asignatura=subject,
                    id_campo=field,
                )
                if created:
                    creados += 1
                else:
                    ya_existentes += 1
        total = creados + ya_existentes + omitidos
        _log_carga("ASIGNATURA_CAMPO", detallado_asig_book.name, total, creados, ya_existentes, omitidos)
        results["step5"] = {"label": "Asignatura-Campo", "total": total, "creados": creados, "actualizados": ya_existentes, "omitidos": omitidos, "msg": f"{creados} creadas, {ya_existentes} ya existian, {omitidos} omitidas"}

    # ── STEP 6: POSGRADO-CAMPO ──
    def _step6_posgrado_campo(self, detallado_maestria_book, posgrado_by_code, posgrado_by_name, field_by_code, field_by_name, results):
        rows = _sheet_rows(detallado_maestria_book, "MAESTRIA_DETALLADO")
        creados = 0
        existentes = 0
        omitidos = 0
        valid_pairs = set()
        source_posgrad_ids = set()
        for row in rows[1:]:
            if not row or len(row) < 5 or not row[1] or not row[3]:
                omitidos += 1
                continue
            posgrado_name = _clean_text(row[1])
            posgrado_code = _clean_text(row[2]) if row[2] else None
            # El codigo original no es unico; el nombre es la identidad real.
            posgrado = posgrado_by_name.get(_normalize_text(posgrado_name))
            if not posgrado and posgrado_code:
                posgrado = posgrado_by_code.get(posgrado_code)
            if not posgrado:
                omitidos += 1
                continue
            field_code = _clean_text(row[4]) if len(row) > 4 and row[4] else None
            field_name_raw = _clean_text(row[3])
            field = None
            if field_code:
                field = field_by_code.get(field_code)
            if not field:
                alias = FIELD_ALIASES.get(_canonical_name(field_name_raw), field_name_raw)
                field = field_by_name.get(_normalize_text(alias))
            if not field:
                omitidos += 1
                continue
            pair = (posgrado.id_posgrado, field.id_campo)
            if pair in valid_pairs:
                continue
            valid_pairs.add(pair)
            source_posgrad_ids.add(posgrado.id_posgrado)
            _, created = RelacionPosgradoCampo.objects.get_or_create(
                id_posgrado=posgrado,
                id_campo=field,
            )
            if created:
                creados += 1
            else:
                existentes += 1
        stale_removed = 0
        for relation in RelacionPosgradoCampo.objects.filter(id_posgrado_id__in=source_posgrad_ids):
            if (relation.id_posgrado_id, relation.id_campo_id) not in valid_pairs:
                relation.delete()
                stale_removed += 1
        total = creados + existentes + omitidos
        detail = f"Relaciones obsoletas retiradas: {stale_removed}" if stale_removed else None
        _log_carga("POSGRADO_CAMPO", detallado_maestria_book.name, total, creados, existentes, omitidos, detail)
        results["step6"] = {"label": "Posgrado-Campo", "total": total, "creados": creados, "actualizados": existentes, "omitidos": omitidos, "msg": f"{creados} creadas, {existentes} existentes, {omitidos} omitidas; {stale_removed} obsoletas retiradas"}

    # ── STEP 7: DOCENTES ──
    def _step7_docentes(self, docentes_book, modalidad_map, modalidad_aliases, dedicacion_map, tipo_docente_default, results):
        rows = _sheet_rows(docentes_book, "MDOCENTES")
        creados = 0
        actualizados = 0
        omitidos = 0
        for row in rows[1:]:
            if not row or not row[0] or not row[1]:
                omitidos += 1
                continue
            cedula = _normalize_cedula(row[0])
            if not cedula:
                omitidos += 1
                continue
            modalidad_key = _normalize_text(row[7]) if len(row) > 7 else ""
            modalidad = modalidad_map.get(modalidad_key) or modalidad_map.get(modalidad_aliases.get(modalidad_key, ""))
            if not modalidad:
                omitidos += 1
                continue
            dedicacion = dedicacion_map.get(_normalize_text(row[6])) if len(row) > 6 else None
            if not dedicacion:
                dedicacion = CatalogoDedicacionHoraria.objects.first()
                if not dedicacion:
                    omitidos += 1
                    continue
            correo = _valid_email(row[3]) if len(row) > 3 else None
            if correo and DocenteFcacc.objects.exclude(cedula_docente=cedula).filter(
                correo_institucional__iexact=correo
            ).exists():
                correo = None
            celular = _normalize_phone(row[4]) if len(row) > 4 else None
            sangre = _valid_tipo_sangre(row[8]) if len(row) > 8 else None
            docente = DocenteFcacc.objects.filter(cedula_docente=cedula).first()
            created = docente is None
            if created:
                docente = DocenteFcacc(cedula_docente=cedula)
            docente.tipo_documento = docente.tipo_documento or "CEDULA"
            if created or not docente.id_tipo_docente_id:
                docente.id_tipo_docente = tipo_docente_default
            docente.id_modalidad = modalidad
            docente.id_dedicacion = dedicacion
            docente.nombres_completos = _clean_text(row[1])
            docente.unidad_organica = _clean_text(row[2]) if len(row) > 2 else docente.unidad_organica
            # Los vacios o valores invalidos del Excel nunca borran un dato
            # valido que ya hubiese sido completado manualmente en el sistema.
            if correo:
                docente.correo_institucional = correo
            if celular:
                docente.numero_celular = celular
            if sangre:
                docente.tipo_sangre = sangre
            docente.docente_activo = True
            if created:
                docente.save()
            else:
                docente.save(update_fields=[
                    "tipo_documento", "id_tipo_docente", "id_modalidad", "id_dedicacion",
                    "nombres_completos", "unidad_organica", "correo_institucional",
                    "numero_celular", "tipo_sangre", "docente_activo",
                ])
            if created:
                creados += 1
            else:
                actualizados += 1
        total = creados + actualizados + omitidos
        _log_carga("DOCENTES", docentes_book.name, total, creados, actualizados, omitidos)
        results["step7"] = {"label": "Docentes", "total": total, "creados": creados, "actualizados": actualizados, "omitidos": omitidos, "msg": f"{creados} creados, {actualizados} actualizados, {omitidos} omitidos"}

    # ── STEP 8: TITULOS DOCENTES ──
    def _step8_titulos(self, nivel_docente_book, main_book, teacher_by_name, teacher_by_cedula, posgrado_by_code, posgrado_by_name, pais_default, results):
        creados = 0
        omitidos = 0
        sources = [("EDU_DOCENTE", nivel_docente_book), ("EDU_DOCENTE", main_book)]
        for sheet_name, book in sources:
            rows = _sheet_rows(book, sheet_name)
            for row in rows[1:]:
                if not row or not row[1]:
                    omitidos += 1
                    continue
                teacher_name = _clean_text(row[1])
                if _normalize_text(teacher_name) in {"SELECCIONE DOCENTE", "NO DEFINIDO", "#N/A", ""}:
                    omitidos += 1
                    continue
                docente = teacher_by_name.get(_normalize_text(teacher_name))
                if not docente and len(row) > 2 and row[2]:
                    docente = teacher_by_cedula.get(_normalize_cedula(row[2]))
                if not docente:
                    omitidos += 1
                    continue
                titulo_nombre = _clean_text(row[3]) if len(row) > 3 and row[3] else ""
                if not titulo_nombre:
                    omitidos += 1
                    continue
                posgrado_code = _clean_text(row[5]) if len(row) > 5 and row[5] else ""
                posgrado = posgrado_by_name.get(_normalize_text(titulo_nombre))
                if not posgrado and posgrado_code:
                    posgrado = posgrado_by_code.get(posgrado_code)
                if not pais_default:
                    omitidos += 1
                    continue
                _, created = DocenteTituloAcademico.objects.update_or_create(
                    id_docente=docente,
                    nombre_titulo=titulo_nombre,
                    defaults={
                        "id_pais": pais_default,
                        "id_posgrado": posgrado,
                        "nivel_titulo": 4,
                        "numero_registro_senescyt": None,
                        "fecha_registro_senescyt": None,
                    },
                )
                if created:
                    creados += 1
        total = creados + omitidos
        _log_carga("TITULOS_DOCENTES", f"NIVEL_DOCENTE.xlsx, {main_book.name}", total, creados, 0, omitidos)
        results["step8"] = {"label": "Titulos Docentes", "total": total, "creados": creados, "actualizados": 0, "omitidos": omitidos, "msg": f"{creados} creados, {omitidos} omitidos"}

    # ── STEP 9: DOCENTE-CAMPO AFINIDAD ──
    def _step9_afinidad(self, revision_book, teacher_by_name, field_by_name, results):
        rows = _sheet_rows(revision_book, "DET_DOCENTE")
        creados = 0
        omitidos = 0
        for row in rows[1:]:
            if not row or not row[1] or not row[2]:
                omitidos += 1
                continue
            teacher = teacher_by_name.get(_normalize_text(row[1]))
            if not teacher:
                omitidos += 1
                continue
            field_values = [f.strip() for f in str(row[2]).split(";")]
            for fv in field_values:
                fv = _clean_text(fv)
                if not fv:
                    continue
                alias_name = FIELD_ALIASES.get(_canonical_name(fv), fv)
                field = field_by_name.get(_normalize_text(alias_name))
                if not field:
                    continue
                _, created = DocenteCampoAfinidad.objects.get_or_create(
                    id_docente=teacher,
                    id_campo=field,
                )
                if created:
                    creados += 1
        total = creados + omitidos
        _log_carga("AFINIDAD_DOCENTE_CAMPO", revision_book.name, total, creados, 0, omitidos)
        results["step9"] = {"label": "Afinidad Docente-Campo", "total": total, "creados": creados, "actualizados": 0, "omitidos": omitidos, "msg": f"{creados} creadas, {omitidos} omitidas"}

    # ── STEP 10: DEMANDAS ACADEMICAS ──
    def _step10_demandas(self, main_book, coor_files, careers_by_name, subjects_by_lookup, periodo, results):
        demand_sources = _planning_sources(main_book, coor_files)
        demand_bucket = {}
        creados = 0
        omitidos = 0
        for source_name, rows in demand_sources:
            for row in rows[1:]:
                if not row or not any(value not in (None, "") for value in row):
                    continue
                if len(row) <= 5:
                    omitidos += 1
                    continue
                if not row[1] and not row[5]:
                    continue
                carrera_name = _clean_text(row[1])
                nivel = _valid_level(row[3])
                paralelo = _clean_text(row[4])
                subject_name = _clean_text(row[5])
                if not carrera_name or not subject_name or nivel is None:
                    # Las actividades complementarias se importan en el paso
                    # 12 y nunca se convierten en demanda academica.
                    continue
                canonical_career = _canonical_name(carrera_name)
                canonical_career = CAREER_ALIASES.get(canonical_career, canonical_career)
                canonical_subject = _canonical_name(subject_name)
                carrera = careers_by_name.get(canonical_career)
                subject_matches = subjects_by_lookup.get((canonical_career, nivel, canonical_subject), [])
                subject = subject_matches[0] if subject_matches else None
                if not carrera or not subject:
                    omitidos += 1
                    continue
                if carrera.es_actividad or subject.es_actividad:
                    continue
                key = (carrera.id_carrera, subject.id_asignatura)
                item = demand_bucket.setdefault(key, {"carrera": carrera, "subject": subject, "parallels": set(), "blank_rows": 0})
                if paralelo:
                    item["parallels"].add(paralelo)
                else:
                    item["blank_rows"] += 1
        for item in demand_bucket.values():
            numero_paralelos = max(1, len(item["parallels"]) + item["blank_rows"])
            _, created = PlanificacionDemandaAcademica.objects.update_or_create(
                id_asignatura=item["subject"],
                id_carrera=item["carrera"],
                id_periodo=periodo,
                defaults={"proyeccion_estudiantes": 0, "numero_paralelos": numero_paralelos},
            )
            if created:
                creados += 1
        total = omitidos + len(demand_bucket)
        _log_carga(
            "DEMANDAS_ACADEMICAS",
            f"{main_book.name} + {len(coor_files)} coordinaciones",
            total, creados, len(demand_bucket) - creados, omitidos,
        )
        results["step10"] = {"label": "Demandas Academicas", "total": total, "creados": creados, "actualizados": len(demand_bucket) - creados, "omitidos": omitidos, "msg": f"{creados} creadas, {len(demand_bucket) - creados} actualizadas, {omitidos} omitidas"}

    # ── STEP 11: ASIGNACIONES DOCENTES ──
    def _step11_asignaciones(self, main_book, coor_files, careers_by_name, subjects_by_lookup, teacher_by_name, teacher_by_cedula, field_by_name, field_by_code, subject_field_by_id, teacher_field_by_id, periodo, sync_assignments, results):
        assignment_sources = _planning_sources(main_book, coor_files)
        default_field = field_by_name.get("UNIDAD BASICA") or CatalogoCampoConocimiento.objects.order_by("id_campo").first()
        assignment_bucket = {}
        invalid_affinity_keys = set()
        skip_count = 0
        conflicts = []
        created_assignments = 0
        updated_assignments = 0
        merged_duplicates = 0
        for source_name, rows in assignment_sources:
            for row in rows[1:]:
                if not row or not any(value not in (None, "") for value in row):
                    continue
                if len(row) <= 22:
                    skip_count += 1
                    continue
                if not row[1] and not row[5]:
                    continue
                carrera_name = _clean_text(row[1])
                nivel = _valid_level(row[3])
                subject_name = _clean_text(row[5])
                teacher_name = _clean_text(row[12])
                if not carrera_name or not subject_name or not teacher_name or nivel is None:
                    skip_count += 1
                    continue
                if _normalize_text(teacher_name) in {"SELECCIONE DOCENTE", "NO DEFINIDO", "#N/A", ""}:
                    skip_count += 1
                    continue
                horas_clase = _positive_int(row[11])
                if horas_clase <= 0:
                    skip_count += 1
                    continue
                canonical_career = _canonical_name(carrera_name)
                canonical_career = CAREER_ALIASES.get(canonical_career, canonical_career)
                canonical_subject = _canonical_name(subject_name)
                carrera = careers_by_name.get(canonical_career)
                subject_matches = subjects_by_lookup.get((canonical_career, nivel, canonical_subject), [])
                subject = subject_matches[0] if subject_matches else None
                docente = teacher_by_name.get(_normalize_text(teacher_name))
                if not docente:
                    docente = teacher_by_cedula.get(_normalize_cedula(row[22]))
                if not carrera or not subject or not docente:
                    skip_count += 1
                    continue
                if carrera.es_actividad or subject.es_actividad:
                    continue
                paralelo = (_clean_text(row[4]) or "A")[:3].upper()
                key = (subject.id_asignatura, carrera.id_carrera, periodo.id_periodo, paralelo)
                required_fields = subject_field_by_id.get(subject.id_asignatura, set())
                teacher_fields = teacher_field_by_id.get(docente.id_docente, set())
                if nivel >= 4 and not (required_fields & teacher_fields):
                    invalid_affinity_keys.add(key)
                    conflicts.append(
                        f"SIN AFINIDAD: {subject.nombre_asignatura} {paralelo} - "
                        f"{docente.nombres_completos} ({source_name})"
                    )
                    skip_count += 1
                    continue
                field_name_raw = _clean_text(row[7])
                field_alias = FIELD_ALIASES.get(_canonical_name(field_name_raw), field_name_raw)
                field = field_by_name.get(_normalize_text(field_alias))
                if not field:
                    field = field_by_code.get(_clean_text(row[7]))
                if not field:
                    field_ids = sorted(subject_field_by_id.get(subject.id_asignatura, set()))
                    field = CatalogoCampoConocimiento.objects.filter(pk=field_ids[0]).first() if field_ids else default_field
                if not field:
                    skip_count += 1
                    continue
                # Una clase se identifica por carrera, asignatura, periodo y
                # paralelo; el docente es un dato reemplazable, no parte de la
                # identidad. Asi un cambio de docente no crea un duplicado.
                previous = assignment_bucket.get(key)
                if previous and previous["docente"].pk != docente.pk:
                    conflicts.append(
                        f"{subject.nombre_asignatura} {paralelo}: "
                        f"{previous['docente'].nombres_completos} -> {docente.nombres_completos} ({source_name})"
                    )
                assignment_bucket[key] = {
                    "docente": docente, "subject": subject, "periodo": periodo,
                    "paralelo": paralelo, "id_carrera": carrera,
                    "id_campo": field, "nivel_semestre_asignado": subject.nivel_semestre,
                    "horas_complementarias": 0, "horas_clase": horas_clase,
                }
        for item in assignment_bucket.values():
            existing = list(PlanificacionAsignacionDocente.objects.filter(
                id_asignatura=item["subject"],
                id_carrera=item["id_carrera"],
                id_periodo=item["periodo"],
                paralelo_asignado__iexact=item["paralelo"],
            ).order_by("id_asignacion"))
            if existing:
                obj = existing[0]
                obj.id_docente = item["docente"]
                obj.id_campo = item["id_campo"]
                obj.nivel_semestre_asignado = item["nivel_semestre_asignado"]
                obj.paralelo_asignado = item["paralelo"]
                obj.horas_clase = item["horas_clase"]
                obj.horas_complementarias = 0
                obj.save()
                updated_assignments += 1
                if len(existing) > 1:
                    duplicate_ids = [obj.pk for obj in existing[1:]]
                    PlanificacionAsignacionDocente.objects.filter(pk__in=duplicate_ids).delete()
                    merged_duplicates += len(duplicate_ids)
            else:
                PlanificacionAsignacionDocente.objects.create(
                    id_docente=item["docente"], id_asignatura=item["subject"],
                    id_carrera=item["id_carrera"], id_periodo=item["periodo"],
                    id_campo=item["id_campo"],
                    nivel_semestre_asignado=item["nivel_semestre_asignado"],
                    paralelo_asignado=item["paralelo"], horas_clase=item["horas_clase"],
                    horas_complementarias=0,
                )
                created_assignments += 1
        quarantined = 0
        for subject_id, career_id, period_id, paralelo in invalid_affinity_keys.difference(assignment_bucket):
            deleted, _ = PlanificacionAsignacionDocente.objects.filter(
                id_asignatura_id=subject_id, id_carrera_id=career_id,
                id_periodo_id=period_id, paralelo_asignado__iexact=paralelo,
            ).delete()
            quarantined += deleted
        # Sanea también asignaciones históricas del período que no estuvieran
        # presentes en los archivos actuales pero incumplan la misma regla.
        existing_level_four = list(PlanificacionAsignacionDocente.objects.filter(
            id_periodo=periodo, nivel_semestre_asignado__gte=4,
        ).select_related("id_docente", "id_asignatura"))
        for assignment in existing_level_four:
            required_fields = subject_field_by_id.get(assignment.id_asignatura_id, set())
            teacher_fields = teacher_field_by_id.get(assignment.id_docente_id, set())
            if required_fields & teacher_fields:
                continue
            conflicts.append(
                f"SIN AFINIDAD HISTORICA: {assignment.id_asignatura.nombre_asignatura} "
                f"{assignment.paralelo_asignado} - {assignment.id_docente.nombres_completos}"
            )
            assignment.delete()
            quarantined += 1
        if sync_assignments:
            valid_keys = set(assignment_bucket.keys())
            for assignment in PlanificacionAsignacionDocente.objects.filter(id_periodo=periodo).only("id_asignacion", "id_docente_id", "id_asignatura_id", "id_periodo_id", "paralelo_asignado"):
                key = (assignment.id_asignatura_id, assignment.id_carrera_id, assignment.id_periodo_id, (assignment.paralelo_asignado or "").strip().upper())
                if key not in valid_keys:
                    assignment.delete()
        total = created_assignments + updated_assignments + skip_count
        detail = "\n".join(conflicts[:100]) or None
        if merged_duplicates:
            detail = (detail + "\n" if detail else "") + f"Duplicados fusionados: {merged_duplicates}"
        if quarantined:
            detail = (detail + "\n" if detail else "") + f"Asignaciones sin afinidad retiradas: {quarantined}"
        _log_carga("ASIGNACIONES_DOCENTES", f"FCACC-PLANIFICACION.xlsx + {len(coor_files)} COOR files", total, created_assignments, updated_assignments, skip_count, detail)
        results["step11"] = {"label": "Asignaciones Docentes", "total": total, "creados": created_assignments, "actualizados": updated_assignments, "omitidos": skip_count, "msg": f"{created_assignments} creadas, {updated_assignments} actualizadas, {skip_count} omitidas; {quarantined} sin afinidad retiradas"}

    # ── STEP 12: MATRIZ F4 ──
    def _step12_actividades(self, main_book, coor_files, teacher_by_name, teacher_by_cedula, periodo, results):
        activity_bucket = {}
        skipped = 0
        for source_name, rows in _planning_sources(main_book, coor_files):
            for row in rows[1:]:
                if not row or len(row) <= 22:
                    continue
                career_code = _clean_text(row[2])
                activity_field = _normalize_text(row[7])
                if career_code not in F4_ACTIVITY_CAREER_CODES and activity_field != "ACTIVIDAD":
                    continue
                teacher_name = _clean_text(row[12])
                activity_name = _clean_text(row[5])
                hours = _positive_int(row[11])
                if (
                    not teacher_name or not activity_name or hours <= 0
                    or _normalize_text(teacher_name) in {"SELECCIONE DOCENTE", "NO DEFINIDO", "#N/A", ""}
                ):
                    skipped += 1
                    continue
                docente = teacher_by_name.get(_normalize_text(teacher_name))
                if not docente:
                    docente = teacher_by_cedula.get(_normalize_cedula(row[22]))
                if not docente:
                    skipped += 1
                    continue
                source_code = _clean_text(row[6]) or activity_name
                code = _fit_code(source_code, "ACT-")
                activity_type = "INVESTIGACION" if "INVESTIG" in _normalize_text(activity_name) else "COMPLEMENTARIA"
                activity_bucket[(docente.pk, code)] = {
                    "docente": docente, "code": code, "name": activity_name,
                    "type": activity_type, "hours": hours, "source": source_name,
                }

        created = 0
        updated = 0
        for item in activity_bucket.values():
            activity, _ = CatalogoActividadComplementaria.objects.update_or_create(
                codigo_actividad=item["code"],
                defaults={
                    "nombre_actividad": item["name"][:150],
                    "tipo_actividad": item["type"],
                    "actividad_activa": True,
                },
            )
            _, was_created = PlanificacionActividadDocente.objects.update_or_create(
                id_docente=item["docente"], id_periodo=periodo, id_actividad=activity,
                defaults={
                    "horas_asignadas": item["hours"],
                    "observaciones": f"Importado desde {item['source']}",
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1
        total = created + updated + skipped
        _log_carga(
            "ACTIVIDADES_DOCENTES",
            f"{main_book.name} + {len(coor_files)} coordinaciones",
            total, created, updated, skipped,
        )
        results["step12"] = {
            "label": "Actividades complementarias", "total": total,
            "creados": created, "actualizados": updated, "omitidos": skipped,
            "msg": f"{created} creadas, {updated} actualizadas, {skipped} omitidas",
        }
