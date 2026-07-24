import re
import unicodedata
from datetime import date
from pathlib import Path
from collections import defaultdict
import warnings

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from catalogos.models import (
    CatalogoCampoConocimiento,
    CatalogoCarrera,
    CatalogoDedicacionHoraria,
    CatalogoGradoAfinidad,
    CatalogoModalidadContratacion,
    CatalogoPeriodoAcademico,
    CatalogoTipoDocente,
)
from curriculo.models import CurriculoAsignatura, CurriculoAsignaturaCampo
from docentes.models import DocenteFcacc, DocenteCampoAfinidad
from planificacion.models import (
    PlanificacionActividadDocente,
    PlanificacionAsignacionDocente,
    PlanificacionCapacidadEspecial,
    PlanificacionDemandaAcademica,
    PlanificacionMatrizF4,
    PlanificacionRepartoHoras,
)


def _normalize_text(value):
    text = str(value or "").replace("\xa0", " ").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def _clean_text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _canonical_name(value):
    text = _normalize_text(value)
    text = re.sub(r"^\d+\s+", "", text)
    text = re.sub(r"\s*\([^)]*\)\s*", " ", text)
    text = text.replace(" - ", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_cedula(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if not digits:
        return None
    if len(digits) > 10:
        digits = digits[:10]
    return digits.zfill(10)


def _positive_int(value):
    try:
        return max(0, int(float(value or 0)))
    except (TypeError, ValueError):
        return 0


def _valid_level(value):
    try:
        level = int(value or 0)
    except (TypeError, ValueError):
        return None
    return level if 0 <= level <= 20 else None


def _parse_parallel_code(value):
    text = _clean_text(value or "")
    match = re.match(r"(\d+)\s*([A-Z])", text)
    if match:
        return f"{match.group(1)}{match.group(2)}"
    return text[:5]


CAREER_NAME_MAP = {
    "MERCADOTECNIA O MARKETING -NS- 2026": "FCACC-MA",
    "MERCADOTECNIA O MARKETING (2024NS)": "FCACC-MA",
    "MERCADOTECNIA (2017)": "FCACC-MA",
    "MERCADOTECNIA": "FCACC-MA",
    "MALLA 2024NS": "FCACC-MA",
    "MALLA 2017": "FCACC-MA",
    "CONTABILIDAD Y AUDITORIA - 2024 NS": "FCACC-CO-AD",
    "CONTABILIDAD Y AUDITORIA 2024 NS": "FCACC-CO-AD",
    "ADMINISTRACION DE EMPRESAS": "FCACC-AD-EM",
    "GESTION DE LA INFORMACION GERENCIAL": "FCACC-GE-IN-GE",
    "COMERCIO EXTERIOR": "FCACC-CO-EX",
}


def _resolve_career(name):
    canonical = _canonical_name(name)
    for pattern, code in CAREER_NAME_MAP.items():
        if _canonical_name(pattern) == canonical:
            return CatalogoCarrera.objects.filter(codigo_carrera=code).first()
    return None


DEDICACION_MAP = {
    "TC": "TC",
    "MT": "MT",
    "TP": "TP",
}


def _resolve_dedicacion(value):
    text = _normalize_text(value)
    code = DEDICACION_MAP.get(text)
    if code:
        return CatalogoDedicacionHoraria.objects.filter(
            codigo_dedicacion=code
        ).first()
    return CatalogoDedicacionHoraria.objects.first()


def _resolve_modalidad(value):
    text = _normalize_text(value)
    if text == "TITULAR":
        return CatalogoModalidadContratacion.objects.filter(
            codigo_modalidad="NOM"
        ).first() or CatalogoModalidadContratacion.objects.first()
    if text in ("CONTRATO", "CONTRATOS OCASIONALES"):
        return CatalogoModalidadContratacion.objects.filter(
            codigo_modalidad="CON"
        ).first() or CatalogoModalidadContratacion.objects.first()
    return CatalogoModalidadContratacion.objects.first()


class Command(BaseCommand):
    help = "Importa datos desde el Excel de Planificacion MKT 2026-2"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default=r"C:\Users\Det-Pc\Downloads\2.- PLANIFICACIÓN MKT 2026-2 V1.xlsx",
            help="Ruta al archivo Excel de planificacion MKT",
        )
        parser.add_argument(
            "--periodo-codigo",
            default="2026-2",
            help="Codigo del periodo academico",
        )
        parser.add_argument(
            "--periodo-nombre",
            default="2026-2",
            help="Nombre del periodo academico",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Solo analiza sin guardar cambios",
        )

    def handle(self, *args, **options):
        warnings.filterwarnings('ignore', message='Data Validation extension is not supported.*')
        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"No se encontro el archivo: {file_path}")

        dry_run = options["dry_run"]
        periodo_codigo = options["periodo_codigo"]

        self.stdout.write(f"Leyendo: {file_path.name}")
        wb = load_workbook(str(file_path), data_only=True)
        sheet_names = wb.sheetnames
        self.stdout.write(f"Hojas disponibles: {', '.join(sheet_names)}")

        stats = defaultdict(int)

        with transaction.atomic():
            periodo, _ = CatalogoPeriodoAcademico.objects.update_or_create(
                codigo_periodo=periodo_codigo,
                defaults={
                    "nombre_periodo": options["periodo_nombre"],
                    "periodo_activo": True,
                    "fecha_inicio_periodo": date(2026, 5, 1),
                    "fecha_fin_periodo": date(2026, 9, 30),
                },
            )
            if not dry_run and periodo.estado_planificacion not in ('BORRADOR', 'EN_REVISION'):
                raise CommandError(
                    f'El periodo {periodo} esta {periodo.get_estado_planificacion_display()} y no admite importaciones.'
                )

            tipo_docente_default = CatalogoTipoDocente.objects.filter(
                codigo_tipo_docente="TIT"
            ).first() or CatalogoTipoDocente.objects.first()
            grado_afinidad_default = CatalogoGradoAfinidad.objects.filter(
                codigo_grado_afinidad="ALTA"
            ).first() or CatalogoGradoAfinidad.objects.order_by("nivel_prioridad").first()
            if not grado_afinidad_default:
                raise CommandError("No existe ningun catalogo de grado de afinidad.")

            careers_by_code = {
                o.codigo_carrera: o for o in CatalogoCarrera.objects.all()
            }
            careers_by_canonical = {
                _canonical_name(o.nombre_carrera): o for o in CatalogoCarrera.objects.all()
            }
            fields_by_name = {
                _normalize_text(o.nombre_campo_conocimiento): o
                for o in CatalogoCampoConocimiento.objects.all()
            }

            teacher_by_cedula = {
                _normalize_cedula(o.cedula_docente): o
                for o in DocenteFcacc.objects.all()
                if _normalize_cedula(o.cedula_docente)
            }

            # ----------------------------------------------------------------
            # 1. DOCENTES 2026-2
            # ----------------------------------------------------------------
            ws_docentes = wb["DOCENTES 2026-2"]
            self._import_docentes(
                ws_docentes, periodo, teacher_by_cedula,
                tipo_docente_default, stats, dry_run,
            )

            # Refresh teacher cache after imports
            teacher_by_cedula = {
                _normalize_cedula(o.cedula_docente): o
                for o in DocenteFcacc.objects.all()
                if _normalize_cedula(o.cedula_docente)
            }

            # ----------------------------------------------------------------
            # 2. DEMANDA 2026-2
            # ----------------------------------------------------------------
            ws_demanda = wb["DEMANDA 2026-2"]
            self._import_demanda(
                ws_demanda, periodo, careers_by_code,
                careers_by_canonical, stats, dry_run,
            )

            # ----------------------------------------------------------------
            # 3. ASIG 2026-1
            # ----------------------------------------------------------------
            ws_asig = wb["ASIG 2026-1"]
            self._import_asignaciones(
                ws_asig, periodo, careers_by_code,
                careers_by_canonical, fields_by_name,
                teacher_by_cedula, grado_afinidad_default,
                stats, dry_run,
            )

            # ----------------------------------------------------------------
            # 4. MATRIZ F4 V1
            # ----------------------------------------------------------------
            ws_f4 = wb["MATRIZ F4 V1"]
            self._import_matriz_f4(
                ws_f4, periodo, careers_by_code,
                fields_by_name, teacher_by_cedula,
                grado_afinidad_default, stats, dry_run,
            )

            # ----------------------------------------------------------------
            # 5. REPARTO HORAS
            # ----------------------------------------------------------------
            ws_reparto = wb["REPARTO HORAS"]
            self._import_reparto_horas(
                ws_reparto, periodo, teacher_by_cedula, stats, dry_run,
            )

            # ----------------------------------------------------------------
            # 6. CAPACIDADES ESPECIALES
            # ----------------------------------------------------------------
            ws_capacidades = wb["CAPACIDADES ESPECIALES"]
            self._import_capacidades(
                ws_capacidades, periodo, careers_by_code, stats, dry_run,
            )

            if dry_run:
                transaction.set_rollback(True)

        wb.close()

        self._print_stats(stats, dry_run, periodo)

    # ------------------------------------------------------------------
    # DOCENTES
    # ------------------------------------------------------------------
    def _import_docentes(self, ws, periodo, teacher_by_cedula,
                         tipo_docente_default, stats, dry_run):
        rows = list(ws.iter_rows(min_row=4, values_only=True))
        for row in rows:
            if not row or not row[2]:
                continue
            cedula = _normalize_cedula(row[2])
            if not cedula:
                stats["docentes_omitidos"] += 1
                continue
            nombre = _clean_text(row[1])
            if not nombre:
                continue
            tipo_text = _clean_text(row[3] or "")
            dedicacion_text = _clean_text(row[4] or "")

            modalidad = _resolve_modalidad(tipo_text)
            dedicacion_obj = _resolve_dedicacion(dedicacion_text)

            if not modalidad:
                stats["docentes_omitidos"] += 1
                continue

            docente, created = DocenteFcacc.objects.update_or_create(
                cedula_docente=cedula,
                defaults={
                    "tipo_documento": "CEDULA",
                    "id_tipo_docente": tipo_docente_default,
                    "id_modalidad": modalidad,
                    "id_dedicacion": dedicacion_obj,
                    "nombres_completos": nombre,
                    "docente_activo": True,
                },
            )
            teacher_by_cedula[cedula] = docente
            if created:
                stats["docentes_creados"] += 1
            else:
                stats["docentes_actualizados"] += 1



        self.stdout.write(f"  Docentes: {stats['docentes_creados']} creados, "
                          f"{stats['docentes_actualizados']} actualizados, "
                          f"{stats['docentes_omitidos']} omitidos")

    # ------------------------------------------------------------------
    # DEMANDA ACADEMICA
    # ------------------------------------------------------------------
    def _import_demanda(self, ws, periodo, careers_by_code,
                        careers_by_canonical, stats, dry_run):
        rows = list(ws.iter_rows(min_row=7, values_only=True))
        subjects_cache = self._build_subject_cache()

        for row in rows:
            if not row or len(row) < 8:
                continue
            carrera_name = _clean_text(row[0] or "")
            asignatura_name = _clean_text(row[1] or "")
            nivel = _valid_level(row[2])
            horas_semanales = _positive_int(row[3])
            proyeccion = _positive_int(row[4])
            num_paralelos = _positive_int(row[5])
            total_horas = _positive_int(row[7])

            if not carrera_name or not asignatura_name or nivel is None:
                continue

            carrera = _resolve_career(carrera_name)
            if not carrera:
                stats["demanda_carreras_no_encontradas"] += 1
                continue

            subject = self._find_or_create_subject(
                carrera, asignatura_name, nivel, horas_semanales,
                subjects_cache, stats, dry_run,
            )
            if not subject:
                continue

            _, created = PlanificacionDemandaAcademica.objects.update_or_create(
                id_asignatura=subject,
                id_carrera=carrera,
                id_periodo=periodo,
                defaults={
                    "proyeccion_estudiantes": proyeccion,
                    "numero_paralelos": max(1, num_paralelos),
                },
            )
            if created:
                stats["demanda_creadas"] += 1
            else:
                stats["demanda_actualizadas"] += 1

        self.stdout.write(f"  Demanda: {stats['demanda_creadas']} creadas, "
                          f"{stats['demanda_actualizadas']} actualizadas")

    # ------------------------------------------------------------------
    # ASIGNACIONES (ASIG 2026-1)
    # ------------------------------------------------------------------
    def _import_asignaciones(self, ws, periodo, careers_by_code,
                             careers_by_canonical, fields_by_name,
                             teacher_by_cedula, grado_afinidad_default,
                             stats, dry_run):
        rows = list(ws.iter_rows(min_row=4, values_only=True))
        subjects_cache = self._build_subject_cache()
        default_field = (
            fields_by_name.get("UNIDAD BASICA")
            or CatalogoCampoConocimiento.objects.order_by("id_campo").first()
        )

        for row in rows:
            if not row or len(row) < 4:
                continue
            nivel_paralelo = _clean_text(row[0] or "")
            asignatura_name = _clean_text(row[1] or "")
            docente_name = _clean_text(row[2] or "")
            horas = _positive_int(row[3])
            comision = _clean_text(row[4] or "")
            carrera_name = _clean_text(row[5] or "")
            dedicacion_text = _clean_text(row[6] or "")
            notas = _clean_text(row[9] or "")

            if not asignatura_name or not docente_name or horas <= 0:
                stats["asig_omitidas"] += 1
                continue
            if not nivel_paralelo:
                stats["asig_omitidas"] += 1
                continue

            nivel = _valid_level(re.sub(r"[A-Za-z]", "", nivel_paralelo))
            paralelo = re.sub(r"\d", "", nivel_paralelo).strip().upper() or "A"

            # Determine carrera
            carrera = _resolve_career(carrera_name) if carrera_name else careers_by_code.get("FCACC-MA")
            if not carrera:
                stats["asig_carreras_no_encontradas"] += 1
                continue

            subject = self._find_or_create_subject(
                carrera, asignatura_name, nivel, horas,
                subjects_cache, stats, dry_run,
            )
            if not subject:
                continue

            # Find teacher by name (ASIG sheet has teacher names in col3)
            names_part = docente_name.split(",")[0].strip().upper().split()
            docente = None
            if names_part:
                last_name = names_part[0]
                docente = DocenteFcacc.objects.filter(
                    nombres_completos__icontains=last_name
                ).first()
            if not docente:
                stats["asig_docentes_no_encontrados"] += 1
                stats[f"asig_no_match_{docente_name[:20]}"] = 1
                continue

            field = default_field
            _, created = PlanificacionAsignacionDocente.objects.update_or_create(
                id_asignatura=subject,
                id_carrera=carrera,
                id_periodo=periodo,
                paralelo_asignado=paralelo[:3],
                defaults={
                    "id_docente": docente,
                    "id_campo": field,
                    "nivel_semestre_asignado": nivel or 0,
                    "horas_clase": horas,
                    "horas_complementarias": 0,
                    "semanas_planificadas": 16,
                    "comision_servicio": comision or None,
                },
            )
            if created:
                stats["asig_creadas"] += 1
            else:
                stats["asig_actualizadas"] += 1

        self.stdout.write(f"  Asignaciones: {stats['asig_creadas']} creadas, "
                          f"{stats['asig_actualizadas']} actualizadas, "
                          f"{stats['asig_omitidas']} omitidas")

    # ------------------------------------------------------------------
    # MATRIZ F4 V1
    # ------------------------------------------------------------------
    def _import_matriz_f4(self, ws, periodo, careers_by_code,
                          fields_by_name, teacher_by_cedula,
                          grado_afinidad_default, stats, dry_run):
        rows = list(ws.iter_rows(min_row=8, values_only=True))
        current_docente = None
        current_cedula = None

        for row in rows:
            if not row:
                continue

            col1 = _clean_text(row[0] or "")
            cedula_val = _normalize_cedula(row[1]) if len(row) > 1 else None

            # New teacher block
            if cedula_val and len(cedula_val) >= 10:
                current_cedula = cedula_val
                current_docente = teacher_by_cedula.get(current_cedula)
                if not current_docente:
                    nombre = _clean_text(row[2] or "") if len(row) > 2 else ""
                    stats["f4_docentes_no_encontrados"] += 1
                    continue
            elif not current_docente:
                continue

            carrera_name = _clean_text(row[9] or "") if len(row) > 9 else ""
            actividad_name = _clean_text(row[10] or "") if len(row) > 10 else ""
            horas = _positive_int(row[12]) if len(row) > 12 else 0
            paralelos_raw = _clean_text(row[13] or "") if len(row) > 13 else ""

            if not carrera_name or not actividad_name:
                continue

            tipo_actividad = "Docencia - Clases"
            if carrera_name.startswith("1."):
                tipo_actividad = "Docencia - Clases"
            elif carrera_name.startswith("2."):
                tipo_actividad = "Investigacion"
            elif carrera_name.startswith("3."):
                tipo_actividad = "Practicas"
            elif carrera_name.startswith("4."):
                tipo_actividad = "Titulacion"
            elif carrera_name.startswith("5."):
                tipo_actividad = "Investigacion"
            elif carrera_name.startswith("6."):
                tipo_actividad = "Investigacion"
            elif carrera_name.startswith("8."):
                tipo_actividad = "Vinculacion"
            elif carrera_name.startswith("9."):
                tipo_actividad = "Gestion"

            carrera = _resolve_career(carrera_name)
            if not carrera and not carrera_name.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "8.", "9.")):
                continue
            if not carrera:
                carrera = careers_by_code.get("FCACC-MA")

            # Extract nivel from actividad name like "COMUNICACIÓN INTEGRAL DE LA MERCADOTECNIA (2A)"
            nivel_match = re.search(r"\((\d+)[A-Z]\)", actividad_name)
            nivel = nivel_match.group(1) if nivel_match else ""

            # Grado de afinidad
            grado_text = _clean_text(row[11] or "") if len(row) > 11 else ""
            grado = grado_afinidad_default
            if "MAESTRIA" in grado_text.upper() or "AFIN" in grado_text.upper():
                alto = CatalogoGradoAfinidad.objects.filter(codigo_grado_afinidad="ALTA").first()
                if alto:
                    grado = alto
            elif "DOCTOR" in grado_text.upper():
                alto = CatalogoGradoAfinidad.objects.filter(codigo_grado_afinidad="ALTA").first()
                if alto:
                    grado = alto

            # Count paralelos (e.g., "2" or "Práctica 2 Paralelo C")
            num_paralelos = 1
            if paralelos_raw and paralelos_raw.replace(".", "").isdigit():
                num_paralelos = max(1, int(float(paralelos_raw)))
            elif "PARALELO" in paralelos_raw.upper():
                # Count parallel mentions like "Práctica 2 Paralelo C"
                num_paralelos = paralelos_raw.upper().count("PARALELO")

            if horas <= 0 and stats.get("f4_con_zero_horas", 0) < 5:
                stats["f4_con_zero_horas"] += 1
                continue
            if horas <= 0:
                continue

            _, created = PlanificacionMatrizF4.objects.update_or_create(
                id_docente=current_docente,
                id_carrera=carrera,
                id_periodo=periodo,
                tipo_actividad=tipo_actividad[:100],
                nombre_asignatura_actividad=actividad_name[:200],
                nivel_semestre_actividad=nivel[:20],
                defaults={
                    "id_grado_afinidad": grado,
                    "horas_actividad": horas,
                    "numero_paralelos_actividad": num_paralelos,
                    "observaciones": f"Importado desde MATRIZ F4 V1 - {carrera_name}",
                },
            )
            if created:
                stats["f4_creadas"] += 1
            else:
                stats["f4_actualizadas"] += 1

        self.stdout.write(f"  Matriz F4: {stats['f4_creadas']} creadas, "
                          f"{stats['f4_actualizadas']} actualizadas")

    # ------------------------------------------------------------------
    # REPARTO HORAS
    # ------------------------------------------------------------------
    def _import_reparto_horas(self, ws, periodo, teacher_by_cedula,
                              stats, dry_run):
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        subjects_cache = self._build_subject_cache()

        current_docente = None
        current_subject = None

        for row in rows:
            if not row or not row[0]:
                continue
            col1 = _clean_text(row[0] or "")
            col2 = row[1]

            # Skip total row
            if "TOTAL" in col1.upper():
                continue

            # Check if it's a teacher row
            docente = DocenteFcacc.objects.filter(
                nombres_completos__icontains=col1
            ).first()
            if docente:
                current_docente = docente
                current_subject = None
                continue

            if not current_docente:
                continue

            horas = _positive_int(col2)
            if horas <= 0:
                continue

            # Check if this is nivel-paralelo like "2A"
            nivel_paralelo_match = re.match(r"(\d+)([A-Z])", col1)
            if nivel_paralelo_match:
                # It's a parallel line
                if current_subject:
                    nivel_code = f"{nivel_paralelo_match.group(1)}{nivel_paralelo_match.group(2)}"
                    _, created = PlanificacionRepartoHoras.objects.update_or_create(
                        id_docente=current_docente,
                        id_asignatura=current_subject,
                        id_periodo=periodo,
                        nivel_paralelo=nivel_code[:5],
                        defaults={
                            "horas_presenciales_asignadas": horas,
                        },
                    )
                    if created:
                        stats["reparto_creadas"] += 1
                    else:
                        stats["reparto_actualizadas"] += 1
            elif current_subject is None:
                # It's a subject line
                current_subject = self._find_subject_by_name(col1, subjects_cache)

        self.stdout.write(f"  Reparto Horas: {stats['reparto_creadas']} creadas, "
                          f"{stats['reparto_actualizadas']} actualizadas")

    def _find_subject_by_name(self, name, subjects_cache):
        canonical = _canonical_name(name)
        for key, subj in subjects_cache.items():
            if _canonical_name(key[2]) == canonical:
                return subj
        return None

    # ------------------------------------------------------------------
    # CAPACIDADES ESPECIALES
    # ------------------------------------------------------------------
    def _import_capacidades(self, ws, periodo, careers_by_code, stats, dry_run):
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        carrera = careers_by_code.get("FCACC-MA")

        for row in rows:
            if not row or len(row) < 3:
                continue
            estudiante = _clean_text(row[2] or "")
            condicion = _clean_text(row[4] or "")
            informes = _clean_text(row[5] or "")
            nivel = _clean_text(row[6] or "")
            paralelo = _clean_text(row[7] or "")

            if not estudiante or not carrera:
                continue

            _, created = PlanificacionCapacidadEspecial.objects.update_or_create(
                id_periodo=periodo,
                id_carrera=carrera,
                estudiante_nombre=estudiante,
                defaults={
                    "condicion": condicion or None,
                    "informes_adjuntos": informes or None,
                    "nivel_asignado": nivel or None,
                    "paralelo_asignado": paralelo or None,
                },
            )
            if created:
                stats["capacidades_creadas"] += 1
            else:
                stats["capacidades_actualizadas"] += 1

        self.stdout.write(f"  Capacidades Especiales: {stats['capacidades_creadas']} creadas, "
                          f"{stats['capacidades_actualizadas']} actualizadas")

    # ------------------------------------------------------------------
    # HELPERS
    # ------------------------------------------------------------------
    def _build_subject_cache(self):
        cache = {}
        for obj in CurriculoAsignatura.objects.select_related("id_carrera"):
            key = (
                _canonical_name(obj.id_carrera.nombre_carrera),
                int(obj.nivel_semestre or 0),
                _canonical_name(obj.nombre_asignatura),
            )
            if key not in cache:
                cache[key] = obj
        return cache

    def _find_or_create_subject(self, carrera, name, nivel, horas,
                                subjects_cache, stats, dry_run):
        canonical_career = _canonical_name(carrera.nombre_carrera)
        canonical_subject = _canonical_name(name)
        key = (canonical_career, nivel or 0, canonical_subject)

        if key in subjects_cache:
            return subjects_cache[key]

        # Try without nivel
        for (cc, nl, cn), obj in subjects_cache.items():
            if cc == canonical_career and cn == canonical_subject:
                return obj

        # Fuzzy match: try partial name match
        for (cc, nl, cn), obj in subjects_cache.items():
            if cc == canonical_career:
                if canonical_subject in cn or cn in canonical_subject:
                    return obj

        # Create subject if not found
        if dry_run:
            stats["asignaturas_(por_crear)"] += 1
            return None

        code = re.sub(r"[^A-Z0-9]", "", canonical_subject)[:12]
        compact_code = f"MKT-{code or 'ASG'}-{nivel or 0}"
        compact_code = compact_code[:20]

        subject, created = CurriculoAsignatura.objects.get_or_create(
            codigo_asignatura=compact_code,
            defaults={
                "id_carrera": carrera,
                "nombre_asignatura": _clean_text(name),
                "horas_semanales_asignatura": horas,
                "nivel_semestre": nivel or 0,
            },
        )
        if created:
            stats["asignaturas_creadas"] += 1
            subjects_cache[key] = subject

        return subject

    def _print_stats(self, stats, dry_run, periodo):
        self.stdout.write("\n" + ("=" * 50))
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN - Sin cambios guardados"))
        else:
            self.stdout.write(self.style.SUCCESS("Importacion completada"))
        self.stdout.write(f"Periodo: {periodo.codigo_periodo} - {periodo.nombre_periodo}")
        self.stdout.write(f"  Docentes creados: {stats.get('docentes_creados', 0)}")
        self.stdout.write(f"  Docentes actualizados: {stats.get('docentes_actualizados', 0)}")
        self.stdout.write(f"  Titulos creados: {stats.get('titulos_creados', 0)}")
        self.stdout.write(f"  Asignaturas creadas: {stats.get('asignaturas_creadas', 0)}")
        self.stdout.write(f"  Demanda creadas: {stats.get('demanda_creadas', 0)}")
        self.stdout.write(f"  Demanda actualizadas: {stats.get('demanda_actualizadas', 0)}")
        self.stdout.write(f"  Asignaciones creadas: {stats.get('asig_creadas', 0)}")
        self.stdout.write(f"  Asignaciones actualizadas: {stats.get('asig_actualizadas', 0)}")
        self.stdout.write(f"  Matriz F4 creadas: {stats.get('f4_creadas', 0)}")
        self.stdout.write(f"  Matriz F4 actualizadas: {stats.get('f4_actualizadas', 0)}")
        self.stdout.write(f"  Reparto Horas creadas: {stats.get('reparto_creadas', 0)}")
        self.stdout.write(f"  Reparto Horas actualizadas: {stats.get('reparto_actualizadas', 0)}")
        self.stdout.write(f"  Capacidades creadas: {stats.get('capacidades_creadas', 0)}")
        self.stdout.write(f"  Capacidades actualizadas: {stats.get('capacidades_actualizadas', 0)}")
