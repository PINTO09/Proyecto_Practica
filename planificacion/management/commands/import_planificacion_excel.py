from collections import defaultdict
from datetime import date
import hashlib
from pathlib import Path
import re
import unicodedata
import warnings

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from catalogos.models import (
    CatalogoCampoConocimiento,
    CatalogoCarrera,
    CatalogoDedicacionHoraria,
    CatalogoGradoAfinidad,
    CatalogoModalidadContratacion,
    CatalogoPeriodoAcademico,
    CatalogoTipoDocente,
)
from curriculo.models import CurriculoAsignatura, CurriculoAsignaturaCampo
from docentes.models import DocenteCampoAfinidad, DocenteFcacc
from planificacion.models import (
    PlanificacionAsignacionDocente,
    PlanificacionDemandaAcademica,
    PlanificacionMatrizF4,
)
from planificacion.services import docente_tiene_afinidad


def _normalize_text(value):
    text = str(value or "").replace("\xa0", " ").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def _clean_text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _fit_code(value, prefix, max_length=20):
    code = _clean_text(value)
    if len(code) <= max_length:
        return code
    digest = hashlib.sha1(code.encode("utf-8")).hexdigest().upper()[: max_length - len(prefix)]
    return f"{prefix}{digest}"


def _sheet_rows(path, sheet_name):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    return rows


def _column_indexes(rows, *columns):
    """Resuelve columnas por encabezado y tolera cambios de posición en el Excel."""
    if not rows:
        raise CommandError('La hoja no contiene encabezados.')
    header = {_normalize_text(value): index for index, value in enumerate(rows[0]) if value is not None}
    result = []
    for column in columns:
        aliases = column if isinstance(column, (tuple, list)) else (column,)
        index = next((header.get(_normalize_text(alias)) for alias in aliases if _normalize_text(alias) in header), None)
        if index is None:
            raise CommandError(f'Falta la columna requerida: {" / ".join(aliases)}')
        result.append(index)
    return result


def _valid_level(value):
    try:
        level = int(value or 0)
    except (TypeError, ValueError):
        return None
    return level if 1 <= level <= 8 else None


def _normalize_cedula(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if not digits:
        return None
    if len(digits) > 10:
        digits = digits[:10]
    return digits.zfill(10)


def _normalize_phone(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if not digits:
        return None
    if len(digits) == 9:
        digits = "0" + digits
    return digits[:15]


def _positive_int(value):
    try:
        number = int(float(value or 0))
    except (TypeError, ValueError):
        return 0
    return number if number > 0 else 0


def _canonical_name(value):
    text = _normalize_text(value)
    text = re.sub(r"^\d+\s+", "", text)
    text = re.sub(r"\s*\([^)]*\)\s*", " ", text)
    text = text.replace(" - ", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


CAREER_ALIASES = {
    "CONTABILIDAD Y ADITORIA 2024": "CONTABILIDAD Y AUDITORIA 2024 NS",
    "AUDITORIA Y CONTROL DE GESTION 2024": "AUDITORIA Y CONTROL DE GESTION 2024 NS",
}


FIELD_ALIASES = {
    "ADMINISTRACION": "ADMINISTRACION",
    "COMPUTACION": "CIENCIAS COMPUTACIONALES",
    "CONTABILIDAD": "CONTABILIDAD Y AUDITORIA",
    "DERECHO": "DERECHO",
    "ESTADISTICA": "ECONOMIA MATEMATICA",
    "INFORMATICA": "CIENCIAS COMPUTACIONALES",
    "MATEMATICA": "ECONOMIA MATEMATICA",
    "SOCIALES": "UNIDAD BASICA",
}


F4_ACTIVITY_CAREER_CODES = {
    "FCACC-1-DO-CL",
    "FCACC-4-DO-TU-TI",
    "FCACC-5-DO-PE-IN",
    "FCACC-6-IVN",
    "FCACC-8-VI-SO",
    "FCACC-9-GE_ED",
}


class Command(BaseCommand):
    help = "Importa datos base de planificacion desde los Excel en _excel_input"

    def add_arguments(self, parser):
        parser.add_argument(
            "--base-dir",
            default=None,
            help="Directorio base con los archivos extraidos de Excel. Por defecto usa _excel_input del proyecto.",
        )
        parser.add_argument(
            "--periodo-codigo",
            default="2026-2",
            help="Codigo del periodo academico a crear/usar para la demanda importada.",
        )
        parser.add_argument(
            "--periodo-nombre",
            default="2026-2",
            help="Nombre del periodo academico a crear/usar para la demanda importada.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Analiza y valida sin guardar cambios.",
        )
        parser.add_argument(
            "--sync-asignaciones",
            action="store_true",
            help="Elimina asignaciones del periodo que no aparecen en la carga actual de Excel.",
        )

    def handle(self, *args, **options):
        warnings.filterwarnings('ignore', message='Data Validation extension is not supported.*')
        warnings.filterwarnings('ignore', message='DateTimeField DocenteFcacc.fecha_creacion_registro received a naive datetime.*')
        base_dir = Path(options["base_dir"] or Path(__file__).resolve().parents[3] / "_excel_input")
        dry_run = options["dry_run"]
        sync_assignments = options["sync_asignaciones"]

        main_book = base_dir / "OneDrive_1" / "FCACC-PLANIFICACION.xlsx"
        docente_book = base_dir / "OneDrive_2" / "REVISIONDocentes.xlsx"

        if not main_book.exists():
            raise CommandError(f"No se encontro el archivo principal: {main_book}")
        if not docente_book.exists():
            raise CommandError(f"No se encontro el archivo de revision docente: {docente_book}")

        self.stdout.write(f"Leyendo datos desde: {base_dir}")

        career_rows = _sheet_rows(main_book, "MAE_CARRERA")
        subject_rows = _sheet_rows(main_book, "MAE_ASIGNATURA")
        field_rows = _sheet_rows(main_book, "MAE_CONOCIMIENTO")
        subject_detail_rows = _sheet_rows(main_book, "DET_ASIG")
        teacher_rows = _sheet_rows(main_book, "MDOCENTES")
        teacher_field_rows = _sheet_rows(docente_book, "DET_DOCENTE")
        source_rows = {
            "FCACC-PLANIFICACION.xlsx": _sheet_rows(main_book, "ASIGNACION"),
            "PLANIFICACION_ADMINISTRACION.xlsx": _sheet_rows(base_dir / "OneDrive_1" / "PLANIFICACION_ADMINISTRACION.xlsx", "ASIGNACION"),
            "PLANIFICACION_COMERCIO.xlsx": _sheet_rows(base_dir / "OneDrive_1" / "PLANIFICACION_COMERCIO.xlsx", "ASIGNACION"),
            "PLANIFICACION_CONTABILIDAD.xlsx": _sheet_rows(base_dir / "OneDrive_1" / "PLANIFICACION_CONTABILIDAD.xlsx", "ASIGNACION"),
        }
        demand_sources = []
        assignment_sources = []
        for source_name, rows in source_rows.items():
            total_hours_name = 'HORAS' if source_name == 'FCACC-PLANIFICACION.xlsx' else 'HORAS_SEMANAS'
            demand_sources.append((
                source_name, rows,
                *_column_indexes(rows, 'CARRERA', 'NIVEL', 'PARALELO', 'ASIGNATURA', total_hours_name, 'TOTAL'),
            ))
            assignment_sources.append((
                source_name, rows,
                *_column_indexes(
                    rows, 'CARRERA', 'NIVEL', 'PARALELO', 'ASIGNATURA', 'CAMPO',
                    'HORAS', 'NOMBRE_DOCENT', 'CEDULA_DOCENTE',
                ),
            ))

        tipo_docente_default = CatalogoTipoDocente.objects.filter(codigo_tipo_docente="TITULAR").first()
        if not tipo_docente_default:
            raise CommandError("No existe el catalogo de tipo docente TITULAR.")
        grado_afinidad_default = CatalogoGradoAfinidad.objects.filter(codigo_grado_afinidad="MEDIA").first()
        if not grado_afinidad_default:
            grado_afinidad_default = CatalogoGradoAfinidad.objects.order_by("nivel_prioridad").first()
        if not grado_afinidad_default:
            raise CommandError("No existe ningun catalogo de grado de afinidad para crear registros F4.")

        modalidad_map = {}
        for obj in CatalogoModalidadContratacion.objects.all():
            modalidad_map[_normalize_text(obj.codigo_modalidad)] = obj
            modalidad_map[_normalize_text(obj.nombre_modalidad)] = obj
        dedicacion_map = {}
        for obj in CatalogoDedicacionHoraria.objects.all():
            dedicacion_map[_normalize_text(obj.codigo_dedicacion)] = obj
            dedicacion_map[_normalize_text(obj.nombre_dedicacion)] = obj
        modalidad_aliases = {
            'CONTRATOS OCASIONALES': 'CONTRATO',
            'CONTRATO OCASIONAL': 'CONTRATO',
            'NOMBRAMIENTO PROVISIONAL': 'NOMBRAMIENTO',
            'NOMBRAMIENTO AUTORIDAD UNIVERS': 'NOMBRAMIENTO',
        }

        field_name_to_code = {}
        imported_careers = 0
        imported_fields = 0
        imported_subjects = 0
        imported_subject_fields = 0
        imported_teachers = 0
        imported_teacher_fields = 0
        imported_demands = 0
        created_assignments = 0
        updated_assignments = 0
        skipped_assignments = 0
        created_f4 = 0
        updated_f4 = 0
        skipped_f4 = 0
        with transaction.atomic():
            periodo, _ = CatalogoPeriodoAcademico.objects.update_or_create(
                codigo_periodo=options["periodo_codigo"],
                defaults={
                    "nombre_periodo": options["periodo_nombre"],
                    "periodo_activo": True,
                    "fecha_inicio_periodo": date(2026, 5, 1),
                    "fecha_fin_periodo": date(2026, 9, 30),
                },
            )
            if not dry_run and periodo.estado_planificacion not in ('BORRADOR', 'EN_REVISION'):
                raise CommandError(
                    f'El periodo {periodo} está {periodo.get_estado_planificacion_display()} y no admite importaciones.'
                )

            for row in career_rows[1:]:
                if not row or not row[0] or not row[1]:
                    continue
                _, created = CatalogoCarrera.objects.update_or_create(
                    codigo_carrera=_clean_text(row[0]),
                    defaults={
                        "nombre_carrera": _clean_text(row[1]),
                        "carrera_activa": _normalize_text(row[6]) != "NO VIGENTE",
                    },
                )
                imported_careers += 1 if created else 0

            career_by_code = {
                obj.codigo_carrera: obj
                for obj in CatalogoCarrera.objects.all()
            }
            careers_by_name = {
                _canonical_name(obj.nombre_carrera): obj
                for obj in CatalogoCarrera.objects.all()
            }

            for row in field_rows[1:]:
                if not row or not row[0] or not row[1]:
                    continue
                codigo = _clean_text(row[0])
                nombre = _clean_text(row[1])
                _, created = CatalogoCampoConocimiento.objects.update_or_create(
                    codigo_campo=codigo,
                    defaults={"nombre_campo_conocimiento": nombre},
                )
                field_name_to_code[_normalize_text(nombre)] = codigo
                imported_fields += 1 if created else 0

            field_by_code = {
                obj.codigo_campo: obj
                for obj in CatalogoCampoConocimiento.objects.all()
            }

            subject_code_map = {}

            for row in subject_rows[1:]:
                if not row or not row[0] or not row[1] or not row[4]:
                    continue
                carrera = career_by_code.get(_clean_text(row[4]))
                if not carrera:
                    continue
                level = _valid_level(row[6])
                if level is None:
                    continue
                source_code = _clean_text(row[0])
                compact_code = _fit_code(source_code, "ASG-")
                subject_code_map[source_code] = compact_code
                _, created = CurriculoAsignatura.objects.update_or_create(
                    codigo_asignatura=compact_code,
                    defaults={
                        "id_carrera": carrera,
                        "nombre_asignatura": _clean_text(row[1]),
                        "horas_semanales_asignatura": int(row[2] or 0),
                        "nivel_semestre": level,
                    },
                )
                imported_subjects += 1 if created else 0

            subject_by_code = {
                obj.codigo_asignatura: obj
                for obj in CurriculoAsignatura.objects.select_related("id_carrera")
            }
            subjects_by_lookup = defaultdict(list)
            for obj in subject_by_code.values():
                key = (
                    _canonical_name(obj.id_carrera.nombre_carrera),
                    int(obj.nivel_semestre or 0),
                    _canonical_name(obj.nombre_asignatura),
                )
                subjects_by_lookup[key].append(obj)

            field_code_by_detail = {}
            for row in subject_detail_rows[1:]:
                if not row or not row[5] or not row[7]:
                    continue
                source_subject_code = _clean_text(row[5])
                compact_subject_code = subject_code_map.get(source_subject_code, _fit_code(source_subject_code, "ASG-"))
                field_code_by_detail[compact_subject_code] = _clean_text(row[7])

            for subject_code, field_code in field_code_by_detail.items():
                subject = subject_by_code.get(subject_code)
                field = field_by_code.get(field_code)
                if not subject or not field:
                    continue
                _, created = CurriculoAsignaturaCampo.objects.get_or_create(
                    id_asignatura=subject,
                    id_campo=field,
                )
                imported_subject_fields += 1 if created else 0

            for row in teacher_rows[1:]:
                if not row or not row[0] or not row[1]:
                    continue
                cedula = _normalize_cedula(row[0])
                if not cedula:
                    continue
                modalidad_key = _normalize_text(row[7])
                modalidad = (
                    modalidad_map.get(modalidad_key) or
                    modalidad_map.get(modalidad_aliases.get(modalidad_key, ''))
                )
                if not modalidad:
                    continue
                dedicacion = dedicacion_map.get(_normalize_text(row[6])) or CatalogoDedicacionHoraria.objects.first()
                tipo_docente = tipo_docente_default
                _, created = DocenteFcacc.objects.update_or_create(
                    cedula_docente=cedula,
                    defaults={
                        "tipo_documento": "CEDULA",
                        "id_tipo_docente": tipo_docente,
                        "id_modalidad": modalidad,
                        "id_dedicacion": dedicacion,
                        "nombres_completos": _clean_text(row[1]),
                        "unidad_organica": _clean_text(row[2]),
                        "correo_institucional": _clean_text(row[3]) or None,
                        "numero_celular": _normalize_phone(row[4]),
                        "tipo_sangre": _clean_text(row[8]) or None,
                        "docente_activo": True,
                    },
                )
                imported_teachers += 1 if created else 0

            teacher_by_name = {
                _normalize_text(obj.nombres_completos): obj
                for obj in DocenteFcacc.objects.all()
            }
            teacher_by_cedula = {
                _normalize_cedula(obj.cedula_docente): obj
                for obj in DocenteFcacc.objects.all()
                if _normalize_cedula(obj.cedula_docente)
            }
            field_by_name = {
                _normalize_text(obj.nombre_campo_conocimiento): obj
                for obj in CatalogoCampoConocimiento.objects.all()
            }

            for row in teacher_field_rows[1:]:
                if not row or not row[1] or not row[2]:
                    continue
                teacher = teacher_by_name.get(_normalize_text(row[1]))
                alias_name = FIELD_ALIASES.get(_canonical_name(row[2]), _clean_text(row[2]))
                field = field_by_name.get(_normalize_text(alias_name))
                if not teacher or not field:
                    continue
                _, created = DocenteCampoAfinidad.objects.get_or_create(
                    id_docente=teacher,
                    id_campo=field,
                )
                imported_teacher_fields += 1 if created else 0

            demand_bucket = {}
            for _, rows, carrera_idx, nivel_idx, paralelo_idx, subject_idx, total_hours_idx, total_idx in demand_sources:
                for row in rows[1:]:
                    if not row or len(row) <= subject_idx:
                        continue
                    carrera_name = _clean_text(row[carrera_idx])
                    nivel = _valid_level(row[nivel_idx]) if len(row) > nivel_idx else None
                    paralelo = _clean_text(row[paralelo_idx]) if len(row) > paralelo_idx else ""
                    subject_name = _clean_text(row[subject_idx])
                    total_hours = row[total_hours_idx] if len(row) > total_hours_idx else None
                    total_value = row[total_idx] if len(row) > total_idx else None
                    if not carrera_name or not subject_name or nivel is None:
                        continue
                    if not paralelo and not total_hours and not total_value:
                        continue

                    canonical_career = _canonical_name(carrera_name)
                    canonical_career = CAREER_ALIASES.get(canonical_career, canonical_career)
                    canonical_subject = _canonical_name(subject_name)

                    carrera = careers_by_name.get(canonical_career)
                    subject_matches = subjects_by_lookup.get((canonical_career, nivel, canonical_subject), [])
                    subject = subject_matches[0] if subject_matches else None
                    if not carrera or not subject:
                        continue

                    key = (carrera.id_carrera, subject.id_asignatura)
                    item = demand_bucket.setdefault(
                        key,
                        {"carrera": carrera, "subject": subject, "parallels": set(), "blank_rows": 0}
                    )
                    if paralelo:
                        item["parallels"].add(paralelo)
                    else:
                        item["blank_rows"] += 1

            for item in demand_bucket.values():
                numero_paralelos = max(1, len(item["parallels"]) + item["blank_rows"])
                _, created = PlanificacionDemandaAcademica.objects.update_or_create(
                    id_asignatura=item["subject"],
                    id_carrera=item["carrera"],
                    id_periodo=periodo,
                    defaults={
                        "proyeccion_estudiantes": 0,
                        "numero_paralelos": numero_paralelos,
                    },
                )
                imported_demands += 1 if created else 0

            subject_field_by_id = {}
            for relation in CurriculoAsignaturaCampo.objects.select_related("id_campo", "id_asignatura"):
                subject_field_by_id.setdefault(relation.id_asignatura_id, relation.id_campo)
            default_field = (
                field_by_name.get("UNIDAD BASICA")
                or CatalogoCampoConocimiento.objects.order_by("id_campo").first()
            )

            assignment_bucket = {}
            for (
                source_name,
                rows,
                carrera_idx,
                nivel_idx,
                paralelo_idx,
                subject_idx,
                field_idx,
                hours_idx,
                teacher_idx,
                cedula_idx,
            ) in assignment_sources:
                for row in rows[1:]:
                    if not row or len(row) <= max(carrera_idx, nivel_idx, subject_idx, teacher_idx):
                        skipped_assignments += 1
                        continue

                    carrera_name = _clean_text(row[carrera_idx])
                    nivel = _valid_level(row[nivel_idx]) if len(row) > nivel_idx else None
                    subject_name = _clean_text(row[subject_idx])
                    teacher_name = _clean_text(row[teacher_idx])
                    horas_clase = _positive_int(row[hours_idx]) if len(row) > hours_idx else 0
                    if source_name != "FCACC-PLANIFICACION.xlsx" and (len(row) <= 9 or _positive_int(row[9]) <= 0):
                        skipped_assignments += 1
                        continue

                    if not carrera_name or not subject_name or not teacher_name or nivel is None or horas_clase <= 0:
                        skipped_assignments += 1
                        continue
                    if _normalize_text(teacher_name) in {"SELECCIONE DOCENTE", "NO DEFINIDO", "#N/A"}:
                        skipped_assignments += 1
                        continue

                    canonical_career = _canonical_name(carrera_name)
                    canonical_career = CAREER_ALIASES.get(canonical_career, canonical_career)
                    canonical_subject = _canonical_name(subject_name)

                    carrera = careers_by_name.get(canonical_career)
                    subject_matches = subjects_by_lookup.get((canonical_career, nivel, canonical_subject), [])
                    subject = subject_matches[0] if subject_matches else None
                    docente = teacher_by_name.get(_normalize_text(teacher_name))
                    if not docente and len(row) > cedula_idx:
                        docente = teacher_by_cedula.get(_normalize_cedula(row[cedula_idx]))
                    if not carrera or not subject or not docente:
                        skipped_assignments += 1
                        continue

                    field_name = _clean_text(row[field_idx]) if len(row) > field_idx else ""
                    field_alias = FIELD_ALIASES.get(_canonical_name(field_name), field_name)
                    field = (
                        field_by_name.get(_normalize_text(field_alias))
                        or subject_field_by_id.get(subject.id_asignatura)
                        or default_field
                    )
                    if not field:
                        skipped_assignments += 1
                        continue
                    if nivel >= 4 and not docente_tiene_afinidad(docente, subject):
                        skipped_assignments += 1
                        continue

                    paralelo = _clean_text(row[paralelo_idx]) if len(row) > paralelo_idx else ""
                    paralelo = (paralelo or "A")[:3]
                    key = (docente.id_docente, subject.id_asignatura, periodo.id_periodo, paralelo)
                    item = assignment_bucket.setdefault(
                        key,
                        {
                            "docente": docente,
                            "subject": subject,
                            "periodo": periodo,
                            "paralelo": paralelo,
                            "id_carrera": carrera,
                            "id_campo": field,
                            "nivel_semestre_asignado": subject.nivel_semestre,
                            "horas_complementarias": 0,
                        },
                    )
                    item["horas_clase"] = item.get("horas_clase", 0) + horas_clase

            for item in assignment_bucket.values():
                _, created = PlanificacionAsignacionDocente.objects.update_or_create(
                    id_asignatura=item["subject"],
                    id_carrera=item["id_carrera"],
                    id_periodo=item["periodo"],
                    paralelo_asignado=item["paralelo"],
                    defaults={
                        "id_docente": item["docente"],
                        "id_campo": item["id_campo"],
                        "nivel_semestre_asignado": item["nivel_semestre_asignado"],
                        "horas_clase": item["horas_clase"],
                        "horas_complementarias": item["horas_complementarias"],
                        "semanas_planificadas": 16,
                    },
                )
                if created:
                    created_assignments += 1
                else:
                    updated_assignments += 1

            if sync_assignments:
                valid_keys = set(assignment_bucket.keys())
                for assignment in PlanificacionAsignacionDocente.objects.filter(id_periodo=periodo).only(
                    "id_asignacion",
                    "id_docente_id",
                    "id_asignatura_id",
                    "id_periodo_id",
                    "paralelo_asignado",
                ):
                    key = (
                        assignment.id_docente_id,
                        assignment.id_asignatura_id,
                        assignment.id_periodo_id,
                        (assignment.paralelo_asignado or "").strip(),
                    )
                    if key not in valid_keys:
                        assignment.delete()

            f4_rows = []
            for row in assignment_sources[0][1][1:]:
                if not row or len(row) <= 12:
                    skipped_f4 += 1
                    continue
                career_code = _clean_text(row[2]) if len(row) > 2 else ""
                teacher_name = _clean_text(row[12])
                activity_name = _clean_text(row[5]) if len(row) > 5 else ""
                activity_field = _normalize_text(row[7]) if len(row) > 7 else ""
                activity_kind = _clean_text(row[1]) if len(row) > 1 else "Actividad"
                hours = _positive_int(row[11]) if len(row) > 11 else 0
                if career_code not in F4_ACTIVITY_CAREER_CODES and activity_field != "ACTIVIDAD":
                    continue
                if not teacher_name or not activity_name or hours <= 0:
                    skipped_f4 += 1
                    continue
                carrera = career_by_code.get(career_code)
                docente = teacher_by_name.get(_normalize_text(teacher_name))
                if not carrera or not docente:
                    skipped_f4 += 1
                    continue
                f4_rows.append({
                    "docente": docente,
                    "carrera": carrera,
                    "tipo_actividad": activity_kind,
                    "nombre": activity_name,
                    "nivel": _clean_text(row[3]) if len(row) > 3 else "",
                    "horas": hours,
                    "observaciones": "Importado desde FCACC-PLANIFICACION.xlsx",
                })

            for source_name, rows, carrera_idx, nivel_idx, _paralelo_idx, subject_idx, _field_idx, _hours_idx, teacher_idx, cedula_idx in assignment_sources[1:]:
                for row in rows[1:]:
                    if not row or len(row) <= max(carrera_idx, subject_idx, teacher_idx, cedula_idx):
                        skipped_f4 += 1
                        continue
                    teacher_name = _clean_text(row[teacher_idx])
                    subject_name = _clean_text(row[subject_idx])
                    carrera_name = _clean_text(row[carrera_idx])
                    if not teacher_name or not subject_name or not carrera_name:
                        skipped_f4 += 1
                        continue

                    canonical_career = CAREER_ALIASES.get(_canonical_name(carrera_name), _canonical_name(carrera_name))
                    carrera = careers_by_name.get(canonical_career)
                    docente = teacher_by_name.get(_normalize_text(teacher_name))
                    if not docente:
                        docente = teacher_by_cedula.get(_normalize_cedula(row[cedula_idx]))
                    if not carrera or not docente:
                        skipped_f4 += 1
                        continue

                    nivel = _clean_text(row[nivel_idx]) if len(row) > nivel_idx else ""
                    h_adicional_1 = _positive_int(row[10]) if len(row) > 10 else 0
                    h_adicional_2 = _positive_int(row[11]) if len(row) > 11 else 0
                    h_investigacion = _positive_int(row[12]) if len(row) > 12 else 0
                    additions = [
                        ("Actividad adicional 1", h_adicional_1),
                        ("Actividad adicional 2", h_adicional_2),
                        ("Investigacion", h_investigacion),
                    ]
                    added_any = False
                    for tipo_actividad, hours in additions:
                        if hours <= 0:
                            continue
                        added_any = True
                        f4_rows.append({
                            "docente": docente,
                            "carrera": carrera,
                            "tipo_actividad": tipo_actividad,
                            "nombre": subject_name,
                            "nivel": nivel,
                            "horas": hours,
                            "observaciones": f"Importado desde {source_name}",
                        })
                    if not added_any:
                        skipped_f4 += 1

            for item in f4_rows:
                _, created = PlanificacionMatrizF4.objects.update_or_create(
                    id_docente=item["docente"],
                    id_carrera=item["carrera"],
                    id_periodo=periodo,
                    tipo_actividad=item["tipo_actividad"][:100],
                    nombre_asignatura_actividad=item["nombre"][:200],
                    nivel_semestre_actividad=item["nivel"][:20],
                    defaults={
                        "id_grado_afinidad": grado_afinidad_default,
                        "horas_actividad": item["horas"],
                        "numero_paralelos_actividad": 1,
                        "observaciones": item["observaciones"],
                    },
                )
                if created:
                    created_f4 += 1
                else:
                    updated_f4 += 1

            if dry_run:
                transaction.set_rollback(True)

        if dry_run:
            self.stdout.write(self.style.WARNING("Dry run completado sin guardar cambios"))
        else:
            self.stdout.write(self.style.SUCCESS("Importacion completada"))
        self.stdout.write(f"Periodo usado: {periodo.codigo_periodo} - {periodo.nombre_periodo}")
        self.stdout.write(f"Carreras nuevas: {imported_careers}")
        self.stdout.write(f"Campos nuevos: {imported_fields}")
        self.stdout.write(f"Asignaturas nuevas: {imported_subjects}")
        self.stdout.write(f"Relaciones asignatura-campo nuevas: {imported_subject_fields}")
        self.stdout.write(f"Docentes nuevos: {imported_teachers}")
        self.stdout.write(f"Afinidades docente-campo nuevas: {imported_teacher_fields}")
        self.stdout.write(f"Demandas nuevas: {imported_demands}")
        self.stdout.write(f"Asignaciones creadas: {created_assignments}")
        self.stdout.write(f"Asignaciones actualizadas: {updated_assignments}")
        self.stdout.write(f"Asignaciones omitidas: {skipped_assignments}")
        self.stdout.write(f"Registros F4 creados: {created_f4}")
        self.stdout.write(f"Registros F4 actualizados: {updated_f4}")
        self.stdout.write(f"Registros F4 omitidos: {skipped_f4}")
