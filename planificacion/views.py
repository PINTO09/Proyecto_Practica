import json
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Sum, Q, Prefetch, F
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.exceptions import PermissionDenied, ValidationError
from core.crud_base import CrudListView, CrudCreateView, CrudUpdateView, CrudDeleteView
from accounts.decorators import (
    ADMIN, has_role, role_required, ROLES_ESCRITURA,
    module_permission_required, allowed_career_ids,
)
from .forms import PlanificacionAsignacionDocenteForm, PlanificacionActividadDocenteForm, PlanificacionAulaHorarioForm
from .services import (
    assert_periodo_editable, docente_tiene_afinidad, normalize_parallel,
    validate_assignment_business_rules,
)
from .models import (
    CatalogoActividadComplementaria, PlanificacionActividadDocente,
    PlanificacionDemandaAcademica, PlanificacionAsignacionDocente,
    PlanificacionRepartoHoras, PlanificacionMatrizF4, PlanificacionAulaHorario,
)
from docentes.models import DocenteFcacc, DocenteCampoAfinidad, DocenteTituloAcademico
from curriculo.models import CurriculoAsignatura, CurriculoAsignaturaCampo, RelacionPosgradoCampo
from catalogos.models import CatalogoCampoConocimiento, CatalogoCarrera, CatalogoDedicacionHoraria, CatalogoModalidadContratacion, CatalogoPeriodoAcademico, LimiteHorario
import re
import unicodedata
import warnings

from django.conf import settings
from openpyxl import load_workbook


F4_ACTIVITY_CAREER_CODES = {
    'FCACC-1-DO-CL',
    'FCACC-4-DO-TU-TI',
    'FCACC-5-DO-PE-IN',
    'FCACC-6-IVN',
    'FCACC-8-VI-SO',
    'FCACC-9-GE_ED',
    'FCACC-8-PRAC',
    'FCACC-8-TITU',
}


class PeriodEditableDeleteMixin:
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            assert_periodo_editable(self.object.id_periodo)
        except ValidationError as exc:
            messages.error(request, exc.messages[0])
            return redirect(f'{self.model._meta.app_label}:{self.model._meta.model_name}_list')
        return super().post(request, *args, **kwargs)


class AdminOnlyMixin:
    def dispatch(self, request, *args, **kwargs):
        if not has_role(request.user, ADMIN):
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)


def _ensure_career_access(request, carrera_id):
    permitted = allowed_career_ids(request.user)
    if permitted is not None and carrera_id and int(carrera_id) not in permitted:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied
    return permitted


def _scope_careers(queryset, request):
    permitted = allowed_career_ids(request.user)
    if permitted is None:
        return queryset
    return queryset.filter(id_carrera__in=permitted)


def _build_parallel_label(index):
    label = ''
    current = index
    while current >= 0:
        label = chr(65 + (current % 26)) + label
        current = (current // 26) - 1
    return label


def _build_parallel_labels(total):
    return [_build_parallel_label(i) for i in range(max(total, 0))]


def _filter_querystring(request):
    params = request.GET.copy()
    params.pop('page', None)
    encoded = params.urlencode()
    return f'{encoded}&' if encoded else ''


def _paginate_items(request, items, per_page):
    paginator = Paginator(items, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))
    return paginator, page_obj, list(page_obj.object_list)


def _style_form_fields(form):
    for field in form.fields.values():
        widget = field.widget
        input_type = getattr(widget, 'input_type', '')
        if input_type == 'checkbox':
            widget.attrs['class'] = 'form-check-input'
        elif widget.__class__.__name__ == 'Select':
            widget.attrs['class'] = 'form-select'
        else:
            widget.attrs['class'] = 'form-control'
    return form


def _build_form_load_summary(docente=None, periodo=None):
    if not docente or not periodo:
        return None

    workload = _build_docente_workload_map(periodo_id=periodo).get(
        docente.id_docente, _empty_workload()
    )
    limite = _get_limite_horario_docente(docente)
    max_total = ((limite.horas_maximas or 0) + (limite.horas_complementarias_maximas or 0)) if limite else 0
    percentage = round((workload['total_horas'] / max_total) * 100, 1) if max_total > 0 else 0
    return {
        'docente': docente,
        'workload': workload,
        'max_total': max_total,
        'max_clase': limite.horas_maximas if limite else 0,
        'max_comp': limite.horas_complementarias_maximas if limite else 0,
        'available': max(0, max_total - workload['total_horas']),
        'percentage': percentage,
        'badge': 'danger' if percentage > 100 else 'warning' if percentage >= 80 else 'success',
    }


class LenientPaginationMixin:
    def paginate_queryset(self, queryset, page_size):
        paginator = self.get_paginator(
            queryset,
            page_size,
            allow_empty_first_page=self.get_allow_empty(),
        )
        page = self.kwargs.get(self.page_kwarg) or self.request.GET.get(self.page_kwarg) or 1
        page_obj = paginator.get_page(page)
        return paginator, page_obj, page_obj.object_list, page_obj.has_other_pages()


class PlanningFlowContextMixin:
    planning_active_section = ''

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['show_planning_flow'] = True
        context['active_section'] = self.planning_active_section
        return context


def _classify_f4_activity(tipo_actividad):
    activity = (tipo_actividad or '').strip().lower()
    if 'invest' in activity:
        return 'investigacion'
    if 'gesti' in activity:
        return 'gestion'
    if 'vincul' in activity:
        return 'vinculacion'
    return 'actividad'


def _empty_workload():
    return {
        'horas_afinidad': 0,
        'horas_no_afinidad': 0,
        'horas_unidad_basica': 0,
        'horas_clase': 0,
        'total_horas_clase': 0,
        'horas_complementarias': 0,
        'horas_investigacion': 0,
        'horas_gestion': 0,
        'horas_vinculacion': 0,
        'horas_actividad': 0,
        'horas_actividades_total': 0,
        'total_horas': 0,
    }


def _knowledge_field_maps():
    subject_fields = {}
    for subject_id, field_id in CurriculoAsignaturaCampo.objects.values_list(
        'id_asignatura_id', 'id_campo_id'
    ):
        subject_fields.setdefault(subject_id, set()).add(field_id)

    teacher_fields = {}
    for teacher_id, field_id in DocenteCampoAfinidad.objects.values_list(
        'id_docente_id', 'id_campo_id'
    ):
        teacher_fields.setdefault(teacher_id, set()).add(field_id)

    postgraduate_fields = {}
    for postgraduate_id, field_id in RelacionPosgradoCampo.objects.values_list(
        'id_posgrado_id', 'id_campo_id'
    ):
        postgraduate_fields.setdefault(postgraduate_id, set()).add(field_id)
    for teacher_id, postgraduate_id in DocenteTituloAcademico.objects.exclude(
        id_posgrado__isnull=True
    ).values_list('id_docente_id', 'id_posgrado_id'):
        teacher_fields.setdefault(teacher_id, set()).update(
            postgraduate_fields.get(postgraduate_id, set())
        )
    return subject_fields, teacher_fields


def _assignment_hour_category(
    docente_id, asignatura_id, field_name, subject_fields=None, teacher_fields=None
):
    """Replica las tres categorías de horas de la hoja ASIGNACION del Excel."""
    if _excel_normalize_text(field_name) == 'UNIDAD BASICA':
        return 'unidad_basica'
    if subject_fields is None or teacher_fields is None:
        subject_fields, teacher_fields = _knowledge_field_maps()
    if subject_fields.get(asignatura_id, set()) & teacher_fields.get(docente_id, set()):
        return 'afinidad'
    return 'no_afinidad'


def _activity_workload_key(docente_id, periodo_id, name, hours):
    return (
        docente_id,
        periodo_id,
        _excel_normalize_text(name),
        int(hours or 0),
    )


def _registered_activity_keys(periodo_id=None):
    qs = PlanificacionActividadDocente.objects.all()
    if periodo_id:
        qs = qs.filter(id_periodo_id=periodo_id)
    return {
        _activity_workload_key(
            row['id_docente_id'], row['id_periodo_id'],
            row['id_actividad__nombre_actividad'], row['horas_asignadas'],
        )
        for row in qs.values(
            'id_docente_id', 'id_periodo_id',
            'id_actividad__nombre_actividad', 'horas_asignadas',
        )
    }


def _build_docente_workload_map(periodo_id=None, carrera_id=None):
    workload = {}
    subject_fields, teacher_fields = _knowledge_field_maps()

    asignaciones_qs = PlanificacionAsignacionDocente.objects.select_related('id_campo')
    if periodo_id:
        asignaciones_qs = asignaciones_qs.filter(id_periodo_id=periodo_id)
    if carrera_id:
        asignaciones_qs = asignaciones_qs.filter(id_carrera_id=carrera_id)
    for row in asignaciones_qs.values(
        'id_docente_id', 'id_asignatura_id', 'id_campo__nombre_campo_conocimiento',
        'horas_clase', 'horas_complementarias',
    ):
        docente_workload = workload.setdefault(row['id_docente_id'], _empty_workload())
        hours = row['horas_clase'] or 0
        category = _assignment_hour_category(
            row['id_docente_id'], row['id_asignatura_id'],
            row['id_campo__nombre_campo_conocimiento'], subject_fields, teacher_fields,
        )
        docente_workload[f'horas_{category}'] += hours
        docente_workload['horas_clase'] += hours
        docente_workload['horas_complementarias'] += row['horas_complementarias'] or 0

    actividades_qs = PlanificacionActividadDocente.objects.select_related('id_actividad')
    if periodo_id:
        actividades_qs = actividades_qs.filter(id_periodo_id=periodo_id)
    activity_keys = _registered_activity_keys(periodo_id)

    f4_qs = PlanificacionMatrizF4.objects.all()
    if periodo_id:
        f4_qs = f4_qs.filter(id_periodo_id=periodo_id)
    if carrera_id:
        f4_qs = f4_qs.filter(id_carrera_id=carrera_id)
    seen_f4 = set()
    f4_totals = {}
    for row in f4_qs.values(
        'id_docente_id', 'id_periodo_id', 'tipo_actividad',
        'nombre_asignatura_actividad', 'horas_actividad',
        'numero_paralelos_actividad',
    ):
        total = (row['horas_actividad'] or 0) * (row['numero_paralelos_actividad'] or 1)
        activity_key = _activity_workload_key(
            row['id_docente_id'], row['id_periodo_id'],
            row['nombre_asignatura_actividad'], total,
        )
        if activity_key in activity_keys:
            continue
        dedupe_key = (
            row['id_docente_id'], row['id_periodo_id'], row['tipo_actividad'],
            _excel_normalize_text(row['nombre_asignatura_actividad']),
            row['horas_actividad'], row['numero_paralelos_actividad'],
        )
        if not carrera_id and dedupe_key in seen_f4:
            continue
        seen_f4.add(dedupe_key)
        key = (row['id_docente_id'], row['tipo_actividad'])
        f4_totals[key] = f4_totals.get(key, 0) + total
    f4_rows = [
        {'id_docente_id': key[0], 'tipo_actividad': key[1], 'total_horas': total}
        for key, total in f4_totals.items()
    ]

    for row in f4_rows:
        docente_workload = workload.setdefault(row['id_docente_id'], _empty_workload())
        bucket = _classify_f4_activity(row['tipo_actividad'])
        if bucket == 'investigacion':
            docente_workload['horas_investigacion'] += row['total_horas'] or 0
        elif bucket == 'gestion':
            docente_workload['horas_gestion'] += row['total_horas'] or 0
        elif bucket == 'vinculacion':
            docente_workload['horas_vinculacion'] += row['total_horas'] or 0
        else:
            docente_workload['horas_actividad'] += row['total_horas'] or 0

    for row in actividades_qs.values('id_docente_id', 'id_actividad__tipo_actividad').annotate(
        total_horas=Sum('horas_asignadas')
    ):
        docente_workload = workload.setdefault(row['id_docente_id'], _empty_workload())
        activity_type = row['id_actividad__tipo_actividad']
        if activity_type == 'INVESTIGACION':
            docente_workload['horas_investigacion'] += row['total_horas'] or 0
        elif activity_type == 'GESTION':
            docente_workload['horas_gestion'] += row['total_horas'] or 0
        elif activity_type == 'VINCULACION':
            docente_workload['horas_vinculacion'] += row['total_horas'] or 0
        else:
            docente_workload['horas_complementarias'] += row['total_horas'] or 0

    for docente_workload in workload.values():
        docente_workload['total_horas_clase'] = docente_workload['horas_clase']
        docente_workload['horas_actividades_total'] = (
            docente_workload['horas_complementarias'] +
            docente_workload['horas_investigacion'] +
            docente_workload['horas_gestion'] +
            docente_workload['horas_vinculacion'] +
            docente_workload['horas_actividad']
        )
        docente_workload['total_horas'] = (
            docente_workload['horas_clase'] + docente_workload['horas_actividades_total']
        )

    return workload


def _get_limite_horario_docente(docente):
    if not docente or not docente.id_modalidad_id:
        return None
    return LimiteHorario.objects.filter(
        id_modalidad_id=docente.id_modalidad_id,
        activo=True,
    ).first()


def _build_limit_config_state():
    modalidades = list(CatalogoModalidadContratacion.objects.order_by('id_modalidad'))
    activos = list(
        LimiteHorario.objects.filter(activo=True)
        .select_related('id_modalidad')
        .order_by('id_modalidad_id')
    )
    configured_ids = {item.id_modalidad_id for item in activos}
    missing = [m for m in modalidades if m.id_modalidad not in configured_ids]
    return {
        'has_limits': bool(activos),
        'all_configured': not missing and bool(modalidades),
        'configured_count': len(activos),
        'modalidad_count': len(modalidades),
        'missing_dedicaciones': missing,
    }


def _excel_clean_text(value):
    return re.sub(r'\s+', ' ', str(value or '').replace('\xa0', ' ')).strip()


def _excel_normalize_text(value):
    text = _excel_clean_text(value)
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r'\s+', ' ', text).strip().upper()


def _excel_positive_int(value):
    try:
        number = int(float(value or 0))
    except (TypeError, ValueError):
        return 0
    return number if number > 0 else 0


def _safe_header_index(headers, name):
    return headers.index(name) if name in headers else None


def _load_workbook_quiet(path):
    with warnings.catch_warnings():
        warnings.simplefilter('ignore', UserWarning)
        return load_workbook(path, read_only=True, data_only=True)


def _build_excel_global_validation(periodo_id=None):
    base_dir = settings.BASE_DIR / '_excel_input' / 'OneDrive_1'
    sources = [
        base_dir / 'FCACC-PLANIFICACION.xlsx',
        base_dir / 'PLANIFICACION_ADMINISTRACION.xlsx',
        base_dir / 'PLANIFICACION_COMERCIO.xlsx',
        base_dir / 'PLANIFICACION_CONTABILIDAD.xlsx',
    ]
    missing_files = [path.name for path in sources if not path.exists()]
    expected = {}

    for path in sources:
        if not path.exists():
            continue

        workbook = _load_workbook_quiet(path)
        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', UserWarning)
                sheet = workbook['ASIGNACION']
                sheet_rows = list(sheet.iter_rows(values_only=True))
            if not sheet_rows:
                continue
            rows = iter(sheet_rows)
            headers = list(next(rows))

            teacher_idx = _safe_header_index(headers, 'NOMBRE_DOCENT')
            total_idx = _safe_header_index(headers, 'TOTAL_HORAS_DOCENTE')
            if total_idx is None:
                total_idx = _safe_header_index(headers, 'TOTAL HORAS')
            subject_idx = _safe_header_index(headers, 'ASIGNATURA')
            career_code_idx = _safe_header_index(headers, 'ID_CARRERA')
            level_idx = _safe_header_index(headers, 'NIVEL')
            row_total_idx = _safe_header_index(headers, 'TOTAL')
            field_idx = _safe_header_index(headers, 'CAMPO')
            hours_idx = _safe_header_index(headers, 'HORAS')
            weekly_hours_idx = _safe_header_index(headers, 'HORAS_SEMANAS')
            extra_1_idx = _safe_header_index(headers, 'H.ADI_01')
            extra_2_idx = _safe_header_index(headers, 'H.ADI_02')
            research_idx = _safe_header_index(headers, 'H.INVESTIGACION')
            is_main_book = path.name == 'FCACC-PLANIFICACION.xlsx'

            required = [teacher_idx, total_idx, subject_idx, career_code_idx, level_idx, row_total_idx, hours_idx]
            if any(index is None for index in required):
                continue
            max_required_idx = max(index for index in required if index is not None)

            for row in rows:
                if not row or len(row) <= max_required_idx:
                    continue

                teacher_name = _excel_clean_text(row[teacher_idx])
                declared_total = _excel_positive_int(row[total_idx])
                if not teacher_name or declared_total <= 0:
                    continue

                item = expected.setdefault(_excel_normalize_text(teacher_name), {
                    'teacher_name': teacher_name,
                    'declared_total': 0,
                    'detail_class': 0,
                    'detail_f4': 0,
                    'detail_f4_seen': set(),
                    'sources': set(),
                })
                item['declared_total'] = max(item['declared_total'], declared_total)
                item['sources'].add(path.name)

                subject_name = _excel_clean_text(row[subject_idx])
                career_code = _excel_clean_text(row[career_code_idx])
                level = _excel_positive_int(row[level_idx])

                if is_main_book:
                    row_total = _excel_positive_int(row[row_total_idx])
                    field_name = _excel_normalize_text(row[field_idx]) if field_idx is not None and len(row) > field_idx else ''
                    if career_code in F4_ACTIVITY_CAREER_CODES or field_name == 'ACTIVIDAD':
                        f4_key = ('Actividad', subject_name, row_total)
                        if row_total > 0 and f4_key not in item['detail_f4_seen']:
                            item['detail_f4_seen'].add(f4_key)
                            item['detail_f4'] += row_total
                    elif 1 <= level <= 8:
                        item['detail_class'] += row_total
                else:
                    weekly_hours = _excel_positive_int(row[weekly_hours_idx]) if weekly_hours_idx is not None and len(row) > weekly_hours_idx else 0
                    class_hours = _excel_positive_int(row[hours_idx])
                    if weekly_hours > 0 and 1 <= level <= 8:
                        item['detail_class'] += class_hours

                    additions = (
                        ('Actividad adicional 1', extra_1_idx),
                        ('Actividad adicional 2', extra_2_idx),
                        ('Investigacion', research_idx),
                    )
                    for activity_type, index in additions:
                        hours = _excel_positive_int(row[index]) if index is not None and len(row) > index else 0
                        f4_key = (activity_type, subject_name, hours)
                        if hours > 0 and f4_key not in item['detail_f4_seen']:
                            item['detail_f4_seen'].add(f4_key)
                            item['detail_f4'] += hours
        finally:
            workbook.close()

    workload_map = _build_docente_workload_map(periodo_id=periodo_id)
    teachers_by_name = {
        _excel_normalize_text(docente.nombres_completos): docente
        for docente in DocenteFcacc.objects.filter(docente_activo=True).select_related('id_dedicacion')
    }

    rows = []
    for key, item in expected.items():
        docente = teachers_by_name.get(key)
        workload = workload_map.get(docente.id_docente, {}) if docente else {}
        actual_class = workload.get('total_horas_clase') or workload.get('horas_clase') or 0
        actual_f4 = workload.get('horas_actividades_total') or 0
        actual_total = actual_class + actual_f4
        detail_total = item['detail_class'] + item['detail_f4']
        declared_diff = actual_total - item['declared_total']
        detail_diff = actual_total - detail_total

        if not docente:
            status = 'missing'
        elif detail_diff == 0:
            status = 'detail_ok'
        elif declared_diff == 0:
            status = 'declared_ok'
        else:
            status = 'mismatch'

        rows.append({
            'docente': docente,
            'teacher_name': item['teacher_name'],
            'declared_total': item['declared_total'],
            'detail_total': detail_total,
            'detail_class': item['detail_class'],
            'detail_f4': item['detail_f4'],
            'actual_total': actual_total,
            'actual_class': actual_class,
            'actual_f4': actual_f4,
            'declared_diff': declared_diff,
            'detail_diff': detail_diff,
            'sources': ', '.join(sorted(item['sources'])),
            'status': status,
        })

    status_order = {'mismatch': 0, 'missing': 1, 'declared_ok': 2, 'detail_ok': 3}
    rows.sort(key=lambda row: (status_order.get(row['status'], 9), row['teacher_name']))
    summary = {
        'total': len(rows),
        'detail_ok': sum(1 for row in rows if row['status'] == 'detail_ok'),
        'declared_ok': sum(1 for row in rows if row['status'] == 'declared_ok'),
        'mismatch': sum(1 for row in rows if row['status'] == 'mismatch'),
        'missing': sum(1 for row in rows if row['status'] == 'missing'),
        'missing_files': missing_files,
    }
    return rows, summary


def _build_asignacion_limit_snapshot(docente, periodo, horas_clase_nuevas, horas_comp_nuevas, exclude_asignacion_id=None):
    limite = _get_limite_horario_docente(docente)
    if not limite:
        return {
            'error': 'No hay un límite horario activo configurado para la dedicación del docente.',
        }

    horas_previas_qs = PlanificacionAsignacionDocente.objects.filter(
        id_docente=docente,
        id_periodo=periodo,
    )
    if exclude_asignacion_id is not None:
        horas_previas_qs = horas_previas_qs.exclude(id_asignacion=exclude_asignacion_id)

    horas_previas = horas_previas_qs.aggregate(
        total_clase=Sum('horas_clase'),
        total_comp=Sum('horas_complementarias'),
    )

    clase_previa = horas_previas['total_clase'] or 0
    comp_previa = horas_previas['total_comp'] or 0
    clase_nueva = horas_clase_nuevas or 0
    comp_nueva = horas_comp_nuevas or 0

    clase_total = clase_previa + clase_nueva
    comp_total = comp_previa + comp_nueva
    total_asignado = clase_total + comp_total
    max_clase = limite.horas_maximas or 0
    max_comp = limite.horas_complementarias_maximas or 0
    max_total = max_clase + max_comp
    workload_map = _build_docente_workload_map(periodo_id=periodo)
    docente_workload = workload_map.get(docente.id_docente, _empty_workload())
    actividad_complementaria_total = max(
        0, docente_workload['horas_complementarias'] - comp_previa
    )
    otras_actividades_total = (
        actividad_complementaria_total +
        docente_workload['horas_investigacion'] +
        docente_workload['horas_gestion'] +
        docente_workload['horas_vinculacion'] +
        docente_workload['horas_actividad']
    )
    total_general = total_asignado + otras_actividades_total

    return {
        'limite': limite,
        'clase_previa': clase_previa,
        'comp_previa': comp_previa,
        'clase_total': clase_total,
        'comp_total': comp_total,
        'total_asignado': total_asignado,
        'max_clase': max_clase,
        'max_comp': max_comp,
        'max_total': max_total,
        'horas_investigacion': docente_workload['horas_investigacion'],
        'horas_gestion': docente_workload['horas_gestion'],
        'horas_vinculacion': docente_workload['horas_vinculacion'],
        'horas_actividad': docente_workload['horas_actividad'],
        'horas_actividad_complementaria': actividad_complementaria_total,
        'otras_actividades_total': otras_actividades_total,
        'total_general': total_general,
        'horas_clase_disponibles': max(0, max_clase - clase_total),
        'horas_complementarias_disponibles': max(0, max_comp - comp_total),
        'horas_totales_disponibles': max(0, max_total - total_general),
    }


def _add_asignacion_limit_errors(form, docente, snapshot, es_actividad=False):
    if es_actividad:
        return False
    if snapshot.get('error'):
        form.add_error(None, snapshot['error'])
        return True

    if snapshot['clase_total'] > snapshot['max_clase']:
        form.add_error(
            'horas_clase',
            (
                f'Se excede el limite de horas clase para {docente.nombres_completos}. '
                f'Asignadas: {snapshot["clase_total"]} / Límite: {snapshot["max_clase"]}.'
            ),
        )

    if snapshot['comp_total'] > snapshot['max_comp']:
        form.add_error(
            'horas_complementarias',
            (
                f'Se excede el limite de horas complementarias para {docente.nombres_completos}. '
                f'Asignadas: {snapshot["comp_total"]} / Límite: {snapshot["max_comp"]}.'
            ),
        )

    if snapshot['total_general'] > snapshot['max_total']:
        form.add_error(
            None,
            (
                f'El total de horas para {docente.nombres_completos} seria {snapshot["total_general"]} '
                f'y supera el limite permitido de {snapshot["max_total"]} horas.'
            ),
        )

    return bool(form.errors)


def _build_f4_limit_snapshot(docente, periodo, horas_actividad, numero_paralelos, exclude_f4_id=None):
    limite = _get_limite_horario_docente(docente)
    if not limite:
        return {
            'error': 'No hay un límite horario activo configurado para la dedicación del docente.',
        }

    asignadas = PlanificacionAsignacionDocente.objects.filter(
        id_docente=docente,
        id_periodo=periodo,
    ).aggregate(
        total_clase=Sum('horas_clase'),
        total_comp=Sum('horas_complementarias'),
    )
    clase_total = asignadas['total_clase'] or 0
    comp_total = asignadas['total_comp'] or 0

    f4_qs = PlanificacionMatrizF4.objects.filter(
        id_docente=docente,
        id_periodo=periodo,
    )
    if exclude_f4_id is not None:
        f4_qs = f4_qs.exclude(id_registro_f4=exclude_f4_id)

    f4_total = 0
    for row in f4_qs.values('horas_actividad', 'numero_paralelos_actividad'):
        f4_total += (row['horas_actividad'] or 0) * (row['numero_paralelos_actividad'] or 1)

    nueva_f4 = (horas_actividad or 0) * (numero_paralelos or 1)
    max_total = (limite.horas_maximas or 0) + (limite.horas_complementarias_maximas or 0)
    total_general = clase_total + comp_total + f4_total + nueva_f4

    return {
        'limite': limite,
        'horas_clase': clase_total,
        'horas_complementarias': comp_total,
        'f4_previa': f4_total,
        'f4_nueva': nueva_f4,
        'max_total': max_total,
        'total_general': total_general,
        'horas_totales_disponibles': max(0, max_total - total_general),
    }


def _add_f4_limit_errors(form, docente, snapshot):
    if snapshot.get('error'):
        form.add_error(None, snapshot['error'])
        return True

    if snapshot['total_general'] > snapshot['max_total']:
        form.add_error(
            None,
            (
                f'El total de horas para {docente.nombres_completos} seria {snapshot["total_general"]} '
                f'y supera el limite permitido de {snapshot["max_total"]} horas.'
            ),
        )

    return bool(form.errors)


class PlanificacionDemandaAcademicaListView(PlanningFlowContextMixin, CrudListView):
    model = PlanificacionDemandaAcademica
    planning_active_section = 'planificaciondemandaacademica_list'


class PlanificacionDemandaAcademicaCreateView(PlanningFlowContextMixin, CrudCreateView):
    model = PlanificacionDemandaAcademica
    planning_active_section = 'planificaciondemandaacademica_list'
    autofill_rules = {
        'id_asignatura': {
            'app': 'curriculo',
            'model': 'CurriculoAsignatura',
            'fields': {
                'id_carrera': 'id_carrera_id',
            },
        },
    }

class PlanificacionDemandaAcademicaUpdateView(PlanningFlowContextMixin, CrudUpdateView):
    model = PlanificacionDemandaAcademica
    planning_active_section = 'planificaciondemandaacademica_list'

class PlanificacionDemandaAcademicaDeleteView(PeriodEditableDeleteMixin, CrudDeleteView):
    model = PlanificacionDemandaAcademica


class PlanificacionAsignacionDocenteListView(LenientPaginationMixin, CrudListView):
    model = PlanificacionAsignacionDocente
    template_name = 'planificacion/planificacionasignaciondocente_list.html'
    paginate_by = 20

    def get_queryset(self):
        qs = PlanificacionAsignacionDocente.objects.select_related(
            'id_docente',
            'id_asignatura',
            'id_carrera',
            'id_periodo',
            'id_campo',
            'id_docente__id_dedicacion',
        ).order_by(
            'id_periodo__nombre_periodo',
            'id_carrera__nombre_carrera',
            'id_asignatura__nivel_semestre',
            'id_asignatura__nombre_asignatura',
            'paralelo_asignado',
        )

        periodo_id = self.request.GET.get('periodo')
        carrera_id = self.request.GET.get('carrera')
        permitted = _ensure_career_access(self.request, carrera_id)
        docente_id = self.request.GET.get('docente')
        nivel = self.request.GET.get('nivel')
        estado = self.request.GET.get('estado')
        search = (self.request.GET.get('q') or '').strip()

        if permitted is not None:
            qs = qs.filter(id_carrera_id__in=permitted)

        if periodo_id:
            qs = qs.filter(id_periodo_id=periodo_id)
        if carrera_id:
            qs = qs.filter(id_carrera_id=carrera_id)
        if docente_id:
            qs = qs.filter(id_docente_id=docente_id)
        if nivel:
            qs = qs.filter(nivel_semestre_asignado=nivel)
        if search:
            qs = qs.filter(
                Q(id_docente__nombres_completos__icontains=search) |
                Q(id_docente__cedula_docente__icontains=search) |
                Q(id_asignatura__codigo_asignatura__icontains=search) |
                Q(id_asignatura__nombre_asignatura__icontains=search)
            )

        if estado:
            filtered_ids = []
            workload_map = _build_docente_workload_map(periodo_id=periodo_id)
            limites = {
                limite.id_modalidad_id: limite
                for limite in LimiteHorario.objects.filter(activo=True).select_related('id_modalidad')
            }
            for asignacion in qs:
                workload = workload_map.get(asignacion.id_docente_id, {'total_horas': 0})
                limite = limites.get(asignacion.id_docente.id_modalidad_id)
                max_total = ((limite.horas_maximas or 0) + (limite.horas_complementarias_maximas or 0)) if limite else 0
                pct = round((workload['total_horas'] / max_total) * 100, 1) if max_total > 0 else 0
                row_status = 'sobrecargado' if pct > 100 else 'alerta' if pct >= 80 else 'balanceado'
                if row_status == estado:
                    filtered_ids.append(asignacion.pk)
            qs = qs.filter(pk__in=filtered_ids)

        return qs

    def get_context_data(self, **kwargs):
        from catalogos.models import CatalogoCarrera, CatalogoPeriodoAcademico

        ctx = super().get_context_data(**kwargs)
        periodo_id = self.request.GET.get('periodo')
        carrera_id = self.request.GET.get('carrera')
        docente_id = self.request.GET.get('docente')
        nivel = self.request.GET.get('nivel')
        estado = self.request.GET.get('estado') or ''
        search = self.request.GET.get('q') or ''

        rows = list(ctx['object_list'])
        workload_map = _build_docente_workload_map(periodo_id=periodo_id)
        limites = {
            limite.id_modalidad_id: limite
            for limite in LimiteHorario.objects.filter(activo=True).select_related('id_modalidad')
        }

        for row in rows:
            row.paralelo_limpio = _excel_clean_text(row.paralelo_asignado)
            row.total_asignacion = (row.horas_clase or 0) + (row.horas_complementarias or 0)
            row.workload = workload_map.get(row.id_docente_id, _empty_workload())
            row.limite = limites.get(row.id_docente.id_modalidad_id)
            row.max_total = ((row.limite.horas_maximas or 0) + (row.limite.horas_complementarias_maximas or 0)) if row.limite else 0
            row.available = max(0, row.max_total - row.workload['total_horas'])
            row.percentage = round((row.workload['total_horas'] / row.max_total) * 100, 1) if row.max_total > 0 else 0
            if row.percentage > 100:
                row.status = 'sobrecargado'
                row.badge = 'danger'
            elif row.percentage >= 80:
                row.status = 'alerta'
                row.badge = 'warning'
            else:
                row.status = 'balanceado'
                row.badge = 'success'
            row.cumple_limite = row.percentage <= 100

        all_filtered = self.get_queryset()
        total_horas_clase = all_filtered.aggregate(total=Sum('horas_clase'))['total'] or 0
        total_horas_comp = all_filtered.aggregate(total=Sum('horas_complementarias'))['total'] or 0

        ctx.update({
            'active_section': 'planificacionasignaciondocente_list',
            'limit_config': _build_limit_config_state(),
            'filter_querystring': _filter_querystring(self.request),
            'rows': rows,
            'periodos': CatalogoPeriodoAcademico.objects.order_by('-fecha_inicio_periodo', '-id_periodo'),
            'carreras': _scope_careers(CatalogoCarrera.objects.filter(carrera_activa=True), self.request).order_by('nombre_carrera'),
            'docentes': DocenteFcacc.objects.filter(docente_activo=True).order_by('nombres_completos'),
            'level_options': list(range(1, 9)),
            'periodo_id': int(periodo_id) if periodo_id else None,
            'carrera_id': int(carrera_id) if carrera_id else None,
            'docente_id': int(docente_id) if docente_id else None,
            'nivel': int(nivel) if nivel else None,
            'estado': estado,
            'search': search,
            'total_asignaciones': all_filtered.count(),
            'total_horas_clase': total_horas_clase,
            'total_horas_comp': total_horas_comp,
            'total_horas': total_horas_clase + total_horas_comp,
        })
        return ctx


class PlanificacionAsignacionDocenteCreateView(CrudCreateView):
    model = PlanificacionAsignacionDocente
    fields = None
    form_class = PlanificacionAsignacionDocenteForm
    template_name = 'planificacion/asignaciondocente_form.html'
    autofill_rules = {
        'id_asignatura': {
            'app': 'curriculo',
            'model': 'CurriculoAsignatura',
            'fields': {
                'horas_clase': 'horas_semanales_asignatura',
            },
        },
    }

    def get_form(self, form_class=None):
        form = _style_form_fields(super().get_form(form_class))
        permitted = allowed_career_ids(self.request.user)
        if permitted is not None:
            form.fields['id_asignatura'].queryset = form.fields['id_asignatura'].queryset.filter(
                id_carrera_id__in=permitted
            )
        return form

    def form_valid(self, form):
        docente = form.cleaned_data.get('id_docente')
        periodo = form.cleaned_data.get('id_periodo')
        asignatura = form.cleaned_data.get('id_asignatura')
        horas_clase_nuevas = form.cleaned_data.get('horas_clase', 0)
        horas_comp_nuevas = form.cleaned_data.get('horas_complementarias', 0)
        snapshot = _build_asignacion_limit_snapshot(
            docente=docente,
            periodo=periodo,
            horas_clase_nuevas=horas_clase_nuevas,
            horas_comp_nuevas=horas_comp_nuevas,
        )

        if _add_asignacion_limit_errors(form, docente, snapshot, es_actividad=getattr(asignatura, 'es_actividad', False)):
            return self.form_invalid(form)

        return super().form_valid(form)

    def get_initial(self):
        initial = super().get_initial()
        # FK fields: doc/id maps to id_docente, etc.
        fk_map = {'docente': 'id_docente', 'asignatura': 'id_asignatura', 'carrera': 'id_carrera', 'periodo': 'id_periodo', 'campo': 'id_campo'}
        for param, field in fk_map.items():
            val = self.request.GET.get(param)
            if val is not None:
                try:
                    initial[field] = int(val)
                except (ValueError, TypeError):
                    pass
        # Direct fields
        direct_map = {'nivel': 'nivel_semestre_asignado', 'horas': 'horas_clase', 'complementarias': 'horas_complementarias', 'paralelo': 'paralelo_asignado'}
        for param, field in direct_map.items():
            val = self.request.GET.get(param)
            if val is not None:
                initial[field] = int(val) if field != 'paralelo_asignado' else val
        # If subject is in GET param, pre-fill derived fields
        subj_id = self.request.GET.get('asignatura')
        if subj_id:
            try:
                subj = CurriculoAsignatura.objects.get(id_asignatura=subj_id)
                _ensure_career_access(self.request, subj.id_carrera_id)
                initial['id_asignatura'] = subj
                initial['id_carrera'] = subj.id_carrera
                initial['nivel_semestre_asignado'] = subj.nivel_semestre
                if 'id_periodo' not in initial:
                    from catalogos.models import CatalogoPeriodoAcademico
                    periodo_activo = CatalogoPeriodoAcademico.objects.filter(periodo_activo=True).first()
                    if periodo_activo:
                        initial['id_periodo'] = periodo_activo
            except CurriculoAsignatura.DoesNotExist:
                pass
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        subj_id = self.request.GET.get('asignatura') or self.request.POST.get('id_asignatura')
        ctx['limit_config'] = _build_limit_config_state()
        ctx['active_section'] = 'planificacionasignaciondocente_list'
        ctx['form_title'] = 'Nueva asignación docente'
        ctx['form_subtitle'] = 'Registra el docente responsable de una asignatura y paralelo.'
        ctx['back_url'] = 'planificacion:planificacionasignaciondocente_list'
        subject_data = {
            str(s.id_asignatura): {'n': s.nivel_semestre, 'c': s.id_carrera_id, 'a': s.es_actividad}
            for s in _scope_careers(CurriculoAsignatura.objects.only('id_asignatura', 'nivel_semestre', 'id_carrera_id', 'es_actividad'), self.request)
        }
        ctx['subject_data_json'] = json.dumps(subject_data)
        career_data = {
            str(c.id_carrera): {'a': c.es_actividad}
            for c in _scope_careers(CatalogoCarrera.objects.only('id_carrera', 'es_actividad'), self.request)
        }
        ctx['career_data_json'] = json.dumps(career_data)
        docente_id = self.request.POST.get('id_docente') or self.request.GET.get('docente')
        periodo_id = self.request.POST.get('id_periodo') or self.request.GET.get('periodo')
        if docente_id and periodo_id:
            try:
                docente = DocenteFcacc.objects.select_related('id_dedicacion').get(id_docente=docente_id)
                ctx['load_summary'] = _build_form_load_summary(docente=docente, periodo=periodo_id)
            except DocenteFcacc.DoesNotExist:
                pass
        if subj_id:
            try:
                subj = CurriculoAsignatura.objects.get(id_asignatura=subj_id)
                recommendations = _compute_teacher_scores(subj)
                ctx['teacher_recommendations'] = recommendations[:10]
                ctx['selected_subject'] = subj
                ctx['affinity_data_missing'] = (
                    subj.nivel_semestre >= 4 and
                    not CurriculoAsignaturaCampo.objects.filter(id_asignatura=subj).exists()
                )
                from catalogos.models import CatalogoPeriodoAcademico
                periodo_activo = CatalogoPeriodoAcademico.objects.filter(periodo_activo=True).first()
                existing = _get_existing_assignment(subj.id_asignatura, getattr(periodo_activo, 'id_periodo', None))
                if existing:
                    ctx['existing_assignment'] = existing
            except CurriculoAsignatura.DoesNotExist:
                pass
        return ctx


class PlanificacionAsignacionDocenteUpdateView(CrudUpdateView):
    model = PlanificacionAsignacionDocente
    fields = None
    form_class = PlanificacionAsignacionDocenteForm
    template_name = 'planificacion/asignaciondocente_form.html'

    def get_form(self, form_class=None):
        form = _style_form_fields(super().get_form(form_class))
        permitted = allowed_career_ids(self.request.user)
        if permitted is not None:
            form.fields['id_asignatura'].queryset = form.fields['id_asignatura'].queryset.filter(
                id_carrera_id__in=permitted
            )
        return form

    def form_valid(self, form):
        docente = form.cleaned_data.get('id_docente')
        periodo = form.cleaned_data.get('id_periodo')
        asignatura = form.cleaned_data.get('id_asignatura')
        horas_clase_nuevas = form.cleaned_data.get('horas_clase', 0)
        horas_comp_nuevas = form.cleaned_data.get('horas_complementarias', 0)
        snapshot = _build_asignacion_limit_snapshot(
            docente=docente,
            periodo=periodo,
            horas_clase_nuevas=horas_clase_nuevas,
            horas_comp_nuevas=horas_comp_nuevas,
            exclude_asignacion_id=self.object.pk,
        )

        if _add_asignacion_limit_errors(form, docente, snapshot, es_actividad=getattr(asignatura, 'es_actividad', False)):
            return self.form_invalid(form)

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['limit_config'] = _build_limit_config_state()
        ctx['active_section'] = 'planificacionasignaciondocente_list'
        ctx['form_title'] = 'Editar asignación docente'
        ctx['form_subtitle'] = 'Actualiza docente, horas o paralelo asignado.'
        ctx['back_url'] = 'planificacion:planificacionasignaciondocente_list'
        ctx['load_summary'] = _build_form_load_summary(
            docente=self.object.id_docente,
            periodo=self.object.id_periodo_id,
        )
        subject_data = {
            str(s.id_asignatura): {'n': s.nivel_semestre, 'c': s.id_carrera_id, 'a': s.es_actividad}
            for s in _scope_careers(CurriculoAsignatura.objects.only('id_asignatura', 'nivel_semestre', 'id_carrera_id', 'es_actividad'), self.request)
        }
        ctx['subject_data_json'] = json.dumps(subject_data)
        career_data = {
            str(c.id_carrera): {'a': c.es_actividad}
            for c in _scope_careers(CatalogoCarrera.objects.only('id_carrera', 'es_actividad'), self.request)
        }
        ctx['career_data_json'] = json.dumps(career_data)
        return ctx

class PlanificacionAsignacionDocenteDeleteView(PeriodEditableDeleteMixin, CrudDeleteView):
    model = PlanificacionAsignacionDocente


class CatalogoActividadComplementariaListView(AdminOnlyMixin, CrudListView):
    model = CatalogoActividadComplementaria


class CatalogoActividadComplementariaCreateView(AdminOnlyMixin, CrudCreateView):
    model = CatalogoActividadComplementaria


class CatalogoActividadComplementariaUpdateView(AdminOnlyMixin, CrudUpdateView):
    model = CatalogoActividadComplementaria


class CatalogoActividadComplementariaDeleteView(AdminOnlyMixin, CrudDeleteView):
    model = CatalogoActividadComplementaria


class PlanificacionActividadDocenteListView(PlanningFlowContextMixin, CrudListView):
    model = PlanificacionActividadDocente
    select_related_fields = ('id_docente', 'id_periodo', 'id_actividad')
    planning_active_section = 'planificacionactividaddocente_list'

    def get_queryset(self):
        qs = super().get_queryset()
        permitted = allowed_career_ids(self.request.user)
        if permitted is not None:
            qs = qs.filter(
                id_docente__docenteasignacioncarreraperiodo__id_carrera_id__in=permitted,
                id_docente__docenteasignacioncarreraperiodo__id_periodo_id=F('id_periodo_id'),
            ).distinct()
        return qs


class PlanificacionActividadDocenteCreateView(PlanningFlowContextMixin, CrudCreateView):
    model = PlanificacionActividadDocente
    fields = None
    form_class = PlanificacionActividadDocenteForm
    planning_active_section = 'planificacionactividaddocente_list'

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        permitted = allowed_career_ids(self.request.user)
        if permitted is not None:
            form.fields['id_docente'].queryset = form.fields['id_docente'].queryset.filter(
                docenteasignacioncarreraperiodo__id_carrera_id__in=permitted
            ).distinct()
        return form

    def form_valid(self, form):
        permitted = allowed_career_ids(self.request.user)
        if permitted is not None and not form.cleaned_data['id_docente'].docenteasignacioncarreraperiodo_set.filter(
            id_carrera_id__in=permitted, id_periodo=form.cleaned_data['id_periodo']
        ).exists():
            form.add_error('id_docente', 'El docente no pertenece a una carrera autorizada en este período.')
            return self.form_invalid(form)
        return super().form_valid(form)


class PlanificacionActividadDocenteUpdateView(PlanningFlowContextMixin, CrudUpdateView):
    model = PlanificacionActividadDocente
    fields = None
    form_class = PlanificacionActividadDocenteForm
    planning_active_section = 'planificacionactividaddocente_list'

    def get_queryset(self):
        qs = super().get_queryset()
        permitted = allowed_career_ids(self.request.user)
        if permitted is not None:
            qs = qs.filter(
                id_docente__docenteasignacioncarreraperiodo__id_carrera_id__in=permitted,
                id_docente__docenteasignacioncarreraperiodo__id_periodo_id=F('id_periodo_id'),
            ).distinct()
        return qs

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        permitted = allowed_career_ids(self.request.user)
        if permitted is not None:
            form.fields['id_docente'].queryset = form.fields['id_docente'].queryset.filter(
                docenteasignacioncarreraperiodo__id_carrera_id__in=permitted
            ).distinct()
        return form

    def form_valid(self, form):
        permitted = allowed_career_ids(self.request.user)
        if permitted is not None and not form.cleaned_data['id_docente'].docenteasignacioncarreraperiodo_set.filter(
            id_carrera_id__in=permitted, id_periodo=form.cleaned_data['id_periodo']
        ).exists():
            form.add_error('id_docente', 'El docente no pertenece a una carrera autorizada en este período.')
            return self.form_invalid(form)
        return super().form_valid(form)


class PlanificacionActividadDocenteDeleteView(PeriodEditableDeleteMixin, CrudDeleteView):
    model = PlanificacionActividadDocente

    def get_queryset(self):
        qs = super().get_queryset()
        permitted = allowed_career_ids(self.request.user)
        if permitted is not None:
            qs = qs.filter(
                id_docente__docenteasignacioncarreraperiodo__id_carrera_id__in=permitted,
                id_docente__docenteasignacioncarreraperiodo__id_periodo_id=F('id_periodo_id'),
            ).distinct()
        return qs


class PlanificacionRepartoHorasListView(AdminOnlyMixin, CrudListView):
    model = PlanificacionRepartoHoras


class PlanificacionRepartoHorasCreateView(AdminOnlyMixin, CrudCreateView):
    model = PlanificacionRepartoHoras

class PlanificacionRepartoHorasUpdateView(AdminOnlyMixin, CrudUpdateView):
    model = PlanificacionRepartoHoras

class PlanificacionRepartoHorasDeleteView(AdminOnlyMixin, PeriodEditableDeleteMixin, CrudDeleteView):
    model = PlanificacionRepartoHoras


class PlanificacionMatrizF4ListView(AdminOnlyMixin, PlanningFlowContextMixin, LenientPaginationMixin, CrudListView):
    planning_active_section = 'planificacionmatrizf4_list'
    model = PlanificacionMatrizF4
    template_name = 'planificacion/planificacionmatrizf4_list.html'
    paginate_by = 20

    def get_queryset(self):
        qs = PlanificacionMatrizF4.objects.select_related(
            'id_docente',
            'id_carrera',
            'id_periodo',
            'id_grado_afinidad',
        ).order_by('id_docente__nombres_completos', 'tipo_actividad', 'nombre_asignatura_actividad')

        periodo_id = self.request.GET.get('periodo')
        carrera_id = self.request.GET.get('carrera')
        permitted = _ensure_career_access(self.request, carrera_id)
        docente_id = self.request.GET.get('docente')
        tipo = self.request.GET.get('tipo')
        search = (self.request.GET.get('q') or '').strip()

        if permitted is not None:
            qs = qs.filter(id_carrera_id__in=permitted)

        if periodo_id:
            qs = qs.filter(id_periodo_id=periodo_id)
        if carrera_id:
            qs = qs.filter(id_carrera_id=carrera_id)
        if docente_id:
            qs = qs.filter(id_docente_id=docente_id)
        if tipo:
            qs = qs.filter(tipo_actividad=tipo)
        if search:
            qs = qs.filter(
                Q(id_docente__nombres_completos__icontains=search) |
                Q(nombre_asignatura_actividad__icontains=search) |
                Q(tipo_actividad__icontains=search)
            )

        return qs

    def get_context_data(self, **kwargs):
        from catalogos.models import CatalogoCarrera, CatalogoPeriodoAcademico

        ctx = super().get_context_data(**kwargs)
        filtered_qs = self.get_queryset()
        rows = list(ctx['object_list'])
        for row in rows:
            row.total_horas_f4 = (row.horas_actividad or 0) * (row.numero_paralelos_actividad or 1)

        total_horas_registros = 0
        total_horas_consolidadas = 0
        seen_f4 = set()
        dedupe_global = not self.request.GET.get('carrera')
        for row in filtered_qs.values(
            'id_docente_id',
            'tipo_actividad',
            'nombre_asignatura_actividad',
            'horas_actividad',
            'numero_paralelos_actividad',
        ):
            horas = (row['horas_actividad'] or 0) * (row['numero_paralelos_actividad'] or 1)
            total_horas_registros += horas
            key = (
                row['id_docente_id'],
                row['tipo_actividad'],
                row['nombre_asignatura_actividad'],
                row['horas_actividad'],
                row['numero_paralelos_actividad'],
            )
            if not dedupe_global or key not in seen_f4:
                seen_f4.add(key)
                total_horas_consolidadas += horas
        tipos = list(
            PlanificacionMatrizF4.objects.exclude(tipo_actividad__isnull=True)
            .exclude(tipo_actividad='')
            .order_by('tipo_actividad')
            .values_list('tipo_actividad', flat=True)
            .distinct()
        )

        ctx.update({
            'active_section': 'planificacionmatrizf4_list',
            'limit_config': _build_limit_config_state(),
            'filter_querystring': _filter_querystring(self.request),
            'periodos': CatalogoPeriodoAcademico.objects.order_by('-fecha_inicio_periodo', '-id_periodo'),
            'carreras': _scope_careers(CatalogoCarrera.objects.filter(carrera_activa=True), self.request).order_by('nombre_carrera'),
            'docentes': DocenteFcacc.objects.filter(docente_activo=True).order_by('nombres_completos'),
            'tipos': tipos,
            'periodo_id': int(self.request.GET['periodo']) if self.request.GET.get('periodo') else None,
            'carrera_id': int(self.request.GET['carrera']) if self.request.GET.get('carrera') else None,
            'docente_id': int(self.request.GET['docente']) if self.request.GET.get('docente') else None,
            'tipo': self.request.GET.get('tipo') or '',
            'search': self.request.GET.get('q') or '',
            'total_registros': filtered_qs.count(),
            'total_horas': total_horas_consolidadas,
            'total_horas_registros': total_horas_registros,
            'horas_duplicadas': max(0, total_horas_registros - total_horas_consolidadas),
            'docentes_con_f4': filtered_qs.values('id_docente_id').distinct().count(),
            'rows': rows,
        })
        return ctx


class PlanificacionMatrizF4CreateView(AdminOnlyMixin, PlanningFlowContextMixin, CrudCreateView):
    model = PlanificacionMatrizF4
    planning_active_section = 'planificacionmatrizf4_list'

    def get_form(self, form_class=None):
        return _style_form_fields(super().get_form(form_class))

    def form_valid(self, form):
        docente = form.cleaned_data.get('id_docente')
        periodo = form.cleaned_data.get('id_periodo')
        horas_actividad = form.cleaned_data.get('horas_actividad', 0)
        numero_paralelos = form.cleaned_data.get('numero_paralelos_actividad', 1)
        snapshot = _build_f4_limit_snapshot(
            docente=docente,
            periodo=periodo,
            horas_actividad=horas_actividad,
            numero_paralelos=numero_paralelos,
        )

        if _add_f4_limit_errors(form, docente, snapshot):
            return self.form_invalid(form)

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['limit_config'] = _build_limit_config_state()
        ctx['form_title'] = 'Nueva actividad F4'
        ctx['form_subtitle'] = 'Registra actividades, investigación u otras horas que suman a la carga docente.'
        return ctx

class PlanificacionMatrizF4UpdateView(AdminOnlyMixin, PlanningFlowContextMixin, CrudUpdateView):
    model = PlanificacionMatrizF4
    planning_active_section = 'planificacionmatrizf4_list'

    def get_form(self, form_class=None):
        return _style_form_fields(super().get_form(form_class))

    def form_valid(self, form):
        docente = form.cleaned_data.get('id_docente')
        periodo = form.cleaned_data.get('id_periodo')
        horas_actividad = form.cleaned_data.get('horas_actividad', 0)
        numero_paralelos = form.cleaned_data.get('numero_paralelos_actividad', 1)
        snapshot = _build_f4_limit_snapshot(
            docente=docente,
            periodo=periodo,
            horas_actividad=horas_actividad,
            numero_paralelos=numero_paralelos,
            exclude_f4_id=self.object.pk,
        )

        if _add_f4_limit_errors(form, docente, snapshot):
            return self.form_invalid(form)

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['limit_config'] = _build_limit_config_state()
        ctx['form_title'] = 'Editar actividad F4'
        ctx['form_subtitle'] = 'Actualiza horas, actividad o docente asociado.'
        ctx['load_summary'] = _build_form_load_summary(
            docente=self.object.id_docente,
            periodo=self.object.id_periodo_id,
        )
        return ctx

class PlanificacionMatrizF4DeleteView(AdminOnlyMixin, PeriodEditableDeleteMixin, CrudDeleteView):
    model = PlanificacionMatrizF4


class PlanificacionAulaHorarioListView(PlanningFlowContextMixin, CrudListView):
    model = PlanificacionAulaHorario
    planning_active_section = 'planificacionaulahorario_list'
    select_related_fields = ('id_periodo', 'id_asignacion', 'id_asignacion__id_docente', 'id_asignacion__id_asignatura')


class PlanificacionAulaHorarioCreateView(PlanningFlowContextMixin, CrudCreateView):
    model = PlanificacionAulaHorario
    fields = None
    form_class = PlanificacionAulaHorarioForm
    planning_active_section = 'planificacionaulahorario_list'

class PlanificacionAulaHorarioUpdateView(PlanningFlowContextMixin, CrudUpdateView):
    model = PlanificacionAulaHorario
    fields = None
    form_class = PlanificacionAulaHorarioForm
    planning_active_section = 'planificacionaulahorario_list'

class PlanificacionAulaHorarioDeleteView(PeriodEditableDeleteMixin, CrudDeleteView):
    model = PlanificacionAulaHorario


# ——— Reporte: Horas por Docente ———————————————————————————————

@login_required
@module_permission_required('planificacion', 'view')
def reporte_horas_docentes(request):
    """Ruta histórica; la carga se consolidó en una sola interfaz."""
    target = reverse('planificacion:planificacion_consolidada_docentes')
    query = request.GET.urlencode()
    return redirect(f'{target}?{query}' if query else target)


@login_required
@module_permission_required('planificacion', 'view')
def validacion_excel_planificacion(request):
    """Compatibilidad con marcadores antiguos; el sistema ya no depende del Excel."""
    query = request.GET.urlencode()
    target = reverse('planificacion:control_calidad')
    return redirect(f'{target}?{query}' if query else target)


@login_required
@module_permission_required('planificacion', 'change')
def sincronizar_excel_planificacion(request):
    messages.info(
        request,
        'Las importaciones Excel son herramientas administrativas opcionales y no forman parte del flujo normal.',
    )
    return redirect('planificacion:control_calidad')


@login_required
@module_permission_required('planificacion', 'view')
def control_calidad_planificacion(request):
    """Audita la planificación usando exclusivamente la base de datos."""
    periodos = CatalogoPeriodoAcademico.objects.order_by('-fecha_inicio_periodo', '-id_periodo')
    periodo_id = request.GET.get('periodo')
    if not periodo_id:
        activo = periodos.filter(periodo_activo=True).first()
        periodo_id = str(activo.id_periodo) if activo else None

    issues = []
    if periodo_id:
        demandas = list(PlanificacionDemandaAcademica.objects.filter(
            id_periodo_id=periodo_id,
        ).select_related('id_asignatura', 'id_carrera'))
        asignaciones = list(PlanificacionAsignacionDocente.objects.filter(
            id_periodo_id=periodo_id,
        ).select_related('id_asignatura', 'id_carrera', 'id_docente', 'id_campo'))
        campos_por_asignatura = {}
        for subject_id, field_id in CurriculoAsignaturaCampo.objects.filter(
            id_asignatura_id__in=[d.id_asignatura_id for d in demandas],
        ).values_list('id_asignatura_id', 'id_campo_id'):
            campos_por_asignatura.setdefault(subject_id, set()).add(field_id)

        asignadas = {
            (a.id_asignatura_id, a.id_carrera_id, normalize_parallel(a.paralelo_asignado)): a
            for a in asignaciones
        }
        demanda_map = {(d.id_asignatura_id, d.id_carrera_id): d for d in demandas}

        for demanda in demandas:
            if not campos_por_asignatura.get(demanda.id_asignatura_id):
                issues.append({
                    'severity': 'error', 'category': 'Campo de conocimiento',
                    'title': demanda.id_asignatura.nombre_asignatura,
                    'detail': f'{demanda.id_carrera.nombre_carrera} · nivel {demanda.id_asignatura.nivel_semestre}: no tiene campo asociado.',
                })
            for paralelo in _build_parallel_labels(demanda.numero_paralelos):
                if (demanda.id_asignatura_id, demanda.id_carrera_id, paralelo) not in asignadas:
                    issues.append({
                        'severity': 'warning', 'category': 'Paralelo pendiente',
                        'title': f'{demanda.id_asignatura.nombre_asignatura} · {paralelo}',
                        'detail': f'{demanda.id_carrera.nombre_carrera} · nivel {demanda.id_asignatura.nivel_semestre}.',
                    })

        for item in asignaciones:
            errors = []
            demanda = demanda_map.get((item.id_asignatura_id, item.id_carrera_id))
            if not demanda:
                errors.append('no existe en la demanda académica')
            elif normalize_parallel(item.paralelo_asignado) not in _build_parallel_labels(demanda.numero_paralelos):
                errors.append('el paralelo está fuera de la demanda')
            if item.id_asignatura.id_carrera_id != item.id_carrera_id:
                errors.append('la carrera no corresponde a la asignatura')
            if item.id_asignatura.nivel_semestre != item.nivel_semestre_asignado:
                errors.append('el nivel no corresponde a la malla')
            if item.id_campo_id not in campos_por_asignatura.get(item.id_asignatura_id, set()):
                errors.append('el campo de conocimiento no corresponde')
            if item.nivel_semestre_asignado >= 4 and not docente_tiene_afinidad(item.id_docente, item.id_asignatura):
                errors.append('el docente no tiene la afinidad requerida')
            if item.horas_clase != (item.id_asignatura.horas_semanales_asignatura or 0):
                errors.append('las horas no coinciden con la malla')
            if errors:
                issues.append({
                    'severity': 'error', 'category': 'Asignación inconsistente',
                    'title': f'{item.id_docente.nombres_completos} · {item.id_asignatura.nombre_asignatura}',
                    'detail': '; '.join(errors).capitalize() + '.',
                })

        workload = _build_docente_workload_map(periodo_id=periodo_id)
        for docente in DocenteFcacc.objects.filter(id_docente__in=workload).select_related('id_modalidad'):
            limite = _get_limite_horario_docente(docente)
            total = workload[docente.id_docente]['total_horas']
            maximo = ((limite.horas_maximas or 0) + (limite.horas_complementarias_maximas or 0)) if limite else 0
            if not limite:
                issues.append({
                    'severity': 'error', 'category': 'Límite horario',
                    'title': docente.nombres_completos,
                    'detail': 'La modalidad del docente no tiene límites configurados.',
                })
            elif total > maximo:
                issues.append({
                    'severity': 'error', 'category': 'Sobrecarga docente',
                    'title': docente.nombres_completos,
                    'detail': f'Carga asignada: {total}h; límite configurado: {maximo}h.',
                })

    summary = {
        'total': len(issues),
        'errors': sum(1 for issue in issues if issue['severity'] == 'error'),
        'warnings': sum(1 for issue in issues if issue['severity'] == 'warning'),
    }
    severity = request.GET.get('severidad') or ''
    if severity:
        issues = [issue for issue in issues if issue['severity'] == severity]
    _, page_obj, page_rows = _paginate_items(request, issues, 25)
    return render(request, 'planificacion/control_calidad.html', {
        'active_section': 'control_calidad', 'show_planning_flow': True,
        'periodos': periodos, 'periodo_id': int(periodo_id) if periodo_id else None,
        'severidad': severity, 'rows': page_rows, 'summary': summary,
        'page_obj': page_obj, 'paginator': page_obj.paginator,
        'is_paginated': page_obj.has_other_pages(),
        'filter_querystring': _filter_querystring(request),
    })


@login_required
@module_permission_required('planificacion', 'change')
def cambiar_estado_periodo(request, periodo_id):
    if request.method != 'POST':
        return redirect('planificacion:planificacion_operativa')
    periodo = get_object_or_404(CatalogoPeriodoAcademico, pk=periodo_id)
    nuevo = request.POST.get('estado')
    transiciones = {
        'BORRADOR': {'EN_REVISION'},
        'EN_REVISION': {'BORRADOR', 'APROBADO'},
        'APROBADO': {'EN_REVISION', 'CERRADO'},
        'CERRADO': set(),
    }
    if nuevo not in transiciones.get(periodo.estado_planificacion, set()):
        messages.error(request, 'La transición de estado solicitada no está permitida.')
        return redirect(f"{reverse('planificacion:planificacion_operativa')}?periodo={periodo_id}")

    if nuevo == 'APROBADO':
        demandas = PlanificacionDemandaAcademica.objects.filter(id_periodo=periodo)
        slots = sum(max(0, item.numero_paralelos) for item in demandas)
        asignadas = PlanificacionAsignacionDocente.objects.filter(id_periodo=periodo).count()
        sin_campo = demandas.filter(id_asignatura__curriculoasignaturacampo__isnull=True).distinct().count()
        workload = _build_docente_workload_map(periodo_id=periodo_id)
        sobrecargados = 0
        for docente in DocenteFcacc.objects.filter(id_docente__in=workload).select_related('id_modalidad'):
            limite = _get_limite_horario_docente(docente)
            maximo = ((limite.horas_maximas or 0) + (limite.horas_complementarias_maximas or 0)) if limite else 0
            if not limite or workload[docente.id_docente]['total_horas'] > maximo:
                sobrecargados += 1
        errores = []
        if asignadas < slots:
            errores.append(f'faltan {slots - asignadas} paralelos por asignar')
        if sin_campo:
            errores.append(f'hay {sin_campo} demandas sin campo de conocimiento')
        if sobrecargados:
            errores.append(f'hay {sobrecargados} docentes sin límite válido o sobrecargados')
        if errores:
            messages.error(request, 'No se puede aprobar: ' + '; '.join(errores) + '.')
            return redirect(f"{reverse('planificacion:planificacion_operativa')}?periodo={periodo_id}")

    estado_anterior = periodo.estado_planificacion
    periodo.estado_planificacion = nuevo
    periodo.save(update_fields=['estado_planificacion'])
    try:
        from core.crud_base import _audit_log
        _audit_log(request, periodo, 'UPDATE', old_values={'estado_planificacion': estado_anterior})
    except Exception:
        pass
    messages.success(request, f'El periodo pasó a {periodo.get_estado_planificacion_display()}.')
    return redirect(f"{reverse('planificacion:planificacion_operativa')}?periodo={periodo_id}")


@login_required
@module_permission_required('planificacion', 'change')
@transaction.atomic
def copiar_planificacion_periodo(request, periodo_id):
    if request.method != 'POST':
        return redirect('planificacion:planificacion_operativa')
    destino = get_object_or_404(CatalogoPeriodoAcademico, pk=periodo_id)
    try:
        assert_periodo_editable(destino)
    except ValidationError as exc:
        messages.error(request, exc.messages[0])
        return redirect(f"{reverse('planificacion:planificacion_operativa')}?periodo={periodo_id}")
    origen_id = request.POST.get('origen')
    if not origen_id:
        messages.error(request, 'Seleccione el período que desea copiar.')
        return redirect(f"{reverse('planificacion:planificacion_operativa')}?periodo={periodo_id}")
    origen = get_object_or_404(CatalogoPeriodoAcademico, pk=origen_id)
    if origen.pk == destino.pk:
        messages.error(request, 'El período de origen y destino deben ser diferentes.')
        return redirect(f"{reverse('planificacion:planificacion_operativa')}?periodo={periodo_id}")

    demandas = asignaciones = actividades = f4_creadas = 0
    for item in PlanificacionDemandaAcademica.objects.filter(id_periodo=origen):
        _, created = PlanificacionDemandaAcademica.objects.update_or_create(
            id_asignatura=item.id_asignatura, id_carrera=item.id_carrera, id_periodo=destino,
            defaults={'proyeccion_estudiantes': item.proyeccion_estudiantes, 'numero_paralelos': item.numero_paralelos},
        )
        demandas += int(created)
    for item in PlanificacionAsignacionDocente.objects.filter(id_periodo=origen):
        _, created = PlanificacionAsignacionDocente.objects.update_or_create(
            id_asignatura=item.id_asignatura, id_carrera=item.id_carrera,
            id_periodo=destino, paralelo_asignado=item.paralelo_asignado,
            defaults={
                'id_docente': item.id_docente, 'id_campo': item.id_campo,
                'nivel_semestre_asignado': item.nivel_semestre_asignado,
                'horas_clase': item.horas_clase, 'horas_complementarias': 0,
                'semanas_planificadas': item.semanas_planificadas,
                'comision_servicio': item.comision_servicio,
            },
        )
        asignaciones += int(created)
    for item in PlanificacionActividadDocente.objects.filter(id_periodo=origen):
        _, created = PlanificacionActividadDocente.objects.update_or_create(
            id_docente=item.id_docente, id_periodo=destino, id_actividad=item.id_actividad,
            defaults={'horas_asignadas': item.horas_asignadas, 'observaciones': item.observaciones},
        )
        actividades += int(created)
    for item in PlanificacionMatrizF4.objects.filter(id_periodo=origen):
        _, created = PlanificacionMatrizF4.objects.get_or_create(
            id_docente=item.id_docente, id_carrera=item.id_carrera, id_periodo=destino,
            id_grado_afinidad=item.id_grado_afinidad, tipo_actividad=item.tipo_actividad,
            nombre_asignatura_actividad=item.nombre_asignatura_actividad,
            nivel_semestre_actividad=item.nivel_semestre_actividad,
            horas_actividad=item.horas_actividad,
            numero_paralelos_actividad=item.numero_paralelos_actividad,
            defaults={'observaciones': item.observaciones},
        )
        f4_creadas += int(created)
    messages.success(
        request,
        f'Planificación copiada desde {origen}: {demandas} demandas, {asignaciones} asignaciones, '
        f'{actividades} actividades y {f4_creadas} registros F4 nuevos.',
    )
    return redirect(f"{reverse('planificacion:planificacion_operativa')}?periodo={periodo_id}")


def _get_existing_assignment(asignatura_id, periodo_id=None):
    from catalogos.models import CatalogoPeriodoAcademico
    if not periodo_id:
        periodo_activo = CatalogoPeriodoAcademico.objects.filter(periodo_activo=True).first()
        if periodo_activo:
            periodo_id = periodo_activo.id_periodo
    if not periodo_id:
        return None

    asignaciones = PlanificacionAsignacionDocente.objects.select_related('id_docente').filter(
        id_asignatura_id=asignatura_id,
        id_periodo_id=periodo_id,
    )
    if not asignaciones.exists():
        return None

    docentes = sorted({
        (asignacion.id_docente_id, asignacion.id_docente.nombres_completos)
        for asignacion in asignaciones
    }, key=lambda item: item[1])
    return {
        'id_docente': docentes[0][0],
        'docente_nombre': docentes[0][1],
        'total_docentes': len(docentes),
        'docente_label': docentes[0][1] if len(docentes) == 1 else f'{docentes[0][1]} (+{len(docentes) - 1} mas)',
    }


# ——— Scoring reutilizable: compatibilidad docente ↔ asignatura ————

def _compute_teacher_scores(subject, periodo_id=None):
    """Return list of teacher dicts with compatibility score for a given subject."""
    from docentes.models import DocenteTituloAcademico

    # Subject campos
    subj_campos = list(CurriculoAsignaturaCampo.objects.filter(id_asignatura=subject).select_related('id_campo'))
    req_campo_ids = {c.id_campo_id for c in subj_campos}

    # All active teachers
    docentes = list(DocenteFcacc.objects.filter(docente_activo=True).select_related('id_dedicacion'))
    if not docentes:
        return []

    # Teacher → campos
    afinidad_rels = list(DocenteCampoAfinidad.objects.all())
    doc_campos = {}
    for ar in afinidad_rels:
        doc_campos.setdefault(ar.id_docente_id, set()).add(ar.id_campo_id)

    # Posgrado → campos
    posgrado_campos = {}
    for rpc in RelacionPosgradoCampo.objects.all():
        posgrado_campos.setdefault(rpc.id_posgrado_id, set()).add(rpc.id_campo_id)

    # Teacher → previous subjects
    prev_dict = {}
    prev_qs = PlanificacionAsignacionDocente.objects.filter(id_asignatura=subject)
    if periodo_id:
        prev_qs = prev_qs.filter(id_periodo_id=periodo_id)
    for p in prev_qs.values('id_docente_id'):
        prev_dict[p['id_docente_id']] = True

    # Teacher titles with posgrado
    doc_titulos = {}
    for t in DocenteTituloAcademico.objects.filter(id_posgrado__isnull=False).values('id_docente_id', 'id_posgrado_id'):
        doc_titulos.setdefault(t['id_docente_id'], []).append(t['id_posgrado_id'])

    # Teacher current hours, including F4 activities for the selected period
    workload_map = _build_docente_workload_map(periodo_id=periodo_id)
    limites = {l.id_modalidad_id: l for l in LimiteHorario.objects.filter(activo=True)}

    results = []
    for d in docentes:
        score = 0
        reasons = []

        # 1. Direct campo affinity (+50)
        teacher_campo_ids = doc_campos.get(d.id_docente, set())
        match_ids = req_campo_ids & teacher_campo_ids
        has_affinity = bool(match_ids)
        if match_ids:
            score += 50
            camp_names = [str(c) for c in subj_campos if c.id_campo_id in match_ids]
            reasons.append(f'Campo afinidad: {", ".join(camp_names)}')

        # 2. Title/Posgrado match (+25)
        for posgrado_id in doc_titulos.get(d.id_docente, []):
            if posgrado_campos.get(posgrado_id, set()) & req_campo_ids:
                score += 25
                reasons.append('Posgrado afín')
                has_affinity = True
                break

        if subject.nivel_semestre >= 4 and not has_affinity and not subject.es_actividad:
            continue
        if subject.nivel_semestre <= 3:
            reasons.append('Nivel 1–3: cualquier docente activo')

        # 3. Previous experience (+20)
        if prev_dict.get(d.id_docente):
            score += 20
            reasons.append('Experiencia previa')

        # 4. Available hours (+10)
        limite = limites.get(d.id_modalidad_id)
        max_h = (limite.horas_maximas or 0) + (limite.horas_complementarias_maximas or 0) if limite else 0
        docente_workload = workload_map.get(d.id_docente, _empty_workload())
        used_h = docente_workload['total_horas']
        available = max_h - used_h
        if available > 0:
            score += 10

        results.append({
                'docente': d,
                'id': d.id_docente,
                'score': score,
                'reasons': reasons,
                'available': max(0, available),
                'used': used_h,
                'max': max_h,
                'class_hours': docente_workload['horas_clase'],
                'affinity_hours': docente_workload['horas_afinidad'],
                'non_affinity_hours': docente_workload['horas_no_afinidad'],
                'basic_unit_hours': docente_workload['horas_unidad_basica'],
                'complementary_hours': docente_workload['horas_complementarias'],
                'investigation_hours': docente_workload['horas_investigacion'],
                'management_hours': docente_workload['horas_gestion'],
                'outreach_hours': docente_workload['horas_vinculacion'],
                'activity_hours': docente_workload['horas_actividad'],
                'activities_total': docente_workload['horas_actividades_total'],
                'status': 'excelente' if score >= 50 else 'bueno' if score >= 25 else 'regular',
                'has_affinity': has_affinity,
            })

    results.sort(key=lambda r: (-r['score'], -r['available']))
    return results


# ——— Asignación Inteligente (página de exploración) ——————————

@login_required
@module_permission_required('planificacion', 'view')
def asignacion_inteligente(request):
    carrera_id = request.GET.get('carrera')
    _ensure_career_access(request, carrera_id)
    subject_id = request.GET.get('asignatura')
    periodo_id = request.GET.get('periodo')
    nivel = request.GET.get('nivel')

    from catalogos.models import CatalogoCarrera, CatalogoPeriodoAcademico
    carreras = _scope_careers(CatalogoCarrera.objects.filter(carrera_activa=True), request)
    periodos = CatalogoPeriodoAcademico.objects.order_by('-fecha_inicio_periodo', '-id_periodo')
    periodo_activo = periodos.filter(periodo_activo=True).first()
    if not periodo_id and periodo_activo:
        periodo_id = str(periodo_activo.id_periodo)

    qs = _scope_careers(CurriculoAsignatura.objects.select_related('id_carrera'), request).order_by('id_carrera_id', 'nivel_semestre', 'nombre_asignatura')
    if carrera_id:
        qs = qs.filter(id_carrera_id=carrera_id)
    if subject_id:
        qs = qs.filter(id_asignatura=subject_id)
    if nivel:
        qs = qs.filter(nivel_semestre=nivel)

    subj_qs = _scope_careers(CurriculoAsignatura.objects.select_related('id_carrera'), request).order_by('id_carrera_id', 'nivel_semestre', 'nombre_asignatura')
    if carrera_id:
        subj_qs = subj_qs.filter(id_carrera_id=carrera_id)
    if nivel:
        subj_qs = subj_qs.filter(nivel_semestre=nivel)
    all_subjects_for_select = list(subj_qs.values('id_asignatura', 'nombre_asignatura', 'codigo_asignatura', 'es_actividad'))

    paginator = Paginator(qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    subjects_data = []
    for a in page_obj:
        req_campos = list(CurriculoAsignaturaCampo.objects.filter(id_asignatura=a).select_related('id_campo'))
        recs = _compute_teacher_scores(a, periodo_id=periodo_id)
        existing = _get_existing_assignment(a.id_asignatura, periodo_id)
        subjects_data.append({
            'asignatura': a,
            'req_campos': req_campos,
            'recomendados': recs[:5],
            'existing_assignment': existing,
        })

    from catalogos.models import CatalogoCampoConocimiento
    campos = CatalogoCampoConocimiento.objects.values('id_campo', 'codigo_campo', 'nombre_campo_conocimiento').order_by('codigo_campo')

    context = {
        'active_section': 'asignacion_inteligente',
        'subjects_data': subjects_data,
        'limit_config': _build_limit_config_state(),
        'carreras': carreras,
        'campos': campos,
        'carrera_id': int(carrera_id) if carrera_id else None,
        'subject_id': int(subject_id) if subject_id else None,
        'periodo_id': int(periodo_id) if periodo_id else None,
        'nivel_selected': int(nivel) if nivel else None,
        'niveles': range(1, 11),
        'page_obj': page_obj,
        'paginator': paginator,
        'total_subjects': paginator.count,
        'all_subjects': all_subjects_for_select,
        'periodo_activo': periodo_activo,
    }
    return render(request, 'planificacion/asignacion_inteligente.html', context)


@login_required
@module_permission_required('planificacion', 'view')
def planificacion_paralelos_matriz(request):
    from catalogos.models import CatalogoCarrera, CatalogoPeriodoAcademico

    periodo_id = request.GET.get('periodo')
    carrera_id = request.GET.get('carrera')
    _ensure_career_access(request, carrera_id)
    nivel = request.GET.get('nivel')
    search = (request.GET.get('q') or '').strip()

    periodos = CatalogoPeriodoAcademico.objects.order_by('-fecha_inicio_periodo', '-id_periodo')
    periodo_activo = periodos.filter(periodo_activo=True).first()
    if not periodo_id and periodo_activo:
        periodo_id = str(periodo_activo.id_periodo)

    carreras = _scope_careers(CatalogoCarrera.objects.filter(carrera_activa=True), request).order_by('nombre_carrera')

    demanda_qs = _scope_careers(PlanificacionDemandaAcademica.objects.select_related(
        'id_asignatura', 'id_carrera', 'id_periodo',
    ), request).order_by('id_asignatura__nivel_semestre', 'id_asignatura__nombre_asignatura')

    if periodo_id:
        demanda_qs = demanda_qs.filter(id_periodo_id=periodo_id)
    if carrera_id:
        demanda_qs = demanda_qs.filter(id_carrera_id=carrera_id)
    if nivel:
        demanda_qs = demanda_qs.filter(id_asignatura__nivel_semestre=nivel)

    nivel_qs = _scope_careers(PlanificacionDemandaAcademica.objects.all(), request)
    if periodo_id:
        nivel_qs = nivel_qs.filter(id_periodo_id=periodo_id)
    if carrera_id:
        nivel_qs = nivel_qs.filter(id_carrera_id=carrera_id)
    level_options = sorted(
        nivel_qs.values_list('id_asignatura__nivel_semestre', flat=True).distinct()
    )

    if search:
        demanda_qs = demanda_qs.filter(
            Q(id_asignatura__nombre_asignatura__icontains=search) |
            Q(id_asignatura__codigo_asignatura__icontains=search)
        )

    demandas = list(demanda_qs)
    subject_ids = [d.id_asignatura_id for d in demandas]

    # Determine all paralelo labels across all filtered demandas
    max_paralelos = max((d.numero_paralelos for d in demandas), default=0)
    all_paralelo_labels = _build_parallel_labels(max_paralelos)

    # Lookup existing assignments for these subjects
    asignaciones_por_clave = {}
    if subject_ids:
        asignaciones_qs = PlanificacionAsignacionDocente.objects.filter(
            id_asignatura_id__in=subject_ids,
        ).select_related('id_docente', 'id_campo')
        if periodo_id:
            asignaciones_qs = asignaciones_qs.filter(id_periodo_id=periodo_id)
        if carrera_id:
            asignaciones_qs = asignaciones_qs.filter(id_carrera_id=carrera_id)
        for asignacion in asignaciones_qs:
            key = (asignacion.id_asignatura_id, _excel_clean_text(asignacion.paralelo_asignado))
            asignaciones_por_clave.setdefault(key, []).append(asignacion)

    # Campos lookup
    campos_por_asignatura = {}
    if subject_ids:
        for rel in CurriculoAsignaturaCampo.objects.filter(
            id_asignatura_id__in=subject_ids
        ).select_related('id_campo'):
            campos_por_asignatura.setdefault(rel.id_asignatura_id, []).append(rel)

    rows = []
    for demanda in demandas:
        paralelo_cells = []
        for label in all_paralelo_labels:
            key = (demanda.id_asignatura_id, label)
            asignaciones = asignaciones_por_clave.get(key, [])
            within_demanda = label in _build_parallel_labels(demanda.numero_paralelos)
            paralelo_cells.append({
                'label': label,
                'asignaciones': asignaciones,
                'has_assignment': bool(asignaciones),
                'exists_in_demanda': within_demanda,
            })

        principal_campo = campos_por_asignatura.get(demanda.id_asignatura_id)
        rows.append({
            'demanda': demanda,
            'principal_campo': principal_campo,
            'paralelo_cells': paralelo_cells,
        })

    context = {
        'active_section': 'planificacion_paralelos_matriz',
        'limit_config': _build_limit_config_state(),
        'periodos': periodos,
        'periodo_activo': periodo_activo,
        'periodo_id': int(periodo_id) if periodo_id else None,
        'carreras': carreras,
        'carrera_id': int(carrera_id) if carrera_id else None,
        'nivel': int(nivel) if nivel else None,
        'search': search,
        'level_options': level_options,
        'all_paralelo_labels': all_paralelo_labels,
        'rows': rows,
        'total_subjects': len(rows),
    }
    return render(request, 'planificacion/planificacion_paralelos_matriz.html', context)


@login_required
@module_permission_required('planificacion', 'view')
def planificacion_operativa(request):
    from catalogos.models import CatalogoCarrera, CatalogoPeriodoAcademico

    periodos = CatalogoPeriodoAcademico.objects.order_by('-fecha_inicio_periodo', '-id_periodo')
    periodo_activo = periodos.filter(periodo_activo=True).first()
    periodo_id = request.GET.get('periodo')
    carrera_id = request.GET.get('carrera')
    _ensure_career_access(request, carrera_id)
    nivel = request.GET.get('nivel')
    estado = request.GET.get('estado')
    search = (request.GET.get('q') or '').strip()

    if not periodo_id and periodo_activo:
        periodo_id = str(periodo_activo.id_periodo)

    carreras = _scope_careers(CatalogoCarrera.objects.filter(carrera_activa=True), request).order_by('nombre_carrera')

    demanda_qs = _scope_careers(PlanificacionDemandaAcademica.objects.select_related(
        'id_asignatura',
        'id_carrera',
        'id_periodo',
    ), request).order_by('id_carrera__nombre_carrera', 'id_asignatura__nivel_semestre', 'id_asignatura__nombre_asignatura')

    if periodo_id:
        demanda_qs = demanda_qs.filter(id_periodo_id=periodo_id)
    if carrera_id:
        demanda_qs = demanda_qs.filter(id_carrera_id=carrera_id)
    # Compute level options BEFORE nivel filter so the dropdown always shows all available levels
    nivel_qs = _scope_careers(PlanificacionDemandaAcademica.objects.filter(
        id_asignatura__nivel_semestre__isnull=False,
    ), request)
    if periodo_id:
        nivel_qs = nivel_qs.filter(id_periodo_id=periodo_id)
    if carrera_id:
        nivel_qs = nivel_qs.filter(id_carrera_id=carrera_id)
    level_options = sorted(
        nivel_qs.values_list('id_asignatura__nivel_semestre', flat=True).distinct()
    )

    if nivel:
        demanda_qs = demanda_qs.filter(id_asignatura__nivel_semestre=nivel)
    if search:
        demanda_qs = demanda_qs.filter(
            Q(id_asignatura__nombre_asignatura__icontains=search) |
            Q(id_asignatura__codigo_asignatura__icontains=search)
        )

    demandas = list(demanda_qs)
    subject_ids = [d.id_asignatura_id for d in demandas]

    campos_por_asignatura = {}
    if subject_ids:
        campos_qs = CurriculoAsignaturaCampo.objects.filter(
            id_asignatura_id__in=subject_ids
        ).select_related('id_campo', 'id_asignatura')
        for rel in campos_qs:
            campos_por_asignatura.setdefault(rel.id_asignatura_id, []).append(rel)

    asignaciones_por_clave = {}
    workload_map = _build_docente_workload_map(periodo_id=periodo_id)
    if demandas:
        asignaciones_qs = PlanificacionAsignacionDocente.objects.filter(
            id_asignatura_id__in=subject_ids,
        ).select_related('id_docente', 'id_campo')
        if periodo_id:
            asignaciones_qs = asignaciones_qs.filter(id_periodo_id=periodo_id)
        if carrera_id:
            asignaciones_qs = asignaciones_qs.filter(id_carrera_id=carrera_id)
        for asignacion in asignaciones_qs:
            docente_workload = workload_map.get(
                asignacion.id_docente_id, _empty_workload()
            )
            asignacion.workload = docente_workload
            key = (
                asignacion.id_asignatura_id,
                asignacion.id_carrera_id,
                _excel_clean_text(asignacion.paralelo_asignado),
            )
            asignaciones_por_clave.setdefault(key, []).append(asignacion)
    rows = []
    total_parallel_slots = 0
    assigned_parallel_slots = 0

    for demanda in demandas:
        campos = campos_por_asignatura.get(demanda.id_asignatura_id, [])
        principal_campo = campos[0].id_campo if campos else None
        parallel_rows = []
        parallel_labels = _build_parallel_labels(demanda.numero_paralelos)
        assigned_count = 0

        for paralelo in parallel_labels:
            key = (demanda.id_asignatura_id, demanda.id_carrera_id, paralelo)
            asignaciones = asignaciones_por_clave.get(key, [])
            has_assignment = bool(asignaciones)
            if has_assignment:
                assigned_count += 1
            parallel_rows.append({
                'label': paralelo,
                'asignaciones': asignaciones,
                'has_assignment': has_assignment,
            })

        total_parallel_slots += len(parallel_rows)
        assigned_parallel_slots += assigned_count
        recommendation_preview = _compute_teacher_scores(demanda.id_asignatura, periodo_id=periodo_id)[:3]
        required_class_hours = demanda.id_asignatura.horas_semanales_asignatura

        status = 'completa' if assigned_count >= demanda.numero_paralelos else 'parcial' if assigned_count > 0 else 'pendiente'
        if estado and status != estado:
            continue

        rows.append({
            'demanda': demanda,
            'campos': campos,
            'principal_campo': principal_campo,
            'principal_campo_id': principal_campo.id_campo if principal_campo else None,
            'principal_campo_nombre': principal_campo.nombre_campo_conocimiento if principal_campo else '',
            'parallel_rows': parallel_rows,
            'assigned_parallel_count': assigned_count,
            'pending_parallel_count': max(0, demanda.numero_paralelos - assigned_count),
            'recommendation_preview': recommendation_preview,
            'required_class_hours': required_class_hours,
            'status': status,
        })

    paginator, page_obj, page_rows = _paginate_items(request, rows, 6)

    demandas_sin_campo = [
        demanda for demanda in demandas
        if not campos_por_asignatura.get(demanda.id_asignatura_id)
    ]
    docentes_sobrecargados = []
    for docente in DocenteFcacc.objects.filter(id_docente__in=workload_map).select_related('id_modalidad'):
        limite = _get_limite_horario_docente(docente)
        carga = workload_map.get(docente.id_docente, _empty_workload())
        maximo = ((limite.horas_maximas or 0) + (limite.horas_complementarias_maximas or 0)) if limite else 0
        if not limite or carga['total_horas'] > maximo:
            docentes_sobrecargados.append({
                'docente': docente,
                'total': carga['total_horas'],
                'limite': maximo,
            })

    context = {
        'active_section': 'planificacion_operativa',
        'limit_config': _build_limit_config_state(),
        'periodo_activo': periodo_activo,
        'periodos': periodos,
        'periodo_id': int(periodo_id) if periodo_id else None,
        'periodo_seleccionado': periodos.filter(pk=periodo_id).first() if periodo_id else None,
        'periodos_origen': periodos.exclude(pk=periodo_id) if periodo_id else periodos.none(),
        'carreras': carreras,
        'carrera_id': int(carrera_id) if carrera_id else None,
        'nivel': int(nivel) if nivel else None,
        'estado': estado,
        'search': search,
        'level_options': level_options,
        'rows': page_rows,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': page_obj.has_other_pages(),
        'filter_querystring': _filter_querystring(request),
        'total_demands': len(rows),
        'total_parallel_slots': total_parallel_slots,
        'assigned_parallel_slots': assigned_parallel_slots,
        'pending_parallel_slots': max(0, total_parallel_slots - assigned_parallel_slots),
        'demandas_sin_campo': demandas_sin_campo,
        'docentes_sobrecargados': docentes_sobrecargados,
    }
    return render(request, 'planificacion/planificacion_operativa.html', context)


@login_required
@module_permission_required('planificacion', 'change')
def asignar_docente_operativa(request):
    if request.method != 'POST':
        return redirect('planificacion:planificacion_operativa')

    next_url = request.POST.get('next') or 'planificacion:planificacion_operativa'
    docente_id = request.POST.get('docente')
    asignatura_id = request.POST.get('asignatura')
    carrera_id = request.POST.get('carrera')
    _ensure_career_access(request, carrera_id)
    periodo_id = request.POST.get('periodo')
    campo_id = request.POST.get('campo')
    paralelo = _excel_clean_text(request.POST.get('paralelo'))[:3]
    nivel = request.POST.get('nivel')
    horas = _excel_positive_int(request.POST.get('horas'))
    complementarias = 0

    required_values = [docente_id, asignatura_id, carrera_id, periodo_id, paralelo, nivel]
    if any(not value for value in required_values):
        messages.error(request, 'Faltan datos para crear la asignación rápida.')
        return redirect(next_url)

    try:
        docente = DocenteFcacc.objects.select_related('id_dedicacion').get(id_docente=docente_id)
        asignatura = CurriculoAsignatura.objects.get(id_asignatura=asignatura_id)
        carrera = CatalogoCarrera.objects.get(id_carrera=carrera_id)
        periodo = CatalogoPeriodoAcademico.objects.get(id_periodo=periodo_id)
        nivel_int = int(nivel)
        if campo_id:
            campo = CatalogoCampoConocimiento.objects.get(id_campo=campo_id)
        else:
            campo_rel = CurriculoAsignaturaCampo.objects.filter(id_asignatura=asignatura).select_related('id_campo').first()
            campo = campo_rel.id_campo if campo_rel else None
        if campo is None:
            raise CatalogoCampoConocimiento.DoesNotExist
    except (
        DocenteFcacc.DoesNotExist, CurriculoAsignatura.DoesNotExist,
        CatalogoCarrera.DoesNotExist, CatalogoPeriodoAcademico.DoesNotExist,
        CatalogoCampoConocimiento.DoesNotExist, ValueError,
    ):
        messages.error(request, 'No se encontro alguno de los datos necesarios para asignar el paralelo.')
        return redirect(next_url)

    if not PlanificacionDemandaAcademica.objects.filter(
        id_asignatura_id=asignatura_id,
        id_carrera_id=carrera_id,
        id_periodo_id=periodo_id,
    ).exists():
        messages.error(request, 'Esta asignatura no está registrada en la demanda académica. No se puede asignar.')
        return redirect(next_url)

    if nivel_int >= 4 and not docente_tiene_afinidad(docente, asignatura):
        messages.error(
            request,
            'Desde cuarto nivel el docente debe tener afinidad registrada con la asignatura.',
        )
        return redirect(next_url)

    existing = None
    existing_qs = PlanificacionAsignacionDocente.objects.filter(
        id_asignatura_id=asignatura_id,
        id_carrera_id=carrera_id,
        id_periodo_id=periodo_id,
    )
    for candidate in existing_qs:
        if _excel_clean_text(candidate.paralelo_asignado) == paralelo:
            existing = candidate
            break

    horas = horas or asignatura.horas_semanales_asignatura or 0
    rule_errors = validate_assignment_business_rules(
        docente=docente, asignatura=asignatura, carrera=carrera,
        periodo=periodo, campo=campo, nivel=nivel_int, paralelo=paralelo,
        horas_clase=horas, instance=existing,
    )
    if rule_errors:
        messages.error(request, ' '.join(dict.fromkeys(rule_errors.values())))
        return redirect(next_url)

    snapshot = _build_asignacion_limit_snapshot(
        docente=docente,
        periodo=periodo_id,
        horas_clase_nuevas=horas,
        horas_comp_nuevas=complementarias,
        exclude_asignacion_id=existing.pk if existing else None,
    )

    if snapshot.get('error'):
        messages.error(request, snapshot['error'])
        return redirect(next_url)
    if snapshot['clase_total'] > snapshot['max_clase']:
        messages.error(request, f'{docente.nombres_completos} excede el límite de horas clase: {snapshot["clase_total"]}/{snapshot["max_clase"]}.')
        return redirect(next_url)
    if snapshot['comp_total'] > snapshot['max_comp']:
        messages.error(request, f'{docente.nombres_completos} excede el límite de horas complementarias: {snapshot["comp_total"]}/{snapshot["max_comp"]}.')
        return redirect(next_url)
    if snapshot['total_general'] > snapshot['max_total']:
        messages.error(request, f'{docente.nombres_completos} excede el límite total: {snapshot["total_general"]}/{snapshot["max_total"]}.')
        return redirect(next_url)

    defaults = {
        'id_docente': docente,
        'id_campo': campo,
        'nivel_semestre_asignado': nivel_int,
        'horas_clase': horas,
        'horas_complementarias': complementarias,
    }
    if existing:
        for field, value in defaults.items():
            setattr(existing, field, value)
        existing.paralelo_asignado = paralelo
        existing.save()
        asignacion = existing
        created = False
    else:
        asignacion = PlanificacionAsignacionDocente.objects.create(
            id_asignatura_id=asignatura_id,
            id_carrera_id=carrera_id,
            id_periodo_id=periodo_id,
            paralelo_asignado=paralelo,
            **defaults,
        )
        created = True

    action = 'Asignado' if created else 'Actualizado'
    messages.success(request, f'{action}: paralelo {paralelo} de {asignatura.nombre_asignatura} para {docente.nombres_completos}.')
    return redirect(next_url)


@login_required
@module_permission_required('planificacion', 'view')
def planificacion_consolidada_docentes(request):
    from catalogos.models import CatalogoCarrera, CatalogoPeriodoAcademico

    periodo_id = request.GET.get('periodo')
    carrera_id = request.GET.get('carrera')
    permitted_careers = _ensure_career_access(request, carrera_id)
    estado = request.GET.get('estado')
    search = (request.GET.get('q') or '').strip()

    periodos = CatalogoPeriodoAcademico.objects.order_by('-fecha_inicio_periodo', '-id_periodo')
    periodo_activo = periodos.filter(periodo_activo=True).first()
    if not periodo_id and periodo_activo:
        periodo_id = str(periodo_activo.id_periodo)

    carreras = _scope_careers(CatalogoCarrera.objects.filter(carrera_activa=True), request).order_by('nombre_carrera')
    docentes_qs = DocenteFcacc.objects.filter(docente_activo=True).select_related(
        'id_dedicacion'
    ).order_by('nombres_completos')

    assignments_qs = PlanificacionAsignacionDocente.objects.select_related(
        'id_docente',
        'id_asignatura',
        'id_carrera',
        'id_periodo',
        'id_campo',
    )
    if permitted_careers is not None:
        assignments_qs = assignments_qs.filter(id_carrera_id__in=permitted_careers)
    if periodo_id:
        assignments_qs = assignments_qs.filter(id_periodo_id=periodo_id)
    if carrera_id:
        assignments_qs = assignments_qs.filter(id_carrera_id=carrera_id)

    f4_qs = PlanificacionMatrizF4.objects.select_related(
        'id_docente',
        'id_carrera',
        'id_periodo',
        'id_grado_afinidad',
    )
    if permitted_careers is not None:
        f4_qs = f4_qs.filter(id_carrera_id__in=permitted_careers)
    if periodo_id:
        f4_qs = f4_qs.filter(id_periodo_id=periodo_id)
    if carrera_id:
        f4_qs = f4_qs.filter(id_carrera_id=carrera_id)

    if carrera_id:
        scoped_teacher_ids = set(assignments_qs.values_list('id_docente_id', flat=True)) | set(
            f4_qs.values_list('id_docente_id', flat=True)
        )
        docentes_qs = docentes_qs.filter(id_docente__in=scoped_teacher_ids)
    docentes = list(docentes_qs)

    # El filtro de carrera define los docentes y el detalle visible. Para
    # comparar con el límite contractual se conserva su carga total del período.
    workload_map = _build_docente_workload_map(periodo_id=periodo_id)
    limites = {l.id_modalidad_id: l for l in LimiteHorario.objects.filter(activo=True).select_related('id_modalidad')}

    assignments_by_doc = {}
    subject_fields, teacher_fields = _knowledge_field_maps()
    for assignment in assignments_qs:
        assignment.hour_category = _assignment_hour_category(
            assignment.id_docente_id,
            assignment.id_asignatura_id,
            assignment.id_campo.nombre_campo_conocimiento,
            subject_fields,
            teacher_fields,
        )
        assignments_by_doc.setdefault(assignment.id_docente_id, []).append(assignment)

    activities_qs = PlanificacionActividadDocente.objects.select_related(
        'id_docente', 'id_periodo', 'id_actividad'
    )
    if periodo_id:
        activities_qs = activities_qs.filter(id_periodo_id=periodo_id)
    if carrera_id:
        activities_qs = activities_qs.filter(
            id_docente_id__in=[docente.id_docente for docente in docentes]
        )
    activities_by_doc = {}
    for activity in activities_qs:
        activities_by_doc.setdefault(activity.id_docente_id, []).append(activity)

    f4_by_doc = {}
    for activity in f4_qs:
        f4_by_doc.setdefault(activity.id_docente_id, []).append(activity)

    rows = []
    for docente in docentes:
        if search:
            hay_match = (
                search.lower() in docente.nombres_completos.lower() or
                search.lower() in docente.cedula_docente.lower()
            )
            if not hay_match:
                continue

        docente_assignments = assignments_by_doc.get(docente.id_docente, [])
        docente_activities = activities_by_doc.get(docente.id_docente, [])
        docente_f4 = f4_by_doc.get(docente.id_docente, [])
        workload = workload_map.get(docente.id_docente, _empty_workload())
        limite = limites.get(docente.id_modalidad_id)
        max_total = ((limite.horas_maximas or 0) + (limite.horas_complementarias_maximas or 0)) if limite else 0
        available = max(0, max_total - workload['total_horas'])
        pct = round((workload['total_horas'] / max_total) * 100, 1) if max_total > 0 else 0

        if pct > 100:
            status = 'sobrecargado'
            badge = 'danger'
        elif pct >= 80:
            status = 'alerta'
            badge = 'warning'
        elif workload['total_horas'] > 0:
            status = 'balanceado'
            badge = 'success'
        else:
            status = 'sin_carga'
            badge = 'secondary'

        if estado and status != estado:
            continue

        rows.append({
            'docente': docente,
            'assignments': docente_assignments,
            'activities': docente_activities,
            'f4_activities': docente_f4,
            'workload': workload,
            'max_total': max_total,
            'available': available,
            'percentage': pct,
            'status': status,
            'badge': badge,
            'assignment_count': len(docente_assignments),
            'activity_count': len(docente_activities),
            'f4_count': len(docente_f4),
        })

    rows.sort(key=lambda item: (-item['workload']['total_horas'], item['docente'].nombres_completos))
    paginator, page_obj, page_rows = _paginate_items(request, rows, 15)

    context = {
        'active_section': 'planificacion_consolidada_docentes',
        'limit_config': _build_limit_config_state(),
        'periodos': periodos,
        'periodo_activo': periodo_activo,
        'periodo_id': int(periodo_id) if periodo_id else None,
        'carreras': carreras,
        'carrera_id': int(carrera_id) if carrera_id else None,
        'estado': estado or '',
        'search': search,
        'rows': page_rows,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': page_obj.has_other_pages(),
        'filter_querystring': _filter_querystring(request),
        'total_docentes': len(rows),
        'sobrecargados': sum(1 for row in rows if row['status'] == 'sobrecargado'),
        'en_alerta': sum(1 for row in rows if row['status'] == 'alerta'),
        'con_carga': sum(1 for row in rows if row['workload']['total_horas'] > 0),
        'sin_carga': sum(1 for row in rows if row['workload']['total_horas'] == 0),
    }
    return render(request, 'planificacion/planificacion_consolidada_docentes.html', context)


# ——— AJAX: endpoints para auto-fill ————————————————————————

@login_required
@module_permission_required('planificacion', 'view')
def api_asignatura_info(request):
    asignatura_id = request.GET.get('id')
    if not asignatura_id:
        return JsonResponse({'error': 'id requerido'}, status=400)
    try:
        subj = CurriculoAsignatura.objects.select_related('id_carrera').get(id_asignatura=asignatura_id)
        _ensure_career_access(request, subj.id_carrera_id)
        campo = CurriculoAsignaturaCampo.objects.filter(id_asignatura=subj).select_related('id_campo').first()
        existing = _get_existing_assignment(subj.id_asignatura)
        niveles = sorted(CurriculoAsignatura.objects.filter(
            id_carrera=subj.id_carrera, es_actividad=False
        ).values_list('nivel_semestre', flat=True).distinct())
        data = {
            'id_asignatura': subj.id_asignatura,
            'id_carrera': subj.id_carrera_id,
            'carrera_nombre': str(subj.id_carrera),
            'nivel_semestre_asignado': subj.nivel_semestre,
            'horas_clase': subj.horas_semanales_asignatura,
            'id_campo': campo.id_campo_id if campo else None,
            'campo_nombre': str(campo.id_campo) if campo else '',
            'existing_assignment': existing,
            'es_actividad': subj.es_actividad,
            'niveles_disponibles': niveles,
        }
        return JsonResponse(data)
    except CurriculoAsignatura.DoesNotExist:
        return JsonResponse({'error': 'no encontrada'}, status=404)


@login_required
@module_permission_required('planificacion', 'view')
def api_recommendations(request):
    """Return top teacher recommendations for a subject as JSON."""
    asignatura_id = request.GET.get('asignatura')
    periodo_id = request.GET.get('periodo')
    if not asignatura_id:
        return JsonResponse({'error': 'asignatura requerida'}, status=400)
    try:
        subj = CurriculoAsignatura.objects.get(id_asignatura=asignatura_id)
        _ensure_career_access(request, subj.id_carrera_id)
    except CurriculoAsignatura.DoesNotExist:
        return JsonResponse({'error': 'no encontrada'}, status=404)

    all_recs = _compute_teacher_scores(subj, periodo_id=periodo_id)
    recs = all_recs[:10]
    if subj.es_actividad:
        all_teachers = DocenteFcacc.objects.filter(docente_activo=True).order_by('nombres_completos')
        data = [{
            'id': d.id_docente,
            'nombre': d.nombres_completos,
            'dedicacion': d.id_dedicacion.codigo_dedicacion if hasattr(d, 'id_dedicacion') and d.id_dedicacion else '',
            'score': 100,
            'reasons': ['Actividad: cualquier docente disponible'],
            'available': 999,
            'used': 0,
            'max': 999,
            'class_hours': 0,
            'complementary_hours': 0,
            'investigation_hours': 0,
            'activity_hours': 0,
            'status': 'excelente',
        } for d in all_teachers]
        docentes_elegibles = [{'id': d.id_docente, 'nombre': d.nombres_completos} for d in all_teachers]
        has_subject_fields = True
        message = ''
    else:
        data = [{
            'id': r['id'],
            'nombre': r['docente'].nombres_completos,
            'dedicacion': r['docente'].id_dedicacion.codigo_dedicacion,
            'score': r['score'],
            'reasons': r['reasons'],
            'available': r['available'],
            'used': r['used'],
            'max': r['max'],
            'class_hours': r['class_hours'],
            'complementary_hours': r['complementary_hours'],
            'investigation_hours': r['investigation_hours'],
            'activity_hours': r['activity_hours'],
            'status': r['status'],
        } for r in recs]
        docentes_elegibles = [
            {'id': item['id'], 'nombre': item['docente'].nombres_completos}
            for item in all_recs
        ]
        has_subject_fields = CurriculoAsignaturaCampo.objects.filter(id_asignatura=subj).exists()
        message = ''
        if subj.nivel_semestre >= 4 and not has_subject_fields:
            message = 'La asignatura no tiene campos de conocimiento configurados; primero complete su afinidad curricular.'
        elif subj.nivel_semestre >= 4 and not data:
            message = 'No existen docentes con afinidad registrada para esta asignatura.'
    return JsonResponse({
        'recomendados': data,
        'docentes_elegibles': docentes_elegibles,
        'affinity_required': subj.nivel_semestre >= 4 and not subj.es_actividad,
        'has_subject_fields': has_subject_fields,
        'message': message,
    })


@login_required
@module_permission_required('planificacion', 'view')
def api_validar_horas_disponibles(request):
    docente_id = request.GET.get('id_docente')
    periodo_id = request.GET.get('id_periodo')

    if not docente_id or not periodo_id:
        return JsonResponse({'error': 'Parámetros insuficientes'}, status=400)

    try:
        docente = DocenteFcacc.objects.select_related('id_dedicacion').get(id_docente=docente_id)
    except DocenteFcacc.DoesNotExist:
        return JsonResponse({'error': 'Docente no encontrado'}, status=404)

    snapshot = _build_asignacion_limit_snapshot(
        docente=docente,
        periodo=periodo_id,
        horas_clase_nuevas=0,
        horas_comp_nuevas=0,
    )
    if snapshot.get('error'):
        return JsonResponse({'error': snapshot['error']}, status=400)

    return JsonResponse({
        'total_asignado': snapshot['total_asignado'],
        'total_general': snapshot['total_general'],
        'horas_disponibles': snapshot['horas_totales_disponibles'],
        'horas_clase_disponibles': snapshot['horas_clase_disponibles'],
        'horas_complementarias_disponibles': snapshot['horas_complementarias_disponibles'],
        'horas_investigacion': snapshot['horas_investigacion'],
        'horas_actividad': snapshot['horas_actividad'],
        'limite_total': snapshot['max_total'],
        'limite_horas_clase': snapshot['max_clase'],
        'limite_horas_complementarias': snapshot['max_comp'],
    })


@login_required
@module_permission_required('planificacion', 'view')
def api_check_affinity(request):
    asignatura_id = request.GET.get('asignatura')
    docente_id = request.GET.get('docente')
    if not asignatura_id or not docente_id:
        return JsonResponse({'error': 'asignatura y docente requeridos'}, status=400)
    try:
        asignatura = CurriculoAsignatura.objects.get(pk=asignatura_id)
        _ensure_career_access(request, asignatura.id_carrera_id)
        docente = DocenteFcacc.objects.get(pk=docente_id)
    except CurriculoAsignatura.DoesNotExist:
        return JsonResponse({'error': 'Asignatura no encontrada.'}, status=404)
    except DocenteFcacc.DoesNotExist:
        return JsonResponse({'error': 'Docente no encontrado.'}, status=404)
    level = asignatura.nivel_semestre
    has_affinity = docente_tiene_afinidad(docente, asignatura)
    return JsonResponse({
        'has_affinity': has_affinity,
        'affinity_required': level >= 4,
        'allowed': level <= 3 or has_affinity,
        'level': level,
    })


@login_required
@module_permission_required('planificacion', 'view')
def api_paralelos_disponibles(request):
    asignatura_id = request.GET.get('asignatura')
    carrera_id = request.GET.get('carrera')
    _ensure_career_access(request, carrera_id)
    periodo_id = request.GET.get('periodo')
    if not all([asignatura_id, carrera_id, periodo_id]):
        return JsonResponse({'paralelos': [], 'message': 'Seleccione asignatura, carrera y período.'})
    demanda = PlanificacionDemandaAcademica.objects.filter(
        id_asignatura_id=asignatura_id,
        id_carrera_id=carrera_id,
        id_periodo_id=periodo_id,
    ).first()
    if not demanda or demanda.numero_paralelos <= 0:
        return JsonResponse({'paralelos': [], 'message': 'No hay demanda registrada para esta combinación.'})
    labels = _build_parallel_labels(demanda.numero_paralelos)
    return JsonResponse({'paralelos': labels, 'message': ''})


@login_required
@module_permission_required('planificacion', 'view')
def api_teacher_load(request):
    docente_id = request.GET.get('docente')
    periodo_id = request.GET.get('periodo')
    if not docente_id or not periodo_id:
        return JsonResponse({'error': 'docente y periodo requeridos'}, status=400)
    try:
        docente = DocenteFcacc.objects.select_related('id_dedicacion', 'id_modalidad').get(id_docente=docente_id)
    except DocenteFcacc.DoesNotExist:
        return JsonResponse({'error': 'docente no encontrado'}, status=404)
    summary = _build_form_load_summary(docente=docente, periodo=periodo_id)
    if not summary:
        return JsonResponse({'error': 'no se pudo calcular la carga'}, status=400)
    return JsonResponse({
        'docente': summary['docente'].nombres_completos,
        'dedicacion': summary['docente'].id_dedicacion.codigo_dedicacion if summary['docente'].id_dedicacion else '',
        'percentage': summary['percentage'],
        'badge': summary['badge'],
        'clase': summary['workload']['horas_clase'],
        'comp': summary['workload']['horas_complementarias'],
        'investigacion': summary['workload']['horas_investigacion'],
        'actividad': summary['workload']['horas_actividad'],
        'available': summary['available'],
        'max_clase': summary['max_clase'],
        'max_comp': summary['max_comp'],
        'max_total': summary['max_total'],
    })


@login_required
@module_permission_required('planificacion', 'change')
@role_required(*ROLES_ESCRITURA)
def api_crear_asignacion(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST requerido'}, status=405)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    docente_id = data.get('id_docente')
    asignatura_id = data.get('id_asignatura')
    carrera_id = data.get('id_carrera')
    periodo_id = data.get('id_periodo')
    campo_id = data.get('id_campo')
    nivel = data.get('nivel_semestre_asignado')
    paralelo = _excel_clean_text(data.get('paralelo_asignado', 'A')).upper()[:3]
    horas_clase = _excel_positive_int(data.get('horas_clase'))
    # Las horas complementarias se asignan como actividades, no como asignaturas.
    horas_comp = 0

    from catalogos.models import CatalogoPeriodoAcademico
    if not periodo_id:
        periodo_activo = CatalogoPeriodoAcademico.objects.filter(periodo_activo=True).first()
        periodo_id = periodo_activo.id_periodo if periodo_activo else None

    if not campo_id:
        campo_rel = CurriculoAsignaturaCampo.objects.filter(id_asignatura_id=asignatura_id).first()
        if campo_rel:
            campo_id = campo_rel.id_campo_id

    if not campo_id:
        return JsonResponse({'error': 'La asignatura no tiene un campo de conocimiento asociado. Seleccione uno manualmente.'}, status=400)

    if not all([docente_id, asignatura_id, carrera_id, periodo_id]):
        return JsonResponse({'error': 'Faltan campos obligatorios'}, status=400)

    try:
        docente = DocenteFcacc.objects.select_related('id_dedicacion').get(id_docente=docente_id)
        asignatura = CurriculoAsignatura.objects.get(id_asignatura=asignatura_id)
        carrera = CatalogoCarrera.objects.get(id_carrera=carrera_id)
        periodo = CatalogoPeriodoAcademico.objects.get(id_periodo=periodo_id)
        campo = CatalogoCampoConocimiento.objects.get(id_campo=campo_id)
        _ensure_career_access(request, carrera.id_carrera)
    except DocenteFcacc.DoesNotExist:
        return JsonResponse({'error': 'Docente no encontrado'}, status=404)
    except CurriculoAsignatura.DoesNotExist:
        return JsonResponse({'error': 'Asignatura no encontrada'}, status=404)
    except CatalogoCarrera.DoesNotExist:
        return JsonResponse({'error': 'Carrera no encontrada'}, status=404)
    except CatalogoPeriodoAcademico.DoesNotExist:
        return JsonResponse({'error': 'Período no encontrado'}, status=404)
    except CatalogoCampoConocimiento.DoesNotExist:
        return JsonResponse({'error': 'Campo de conocimiento no encontrado'}, status=404)

    try:
        nivel = int(nivel or asignatura.nivel_semestre)
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Nivel inválido.'}, status=400)
    if nivel >= 4 and not docente_tiene_afinidad(docente, asignatura):
        return JsonResponse({
            'error': 'Desde cuarto nivel solo se permiten docentes con afinidad registrada para la asignatura.'
        }, status=400)

    existing = list(PlanificacionAsignacionDocente.objects.filter(
        id_asignatura_id=asignatura_id,
        id_carrera_id=carrera_id,
        id_periodo_id=periodo_id,
        paralelo_asignado__iexact=paralelo,
    ).select_related('id_docente'))
    if existing:
        return JsonResponse({
            'error': 'Este paralelo ya tiene un docente asignado en el período actual.',
            'existing': [{'id': e.id_docente_id, 'nombre': e.id_docente.nombres_completos} for e in existing],
        }, status=409)

    horas_clase = horas_clase or asignatura.horas_semanales_asignatura or 0
    rule_errors = validate_assignment_business_rules(
        docente=docente, asignatura=asignatura, carrera=carrera,
        periodo=periodo, campo=campo, nivel=nivel, paralelo=paralelo,
        horas_clase=horas_clase,
    )
    if rule_errors:
        return JsonResponse({'error': ' '.join(dict.fromkeys(rule_errors.values())), 'fields': rule_errors}, status=400)

    snapshot = _build_asignacion_limit_snapshot(
        docente=docente,
        periodo=periodo_id,
        horas_clase_nuevas=horas_clase,
        horas_comp_nuevas=horas_comp,
    )
    if snapshot.get('error'):
        return JsonResponse({'error': snapshot['error']}, status=400)
    if snapshot['clase_total'] > snapshot['max_clase']:
        return JsonResponse({'error': f'Se excede el límite de horas clase: {snapshot["clase_total"]}/{snapshot["max_clase"]}.'}, status=400)
    if snapshot['comp_total'] > snapshot['max_comp']:
        return JsonResponse({'error': f'Se excede el límite de horas complementarias: {snapshot["comp_total"]}/{snapshot["max_comp"]}.'}, status=400)
    if snapshot['total_general'] > snapshot['max_total']:
        return JsonResponse({'error': f'Se excede el límite total: {snapshot["total_general"]}/{snapshot["max_total"]}.'}, status=400)

    asignacion = PlanificacionAsignacionDocente.objects.create(
        id_docente_id=docente_id,
        id_asignatura_id=asignatura_id,
        id_carrera_id=carrera_id,
        id_periodo_id=periodo_id,
        id_campo_id=campo_id,
        nivel_semestre_asignado=nivel or 0,
        paralelo_asignado=paralelo or 'A',
        horas_clase=horas_clase,
        horas_complementarias=horas_comp,
    )
    try:
        from core.crud_base import _audit_log
        _audit_log(request, asignacion, 'INSERT')
    except Exception:
        pass

    return JsonResponse({
        'success': True,
        'id_asignacion': asignacion.id_asignacion,
        'message': 'Asignación creada correctamente.',
    })
